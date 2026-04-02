"""Unit tests for dig package anchor + offset layout resolution."""

import io

import pytest
from openpyxl import Workbook

from backend.pipeline.dig_package_layout import (
    AnchorNotFoundError,
    find_anchor_cell,
    find_anchor_cell_candidates,
    load_layout_manifest,
    resolve_field_cell,
    verify_layout_against_workbook,
)


def test_find_anchor_cell_equals():
    ws = Workbook().active
    ws.title = "Dig Package"
    ws["C5"] = "Dig Name"
    r, c = find_anchor_cell(ws, "Dig Name", "equals")
    assert (r, c) == (5, 3)


def test_find_anchor_cell_contains():
    ws = Workbook().active
    ws["B2"] = "Rev # 0"
    r, c = find_anchor_cell(ws, "Rev #", "contains")
    assert (r, c) == (2, 2)


def test_find_anchor_cell_raises():
    ws = Workbook().active
    with pytest.raises(AnchorNotFoundError):
        find_anchor_cell(ws, "NoSuchLabel", "equals")


def test_find_anchor_cell_candidates_uses_fallback():
    ws = Workbook().active
    ws["D4"] = "Nominal Pipe Size (in)"
    r, c = find_anchor_cell_candidates(
        ws,
        [("Pipe NPS", "contains"), ("Nominal Pipe Size", "contains")],
    )
    assert (r, c) == (4, 4)


def test_resolve_field_cell_pipeline_minimal_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dig Package"
    ws["A3"] = "Pipeline Name"
    manifest = load_layout_manifest()
    cell = resolve_field_cell(wb, manifest, "pipeline_name")
    assert cell.coordinate == "B3"


def test_resolve_field_cell_value_after_merged_label():
    """Merged label row (e.g. A:B): value must start at C, not B (B snaps to A and overwrites)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dig Package"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Dig Name"
    manifest = load_layout_manifest()
    cell = resolve_field_cell(wb, manifest, "dig_display")
    assert cell.coordinate == "C1"


def test_resolve_field_cell_pipe_od_does_not_match_nps_inside_dig_name():
    """Bare 'NPS' must not anchor on Dig Name values like …NPS10… (pipe row must win)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dig Package"
    ws["A7"] = "Dig Name"
    ws["C7"] = "ID6000_R1R2_MP3_NPS10_GW3180_ML"
    ws["A14"] = "Pipe Outside Diameter (mm)"
    manifest = load_layout_manifest()
    cell = resolve_field_cell(wb, manifest, "pipe_od")
    assert cell.row == 14


def _minimal_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dig Package"
    for i, lab in enumerate(
        [
            "Dig Name",
            "Rev #",
            "Pipeline Name",
            "Target Girth Weld (TGW)",
            "Pipe NPS",
            "Pipe NWT",
            "MOP",
            "SEP",
            "Lat (deg)",
            "Long (deg)",
            "Milepost",
            "Pipe Year",
            "Pipe Grade",
            "Originating ILI",
            "ILI Time",
            "Upstream AGM",
            "Downstream AGM",
            "Number of Excavations",
            "Issue Date",
        ],
        start=1,
    ):
        ws.cell(i, 1).value = lab
    ws["A20"] = "Excavation"
    ws["A25"] = "Exposure"
    ws["A30"] = "Feature ID"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_verify_layout_all_ok_on_minimal_template():
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(_minimal_workbook_bytes()))
    manifest = load_layout_manifest()
    results = verify_layout_against_workbook(wb, manifest)
    bad = [r for r in results if not r["ok"]]
    assert not bad, bad
