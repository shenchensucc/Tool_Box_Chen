import io
import time

import pytest
from openpyxl import Workbook

from backend.pipeline.ili_parse import ILIParseTimeoutError, parse_ili_file, parse_ili_file_with_timeout

def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_ili_file_uses_shared_detection_for_rosen_anomalies_sheet():
    workbook = Workbook()
    workbook.active.title = "Cover"
    ws = workbook.create_sheet("Anomalies Listing 2025")
    ws["A1"] = "Rosen export"
    ws.append(
        [
            "Feature ID",
            "Description",
            "Length (mm)",
            "Width (mm)",
            "Peak Depth",
            "Orientation (hh:mm)",
            "ILI Chainage (m)",
            "Joint No. or US GW No.",
        ]
    )
    ws.append(["F-1", "Metal Loss", 25.4, 12.7, 42, "03:30", 1000.5, 3150])

    df, column_mapping, sheet_name = parse_ili_file(_workbook_bytes(workbook), "Rosen-MFLA")

    assert sheet_name == "Anomalies Listing 2025"
    assert column_mapping["feature_id"] == "Feature ID"
    assert column_mapping["distance"] == "ILI Chainage (m)"
    assert column_mapping["joint_number"] == "Joint No. or US GW No."
    assert len(df) == 1


def test_ili_vendor_keyword_overrides_exported_for_callers():
    from backend.pipeline.ili_parse import ILI_VENDOR_KEYWORD_OVERRIDES

    assert "TDW" in ILI_VENDOR_KEYWORD_OVERRIDES
    assert "Rosen-MFLA" in ILI_VENDOR_KEYWORD_OVERRIDES


def test_parse_ili_file_with_timeout_aborts_slow_parse(monkeypatch):
    def slow_parse(*args, **kwargs):
        time.sleep(30)

    monkeypatch.setattr("backend.pipeline.ili_parse.parse_ili_file", slow_parse)
    with pytest.raises(ILIParseTimeoutError) as exc_info:
        parse_ili_file_with_timeout(b"x", "TDW", timeout_sec=0.15)
    assert "timed out" in str(exc_info.value).lower()


def test_parse_ili_file_with_timeout_zero_disables_limit(monkeypatch):
    import pandas as pd

    calls = []

    def fast_parse(fc, vf):
        calls.append((vf,))
        return pd.DataFrame(), {}, "S"

    monkeypatch.setattr("backend.pipeline.ili_parse.parse_ili_file", fast_parse)
    parse_ili_file_with_timeout(b"x", "Rosen-MFLA", timeout_sec=0)
    assert calls == [("Rosen-MFLA",)]
