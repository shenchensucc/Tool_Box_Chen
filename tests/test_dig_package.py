import io
import json
import zipfile

import pandas as pd
from openpyxl import Workbook

from backend.pipeline.dig_package import (
    _get_mdl_value,
    extract_dig_ids,
    filter_ili_by_gw_count,
    generate_dig_packages,
    get_target_gw_chainage,
    is_valid_dig_id,
    match_features_by_dimensions,
    package_output_stem,
    parse_mdl_file,
    populate_excavation_summary,
)
from backend.pipeline.dig_package_layout import load_layout_manifest
from backend.pipeline.dig_package_reader import (
    _build_joint_context,
    _find_header_row_after,
    _match_joint_source_name,
    _parse_joint_summary_matrix,
    _extract_sources_from_block_label,
    _reshape_joint_summary_dataframe,
    build_feature_map_from_dig_package,
    logger as dig_package_reader_logger,
    parse_dig_package_excel,
)
from backend.pipeline.feature_map_builder import parse_orientation_to_hours


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_mdl_file_detects_header_after_title_rows():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "2025 Features&Dig Export"
    ws["A1"] = "Master Dig List"
    ws.append([])
    ws.append(["Dig ID", "Feature ID", "Start Assessment", "End Assessment", "Target Girth Weld", "Pipeline Name", "Length (mm)", "Width (mm)"])
    ws.append(["GW-100", "F-1", 12, 18, 3150, "Main Line", 25.4, 12.7])

    df, column_mapping = parse_mdl_file(_workbook_bytes(workbook))

    assert column_mapping["dig_id"] == "Dig ID"
    assert column_mapping["target_girth_weld"] == "Target Girth Weld"
    assert extract_dig_ids(df, column_mapping) == ["GW-100"]


def test_extract_dig_ids_accepts_png_integrity_numeric_dig_ids():
    df = pd.DataFrame(
        {
            "Dig ID": [6000, 6001],
            "Dig Name": ["ID6000_R1R2_MP3_NPS10_GW3180_ML", "ID6001_R1R2_MP4_NPS10_GW4580_ML"],
            "Target Girth Weld (TGW)": [3180, 4580],
        }
    )
    column_mapping = {
        "dig_id": "Dig ID",
        "dig_name": "Dig Name",
        "target_girth_weld": "Target Girth Weld (TGW)",
    }
    got = extract_dig_ids(df, column_mapping)
    assert [int(x) for x in got] == [6000, 6001]


def test_package_output_stem_prefers_dig_name():
    row = pd.Series(
        {
            "Dig Name": "ID6000_R1R2_MP3_NPS10_GW3180_ML",
            "Dig ID": 6000,
        }
    )
    col_map = {"dig_name": "Dig Name", "dig_id": "Dig ID"}
    assert package_output_stem(row, col_map, 6000) == "ID6000_R1R2_MP3_NPS10_GW3180_ML"


def test_is_valid_dig_id_numeric_and_legacy_gw():
    assert is_valid_dig_id(6000) is True
    assert is_valid_dig_id("GW-100") is True
    assert is_valid_dig_id("") is False


def test_populate_excavation_summary_writes_via_layout_anchors():
    """Excavation / exposure blocks use anchor labels Excavation / Exposure + column B values."""
    workbook = Workbook()
    intro_ws = workbook.active
    intro_ws.title = "Intro"
    template_ws = workbook.create_sheet("Dig Package")
    template_ws["A20"] = "Excavation"
    template_ws["A25"] = "Exposure"

    mdl_row = pd.Series(
        {
            "assessment_length_col": 30,
            "start_assessment_col": 10,
            "end_assessment_col": 20,
            "exposure_length_col": 25,
            "start_exposure_col": 11,
            "end_exposure_col": 21,
        }
    )
    mdl_col_map = {
        "assessment_length": "assessment_length_col",
        "start_assessment": "start_assessment_col",
        "end_assessment": "end_assessment_col",
        "exposure_length": "exposure_length_col",
        "start_exposure": "start_exposure_col",
        "end_exposure": "end_exposure_col",
    }

    manifest = load_layout_manifest()
    populate_excavation_summary(workbook, mdl_row, mdl_col_map, excavation_num=7, layout_manifest=manifest)

    assert template_ws["B20"].value == "Excavation #7"
    assert template_ws["B21"].value == 30
    assert template_ws["B22"].value == 10
    assert template_ws["B23"].value == 20
    assert template_ws["B25"].value == "Excavation #7"
    assert template_ws["B26"].value == 25
    assert template_ws["B27"].value == 11
    assert template_ws["B28"].value == 21
    assert intro_ws["B2"].value is None


def test_get_target_gw_chainage_matches_normalized_joint_values():
    mdl_row = pd.Series({"target_gw_col": "3150"})
    mdl_col_map = {"target_girth_weld": "target_gw_col"}
    ili_df = pd.DataFrame(
        {
            "Joint Number": [3150.0, 3160.0],
            "ILI Chainage (m)": [1000.5, 1012.0],
        }
    )
    ili_col_map = {"joint_number": "Joint Number", "distance": "ILI Chainage (m)"}

    chainage = get_target_gw_chainage(mdl_row, ili_df, mdl_col_map, ili_col_map)

    assert chainage == 1000.5


def test_match_features_by_dimensions_handles_mm_to_inch_conversion():
    ili_df = pd.DataFrame(
        {
            "Length (in)": [1.0, 2.0],
            "Width (in)": [0.5, 0.75],
        }
    )
    ili_col_map = {"length": "Length (in)", "width": "Width (in)"}

    matched = match_features_by_dimensions(25.4, 12.7, ili_df, ili_col_map)

    assert len(matched) == 1
    assert matched.index.tolist() == [0]


def test_parse_dig_package_excel_propagates_merged_joint_summary_labels():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Dig Package"
    ws["A1"] = "Joint Summary"
    ws["A2"] = "Girth Weld No."
    ws["B2"] = 4570
    ws["C2"] = 4580
    ws["D2"] = 4590

    ws.merge_cells("A3:A4")
    ws["A3"] = "Long Seam Orientation (hh:mm)\n(2022 Rosen)\n(2025 TDW)"
    ws["B3"] = "11:28"
    ws["C3"] = "12:40"
    ws["D3"] = "11:22"
    ws["B4"] = "11:38"
    ws["C4"] = "12:56"
    ws["D4"] = "11:28"

    ws.merge_cells("A5:A6")
    ws["A5"] = "Joint Length (m)\n(2022 Rosen)\n(2025 TDW)"
    ws["B5"] = 19.560
    ws["C5"] = 18.860
    ws["D5"] = 19.340
    ws["B6"] = 19.650
    ws["C6"] = 18.910
    ws["D6"] = 19.394

    _, joint_df, metadata = parse_dig_package_excel(_workbook_bytes(workbook))

    assert metadata["joint_section_found"] is True
    assert joint_df is not None
    assert joint_df.iloc[0, 0] == "Long Seam Orientation (hh:mm)"
    assert joint_df.iloc[0, 1] == "2022 Rosen"
    assert joint_df.iloc[1, 1] == "2025 TDW"
    assert joint_df.iloc[2, 0] == "Joint Length (m)"
    assert list(joint_df.columns[:5]) == ["Metric", "Source", "4570", "4580", "4590"]


def test_parse_joint_summary_matrix_reads_sources_from_merged_label_block():
    joint_df = pd.DataFrame(
        [
            ["Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)", "11:28", "12:40", "11:22"],
            ["Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)", "11:38", "12:56", "11:28"],
            ["Joint Length (m) (2022 Rosen) (2025 TDW)", 19.560, 18.860, 19.340],
            ["Joint Length (m) (2022 Rosen) (2025 TDW)", 19.650, 18.910, 19.394],
        ],
        columns=["Girth Weld No.", "4570", "4580", "4590"],
    )

    parsed = _parse_joint_summary_matrix(joint_df, parse_orientation_to_hours, logger=dig_package_reader_logger)

    assert parsed is not None
    by_source = parsed["gwd_by_source"]
    assert set(by_source.keys()) == {"2022 Rosen", "2025 TDW"}
    assert by_source["2022 Rosen"][4580] == 12 + 40 / 60
    assert by_source["2025 TDW"][4580] == 12 + 56 / 60
    assert all(value <= 12.95 for seam_map in by_source.values() for value in seam_map.values())


def test_parse_joint_summary_matrix_interleaved_duplicate_column_headers():
    """PNG-style Joint Summary: two ILI columns per GWD (30900 / 30900__2), one longseam row."""
    joint_df = pd.DataFrame(
        [
            [
                "Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)",
                "10:40",
                "10:30",
                "10:45",
                "10:35",
            ],
            [
                "Joint Length (m) (2022 Rosen) (2025 TDW)",
                13.5,
                13.6,
                13.84,
                13.92,
            ],
        ],
        columns=["Metric", "30900", "30900__2", "30930", "30930__2"],
    )

    parsed = _parse_joint_summary_matrix(joint_df, parse_orientation_to_hours, logger=dig_package_reader_logger)

    assert parsed is not None
    by_source = parsed["gwd_by_source"]
    assert set(by_source.keys()) == {"2022 Rosen", "2025 TDW"}
    assert by_source["2022 Rosen"][30900] == 10 + 40 / 60
    assert by_source["2025 TDW"][30900] == 10 + 30 / 60
    assert by_source["2022 Rosen"][30930] == 10 + 45 / 60
    assert by_source["2025 TDW"][30930] == 10 + 35 / 60
    jl = parsed["joint_lengths"]
    assert jl[30900] == round((13.5 + 13.6) / 2, 2)
    assert jl[30930] == round((13.84 + 13.92) / 2, 2)


def test_reshape_joint_summary_dataframe_splits_metric_and_source():
    joint_df = pd.DataFrame(
        [
            ["Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)", "Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)", "11:28", "12:40"],
            ["Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)", "Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)", "11:38", "12:56"],
            ["Joint Length (m) (2022 Rosen) (2025 TDW)", "Joint Length (m) (2022 Rosen) (2025 TDW)", 19.560, 18.860],
            ["Joint Length (m) (2022 Rosen) (2025 TDW)", "Joint Length (m) (2022 Rosen) (2025 TDW)", 19.650, 18.910],
        ],
        columns=["Girth Weld No.", "Girth Weld No.__2", "4570", "4580"],
    )

    reshaped = _reshape_joint_summary_dataframe(joint_df)

    assert list(reshaped.columns[:4]) == ["Metric", "Source", "4570", "4580"]
    assert reshaped.iloc[0, 0] == "Long Seam Orientation (hh:mm)"
    assert reshaped.iloc[0, 1] == "2022 Rosen"
    assert reshaped.iloc[1, 1] == "2025 TDW"
    assert reshaped.iloc[2, 0] == "Joint Length (m)"


def test_extract_sources_from_block_label_does_not_require_known_vendor_names():
    label = "Long Seam Orientation (hh:mm)\n(Alpha Tool)\n(Beta Inline 2027)"

    sources = _extract_sources_from_block_label(label)

    assert sources == ["Alpha Tool", "Beta Inline 2027"]


def test_match_joint_source_name_matches_joint_source_to_feature_source_generically():
    available = ["2022 Rosen", "2025 TDW"]

    assert _match_joint_source_name("2022 Rosen-MFLA", available) == "2022 Rosen"
    assert _match_joint_source_name("2022 Rosen EMAT", available) == "2022 Rosen"
    assert _match_joint_source_name("2025 TDW", available) == "2025 TDW"


def test_build_joint_context_returns_upstream_target_downstream_per_source():
    girth_welds = [
        {"chainage": -19.56, "gwd_number": 4570, "label": "GWD 4570", "source": "2022 Rosen-MFLA"},
        {"chainage": 0.0, "gwd_number": 4580, "label": "GWD 4580", "source": "2022 Rosen-MFLA"},
        {"chainage": 18.86, "gwd_number": 4590, "label": "GWD 4590", "source": "2022 Rosen-MFLA"},
    ]
    seam_map_by_joint_source = {
        "2022 Rosen": {4570: 11 + 28 / 60, 4580: 12 + 40 / 60, 4590: 11 + 22 / 60},
    }

    context = _build_joint_context(girth_welds, seam_map_by_joint_source, ["2022 Rosen-MFLA"])

    assert context["2022 Rosen-MFLA"]["joint_source"] == "2022 Rosen"
    assert context["2022 Rosen-MFLA"]["upstream"]["gwd_number"] == 4570
    assert context["2022 Rosen-MFLA"]["target"]["gwd_number"] == 4580
    assert context["2022 Rosen-MFLA"]["downstream"]["gwd_number"] == 4590
    assert context["2022 Rosen-MFLA"]["target"]["longseam_label"] == "12:40"


def test_build_feature_map_from_dig_package_returns_joint_summary_rows_and_context():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Dig Package"
    ws["A1"] = "Joint Summary"
    ws["A2"] = "Girth Weld No."
    ws["B2"] = "Girth Weld No."
    ws["C2"] = 4570
    ws["D2"] = 4580
    ws["E2"] = 4590
    ws.merge_cells("A3:B3")
    ws["A3"] = "Long Seam Orientation (hh:mm)\n(2022 Rosen)\n(2025 TDW)"
    ws["C3"] = "11:28"
    ws["D3"] = "12:40"
    ws["E3"] = "11:22"
    ws.merge_cells("A4:B4")
    ws["A4"] = "Long Seam Orientation (hh:mm)\n(2022 Rosen)\n(2025 TDW)"
    ws["C4"] = "11:38"
    ws["D4"] = "12:56"
    ws["E4"] = "11:28"

    ws["A7"] = "Feature Summary"
    ws["A8"] = "Distance from TGW (m)"
    ws["B8"] = "Feature Type"
    ws["C8"] = "Joint Number"
    ws["D8"] = "ILI Source"
    ws["A9"] = -19.56
    ws["B9"] = "Girth Weld"
    ws["C9"] = 4570
    ws["D9"] = "2022 Rosen-MFLA"
    ws["A10"] = 0.0
    ws["B10"] = "Girth Weld"
    ws["C10"] = 4580
    ws["D10"] = "2022 Rosen-MFLA"
    ws["A11"] = 18.86
    ws["B11"] = "Girth Weld"
    ws["C11"] = 4590
    ws["D11"] = "2022 Rosen-MFLA"

    _, scatter_data, _, _, joint_summary_parsed, feature_summary_raw = build_feature_map_from_dig_package(_workbook_bytes(workbook))

    assert len(joint_summary_parsed) == 6
    ctx = scatter_data["joint_context_by_source"]["2022 Rosen-MFLA"]
    assert ctx["target"]["gwd_number"] == 4580
    assert ctx["upstream"]["longseam_label"] == "11:28"
    assert feature_summary_raw["joint_context_by_source"]["2022 Rosen-MFLA"]["downstream"]["gwd_number"] == 4590


def test_build_feature_map_tgw_layout_uses_joint_lengths_for_four_red_lines():
    """When Joint Summary has ≥4 GWDs + joint lengths, girth welds sit at 0 and ±joint lengths."""
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Dig Package"
    ws["A1"] = "Joint Summary"
    ws["A2"] = "Girth Weld No."
    ws["B2"] = "Girth Weld No."
    ws["C2"] = 4560
    ws["D2"] = 4570
    ws["E2"] = 4580
    ws["F2"] = 4590
    ws.merge_cells("A3:B3")
    ws["A3"] = "Long Seam Orientation (hh:mm)\n(2022 Rosen)\n(2025 TDW)"
    ws["C3"] = "10:00"
    ws["D3"] = "10:10"
    ws["E3"] = "10:20"
    ws["F3"] = "10:30"
    ws.merge_cells("A4:B4")
    ws["A4"] = "Long Seam Orientation (hh:mm)\n(2022 Rosen)\n(2025 TDW)"
    ws["C4"] = "10:05"
    ws["D4"] = "10:15"
    ws["E4"] = "10:25"
    ws["F4"] = "10:35"

    ws.merge_cells("A5:B5")
    ws["A5"] = "Joint Length (m)\n(2022 Rosen)\n(2025 TDW)"
    ws["C5"] = 12.0
    ws["D5"] = 17.5
    ws["E5"] = 19.0
    ws["F5"] = 18.5

    ws["A9"] = "Feature Summary"
    ws["A10"] = "Distance from TGW (m)"
    ws["B10"] = "Feature Type"
    ws["C10"] = "Joint Number"
    ws["D10"] = "ILI Source"
    ws["A11"] = -5.0
    ws["B11"] = "Metal Loss"
    ws["C11"] = 1
    ws["D11"] = "2022 Rosen-MFLA"

    _, scatter_data, _, _, _, _ = build_feature_map_from_dig_package(_workbook_bytes(workbook))

    assert scatter_data.get("joint_summary_tgw_layout") is True
    gws = sorted(scatter_data["girth_welds"], key=lambda g: g["chainage"])
    assert len(gws) == 4
    # Joint length under upstream GWD of each span: L(4560–4570)=12, L(4570–4580)=17.5, L(4580–4590)=19
    assert abs(gws[0]["chainage"] - (-12.0)) < 0.01
    assert abs(gws[1]["chainage"] - 0.0) < 0.01
    assert abs(gws[2]["chainage"] - 17.5) < 0.01
    assert abs(gws[3]["chainage"] - 36.5) < 0.01
    seams = scatter_data.get("seam_welds", [])
    assert len(seams) >= 3
    tol = 0.02

    def _has_span(a: float, b: float) -> bool:
        return any(
            abs(s["chainage_start"] - a) < tol and abs(s["chainage_end"] - b) < tol for s in seams
        )

    assert _has_span(-12.0, 0.0)
    assert _has_span(0.0, 17.5)


def test_build_feature_map_from_dig_package_parses_nde_limits_tgw_strip():
    """NDE Limits block (between Joint and Feature sections) → scatter_data.nde_region for the map."""
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Dig Package"
    ws["A1"] = "NDE Limits"
    ws["A2"] = "Target Girth Weld (U/S)"
    ws["B2"] = 3180
    ws["A3"] = "NDE Assessment Length (m)"
    ws["B3"] = 9.304
    ws["A4"] = "NDE Assessment Start from TGW (m)"
    ws["B4"] = -0.5
    ws["A5"] = "NDE Assessment End from TGW (m)"
    ws["B5"] = 8.804

    ws["A8"] = "Feature Summary"
    ws["A9"] = "Distance from TGW (m)"
    ws["B9"] = "Feature Type"
    ws["C9"] = "Depth (%wt)"
    ws["A10"] = 1.0
    ws["B10"] = "Metal Loss"
    ws["C10"] = 42.0

    _, scatter_data, _, _, _, feature_summary_raw = build_feature_map_from_dig_package(_workbook_bytes(workbook))

    nr = scatter_data.get("nde_region")
    assert nr is not None
    assert abs(nr["x0"] - (-0.5)) < 1e-6
    assert abs(nr["x1"] - 8.804) < 1e-6
    assert abs(nr["length_m"] - 9.304) < 1e-6
    assert nr.get("target_gwd_number") == 3180
    assert feature_summary_raw.get("nde_limits", {}).get("target_gwd_number") == 3180


def test_find_header_row_after_skips_repeated_merged_section_title():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Dig Package"
    ws.merge_cells("A1:J1")
    ws["A1"] = "Feature Summary"
    ws["A2"] = "Distance from TGW (m)"
    ws["B2"] = "Feature ID"
    ws["C2"] = "Depth (%)"
    ws["A3"] = 0.0
    ws["B3"] = "F-1"
    ws["C3"] = 25

    header_row = _find_header_row_after(ws, 1)

    assert header_row == 2


# ============================================================================
# New tests — Lane D
# ============================================================================


# -----------------------------------------------------------------------
# _get_mdl_value helper
# -----------------------------------------------------------------------

def test_get_mdl_value_returns_value_when_present():
    row = pd.Series({"Pipeline Name": "Main Line"})
    col_map = {"pipeline_name": "Pipeline Name"}
    assert _get_mdl_value(row, col_map, "pipeline_name") == "Main Line"


def test_get_mdl_value_returns_dash_when_col_not_in_map():
    row = pd.Series({"Pipeline Name": "Main Line"})
    assert _get_mdl_value(row, {}, "pipeline_name") == "-"


def test_get_mdl_value_returns_dash_for_nan_cell():
    row = pd.Series({"Pipeline Name": float("nan")})
    col_map = {"pipeline_name": "Pipeline Name"}
    assert _get_mdl_value(row, col_map, "pipeline_name") == "-"


# -----------------------------------------------------------------------
# filter_ili_by_gw_count
# -----------------------------------------------------------------------

def _make_ili_with_girth_welds() -> tuple:
    """Return (DataFrame, col_map) with a mix of girth weld and anomaly rows."""
    rows = [
        # GWDs sorted by chainage
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 900.0, "ID#": "GW-1"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 920.0, "ID#": "GW-2"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 940.0, "ID#": "GW-3"},
        # TGW
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 1000.0, "ID#": "GW-TGW"},
        {"Feature Type": "Metal Loss",  "ILI Chainage (m)": 1005.0, "ID#": "F-1"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 1060.0, "ID#": "GW-5"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 1080.0, "ID#": "GW-6"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 1100.0, "ID#": "GW-7"},
        {"Feature Type": "Metal Loss",  "ILI Chainage (m)": 1200.0, "ID#": "F-2"},
    ]
    df = pd.DataFrame(rows)
    col_map = {
        "feature_type": "Feature Type",
        "distance": "ILI Chainage (m)",
        "feature_id": "ID#",
    }
    return df, col_map


def test_filter_ili_by_gw_count_happy_path_3_each_side():
    df, col_map = _make_ili_with_girth_welds()
    result = filter_ili_by_gw_count(df, target_gw_chainage=1000.0, ili_col_map=col_map)
    chainages = result["ILI Chainage (m)"].tolist()
    # Should include the 3 upstream GWDs (900, 920, 940) and 3 downstream (1060, 1080, 1100)
    # plus the TGW itself (1000) and the anomaly between TGW and first downstream (1005)
    assert 900.0 in chainages
    assert 1000.0 in chainages
    assert 1005.0 in chainages
    assert 1100.0 in chainages
    # The far-downstream anomaly at 1200m should be excluded
    assert 1200.0 not in chainages


def test_filter_ili_by_gw_count_clamps_when_fewer_upstream_gws():
    """Only 1 GWD upstream of TGW — should clamp without error."""
    rows = [
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 990.0, "ID#": "GW-only-upstream"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 1000.0, "ID#": "GW-TGW"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 1020.0, "ID#": "GW-DS1"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 1040.0, "ID#": "GW-DS2"},
        {"Feature Type": "Girth Weld", "ILI Chainage (m)": 1060.0, "ID#": "GW-DS3"},
    ]
    df = pd.DataFrame(rows)
    col_map = {"feature_type": "Feature Type", "distance": "ILI Chainage (m)", "feature_id": "ID#"}
    result = filter_ili_by_gw_count(df, target_gw_chainage=1000.0, ili_col_map=col_map)
    chainages = result["ILI Chainage (m)"].tolist()
    assert 990.0 in chainages   # the only upstream GWD is still included
    assert 1060.0 in chainages  # 3 downstream GWDs present


def test_filter_ili_by_gw_count_falls_back_when_no_gwd_rows():
    """No girth-weld rows at all → fall back to ±DEFAULT_ASSESSMENT_RANGE_M."""
    from backend.pipeline.dig_package import DEFAULT_ASSESSMENT_RANGE_M
    rows = [
        {"Feature Type": "Metal Loss", "ILI Chainage (m)": 960.0, "ID#": "F-1"},
        {"Feature Type": "Metal Loss", "ILI Chainage (m)": 1000.0, "ID#": "F-2"},
        {"Feature Type": "Metal Loss", "ILI Chainage (m)": 1050.0, "ID#": "F-3"},
    ]
    df = pd.DataFrame(rows)
    col_map = {"feature_type": "Feature Type", "distance": "ILI Chainage (m)", "feature_id": "ID#"}
    result = filter_ili_by_gw_count(df, target_gw_chainage=1000.0, ili_col_map=col_map)
    # ±30m window: 960 is in range (1000-30=970? no — 960 < 970, excluded)
    # 1000 and 1050 (1050 > 1030) out as well — only 1000 is in [970, 1030]
    chainages = result["ILI Chainage (m)"].tolist()
    assert 1000.0 in chainages
    assert 960.0 not in chainages   # outside ±30m


def test_filter_ili_by_gw_count_returns_full_df_when_chainage_col_missing():
    df = pd.DataFrame({"ID#": ["F-1", "F-2"], "Feature Type": ["Metal Loss", "Metal Loss"]})
    col_map = {"feature_type": "Feature Type", "feature_id": "ID#"}
    result = filter_ili_by_gw_count(df, target_gw_chainage=1000.0, ili_col_map=col_map)
    assert len(result) == len(df)


# -----------------------------------------------------------------------
# End-to-end: generate_dig_packages
# -----------------------------------------------------------------------

def _build_minimal_template() -> bytes:
    """Minimal template: column A labels (anchors), column B values; excavation/exposure/feature blocks."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dig Package"

    row_labels = [
        "Dig Name",
        "Rev #",
        "Pipeline Name",
        "Pipe NPS",
        "Pipe NWT",
        "MOP",
        "SEP",
        "Lat (deg)",
        "Long (deg)",
        "Milepost",
        "Pipe Year",
        "Pipe Grade",
        "Originating ILI",
        "ILI Time",
        "Upstream AGM",
        "Downstream AGM",
        "Number of Excavations",
        "Issue Date",
        "Target Girth Weld (TGW)",
    ]
    for i, lab in enumerate(row_labels, start=1):
        ws.cell(i, 1).value = lab

    ws["A20"] = "Excavation"
    ws["A25"] = "Exposure"
    ws["A30"] = "Feature ID"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_minimal_mdl(dig_id: int = 6000, pipeline_name: str = "Test Line") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dig Notification Log"
    ws.append(
        [
            "Dig ID",
            "Dig Name",
            "Feature ID",
            "Pipeline Name",
            "Length (mm)",
            "Width (mm)",
            "Target Girth Weld (TGW)",
            "ILI Run Name",
            "Total Assessment Length (m)",
            "Start Assessment to TGW (m)",
            "End Assessment to TGW (m)",
        ]
    )
    ws.append(
        [
            dig_id,
            f"ID{dig_id}_TestDig",
            "F-001",
            pipeline_name,
            25.4,
            12.7,
            3180,
            "2025 Test Run",
            9.0,
            -0.5,
            8.5,
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_minimal_ili(tgw_joint: int = 3180) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies Listing"
    ws.append(["Feature ID", "Feature Type", "Description", "Length (mm)", "Width (mm)",
               "Peak Depth", "Orientation (hh:mm)", "ILI Chainage (m)", "Joint No. or US GW No."])
    # Three GWDs upstream of TGW
    for gwd, ch in [(3177, 940.0), (3178, 960.0), (3179, 980.0)]:
        ws.append([f"GW-{gwd}", "Girth Weld", "GWD", 0, 0, 0, "12:00", ch, gwd])
    # TGW
    ws.append(["GW-3180", "Girth Weld", "Target GWD", 0, 0, 0, "12:00", 1000.0, tgw_joint])
    # Target anomaly
    ws.append(["F-001", "Metal Loss", "Corrosion", 25.4, 12.7, 42, "03:30", 1005.0, tgw_joint])
    # Three GWDs downstream
    for gwd, ch in [(3181, 1020.0), (3182, 1040.0), (3183, 1060.0)]:
        ws.append([f"GW-{gwd}", "Girth Weld", "GWD", 0, 0, 0, "12:00", ch, gwd])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_minimal_mdl_two_digs() -> bytes:
    """MDL with two dig IDs to test max_digs limiting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dig Notification Log"
    hdr = [
        "Dig ID",
        "Dig Name",
        "Feature ID",
        "Pipeline Name",
        "Length (mm)",
        "Width (mm)",
        "Target Girth Weld (TGW)",
        "ILI Run Name",
        "Total Assessment Length (m)",
        "Start Assessment to TGW (m)",
        "End Assessment to TGW (m)",
    ]
    ws.append(hdr)
    row = [
        6000,
        "ID6000_TestDig",
        "F-001",
        "Test Line",
        25.4,
        12.7,
        3180,
        "2025 Test Run",
        9.0,
        -0.5,
        8.5,
    ]
    ws.append(row)
    ws.append(
        [
            6001,
            "ID6001_OtherDig",
            "F-002",
            "Test Line",
            10.0,
            10.0,
            3180,
            "2025 Test Run",
            9.0,
            -0.5,
            8.5,
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_generate_dig_packages_skip_pdf_skips_pdf_file():
    """skip_pdf=True produces Excel only (no PDF in ZIP)."""
    mdl_content = _build_minimal_mdl(dig_id=6000)
    ili_content = _build_minimal_ili(tgw_joint=3180)
    template_content = _build_minimal_template()

    zip_buffer = generate_dig_packages(
        mdl_content=mdl_content,
        ili_contents=[ili_content],
        template_content=template_content,
        revision="0",
        ili_formats=["Rosen-MFLA"],
        skip_pdf=True,
    )

    with zipfile.ZipFile(zip_buffer) as zf:
        names = zf.namelist()
        assert not any(n.lower().endswith(".pdf") for n in names)
        summary = json.loads(zf.read([n for n in names if n.startswith("Dig_Package_Generation_Summary")][0]))
        assert summary.get("skip_pdf") is True
        assert summary["generated"][0].get("pdf_generated") is False


def test_generate_dig_packages_max_digs_processes_first_id_only():
    """max_digs=1 yields one xlsx / one generated row when MDL lists multiple digs."""
    mdl_content = _build_minimal_mdl_two_digs()
    ili_content = _build_minimal_ili(tgw_joint=3180)
    template_content = _build_minimal_template()

    zip_buffer = generate_dig_packages(
        mdl_content=mdl_content,
        ili_contents=[ili_content],
        template_content=template_content,
        revision="1",
        ili_formats=["Rosen-MFLA"],
        max_digs=1,
    )

    with zipfile.ZipFile(zip_buffer) as zf:
        xlsx_files = [n for n in zf.namelist() if n.endswith(".xlsx")]
        assert len(xlsx_files) == 1
        summary_files = [n for n in zf.namelist() if n.startswith("Dig_Package_Generation_Summary")]
        summary = json.loads(zf.read(summary_files[0]))
        assert summary["max_digs"] == 1
        assert len(summary["dig_ids_in_mdl"]) == 2
        assert summary["dig_ids_requested"] == ["6000"]
        assert len(summary["generated"]) == 1
        assert summary["generated"][0]["dig_id"] == "6000"


def test_generate_dig_packages_end_to_end_produces_zip_with_xlsx():
    """Full pipeline: MDL + ILI + template → ZIP containing an xlsx for dig 6000."""
    mdl_content = _build_minimal_mdl(dig_id=6000)
    ili_content = _build_minimal_ili(tgw_joint=3180)
    template_content = _build_minimal_template()

    zip_buffer = generate_dig_packages(
        mdl_content=mdl_content,
        ili_contents=[ili_content],
        template_content=template_content,
        revision="1",
        ili_formats=["Rosen-MFLA"],
    )

    assert zip_buffer is not None
    with zipfile.ZipFile(zip_buffer) as zf:
        names = zf.namelist()
        xlsx_files = [n for n in names if n.endswith(".xlsx")]
        assert len(xlsx_files) == 1
        assert "ID6000_TestDig_DP_R1.xlsx" in xlsx_files[0] or "6000" in xlsx_files[0]

        # Verify summary JSON is present and has the correct dig_id
        summary_files = [n for n in names if n.startswith("Dig_Package_Generation_Summary")]
        assert summary_files
        summary = json.loads(zf.read(summary_files[0]))
        assert len(summary["generated"]) == 1
        assert str(summary["generated"][0]["dig_id"]) == "6000"
        assert summary["generated"][0]["features_matched"] >= 1


def test_generate_dig_packages_progress_callback_is_called():
    """progress_callback receives phase-tagged updates (parse + dig_generation)."""
    calls = []

    def _cb(current, total, **kwargs):
        calls.append((current, total, kwargs.get("phase", "")))

    mdl_content = _build_minimal_mdl(dig_id=6000)
    ili_content = _build_minimal_ili(tgw_joint=3180)
    template_content = _build_minimal_template()

    generate_dig_packages(
        mdl_content=mdl_content,
        ili_contents=[ili_content],
        template_content=template_content,
        revision="0",
        ili_formats=["Rosen-MFLA"],
        progress_callback=_cb,
    )

    phases = [ph for _, _, ph in calls]
    assert "parse_mdl" in phases
    assert "parse_ili" in phases
    dig_calls = [(c, t) for c, t, ph in calls if ph == "dig_generation"]
    assert dig_calls == [(0, 1), (1, 1)]


def test_generate_dig_packages_handles_failed_ili_parse_gracefully():
    """A corrupted second ILI file should not abort generation of the first."""
    mdl_content = _build_minimal_mdl(dig_id=6000)
    good_ili = _build_minimal_ili(tgw_joint=3180)
    bad_ili = b"this is not an xlsx file"
    template_content = _build_minimal_template()

    zip_buffer = generate_dig_packages(
        mdl_content=mdl_content,
        ili_contents=[good_ili, bad_ili],
        template_content=template_content,
        revision="0",
        ili_formats=["Rosen-MFLA", "TDW"],
    )

    with zipfile.ZipFile(zip_buffer) as zf:
        xlsx_files = [n for n in zf.namelist() if n.endswith(".xlsx")]
        assert len(xlsx_files) == 1  # still generated from the good ILI


def test_generate_dig_packages_template_cell_value_populated():
    """Named range tmp_pipNme in generated xlsx matches MDL pipeline name."""
    from openpyxl import load_workbook as _load

    mdl_content = _build_minimal_mdl(dig_id=6000, pipeline_name="Alpha Pipeline")
    ili_content = _build_minimal_ili(tgw_joint=3180)
    template_content = _build_minimal_template()

    zip_buffer = generate_dig_packages(
        mdl_content=mdl_content,
        ili_contents=[ili_content],
        template_content=template_content,
        revision="0",
        ili_formats=["Rosen-MFLA"],
    )

    with zipfile.ZipFile(zip_buffer) as zf:
        xlsx_names = [n for n in zf.namelist() if n.endswith(".xlsx")]
        wb = _load(io.BytesIO(zf.read(xlsx_names[0])))

    # tmp_pipNme is mapped to cell B3 in our minimal template
    ws = wb["Dig Package"]
    assert ws["B3"].value == "Alpha Pipeline"
