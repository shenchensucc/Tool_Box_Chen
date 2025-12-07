"""
DataProcessor Module

This module handles data processing operations for TML (Thickness Monitoring Location) data.
It provides functionality for cleaning data, filtering records, and saving processed data to Excel files.

Key Features:
- Removes duplicate columns from DataFrames
- Filters records based on various conditions
- Processes and appends data to Excel files while preserving existing sheets
- Maintains consistent column formatting in output files

Dependencies:
- pandas: For data manipulation and Excel file handling
- openpyxl: For Excel file operations and formatting
"""

import pandas as pd
import os
from openpyxl import load_workbook
from typing import Dict, Optional


class DataProcessor:
    """
    A class to handle data processing operations for TML data.
    
    This class provides methods for:
    1. Cleaning data by removing duplicate columns
    2. Filtering records based on various conditions
    3. Processing and saving data to Excel files while preserving existing sheets
    4. Maintaining consistent column formatting
    
    Attributes:
        None (all methods are static)
    """
    
    @staticmethod
    def clean_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
        """
        Removes duplicate columns from a DataFrame.
        
        Args:
            df: Input DataFrame with potential duplicate columns
            
        Returns:
            DataFrame with duplicate columns removed
            
        Note:
            - Keeps the first occurrence of each column name
            - Useful for cleaning data loaded from Excel files where column names might be duplicated
        """
        return df.loc[:, ~df.columns.duplicated()]

    @staticmethod
    def filter_records(df: pd.DataFrame, column: str, filter_condition: str, 
                      constant_values: Optional[Dict] = None) -> pd.DataFrame:
        """
        Filters records based on specified conditions.
        
        Args:
            df: Input DataFrame to filter
            column: Column name to apply filter on
            filter_condition: Filter condition to apply
                - "not_empty": Keep records where column is not empty
                - "empty": Keep records where column is empty
                - "True": Keep records where column is True
                - "False": Keep records where column is False
            constant_values: Dictionary of constant values to assign (optional)
            
        Returns:
            Filtered DataFrame
            
        Note:
            - Supports various filter conditions for different data types
            - Can assign constant values to filtered records
            - Prints the number of rows found for the filter condition
        """
        if filter_condition == "not_empty":
            filtered_df = df[df[column].notna()]
        elif filter_condition == "empty":
            filtered_df = df[df[column].isna()]
        elif filter_condition == "True":
            filtered_df = df[df[column] == True]
        elif filter_condition == "False":
            filtered_df = df[df[column] == False]
        else:
            raise ValueError(f"Invalid filter condition: {filter_condition}")

        if constant_values:
            for col, value in constant_values.items():
                filtered_df[col] = value

        print(f"Found {len(filtered_df)} rows for filter condition '{filter_condition}' on column '{column}'")
        return filtered_df

    @staticmethod
    def append_and_save(loader_assets: pd.DataFrame, loader_tml: pd.DataFrame, 
                       additional_data: pd.DataFrame, column_map: Dict[str, str], 
                       output_file: str, asset_sheet_name: str, tml_sheet_name: str) -> int:
        """
        Appends processed data to existing Excel files while preserving other sheets.
        
        Args:
            loader_assets: Template Assets data
            loader_tml: Template TML data
            additional_data: New data to append
            column_map: Mapping of source columns to target columns
            output_file: Path to output Excel file
            asset_sheet_name: Name of the Assets sheet
            tml_sheet_name: Name of the TML sheet
            
        Returns:
            Number of TML records added (excluding template rows)
            
        Note:
            - Preserves all existing sheets in the output file
            - Only updates the specified Assets and TML sheets
            - Sets consistent column widths (20) for updated sheets
            - Creates output directory if it doesn't exist
            - Handles both new and existing Excel files
            - Returns 0 if no new records to add (won't create file)
        """
        # If no new data, don't create file
        if additional_data.empty:
            print(f"No new records to add, skipping file creation for {output_file}")
            return 0
            
        cmms_system = "P1R-100"
        records_added = len(additional_data)

        # Create output directory if it doesn't exist
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        # Convert 'Equipment ID' to string to preserve leading zeros
        additional_data['Equipment ID'] = additional_data['Equipment ID'].astype(str)

        appended_tml = pd.concat([
            loader_tml.reset_index(drop=True),
            additional_data.rename(columns=column_map).assign(**{
                "CMMS System": cmms_system,
                "TML Analysis Type": "TML"
            }).reset_index(drop=True)
        ], ignore_index=True)

        appended_assets = pd.concat([
            loader_assets.reset_index(drop=True),
            additional_data["Equipment ID"].drop_duplicates().to_frame()
               .assign(**{"CMMS System": cmms_system})
               .reset_index(drop=True)
        ], ignore_index=True).drop_duplicates()

        # Check if file exists, if not create it first
        if not os.path.exists(output_file):
            # Create a new Excel file with the sheets
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                appended_tml.to_excel(writer, sheet_name=tml_sheet_name, index=False)
                appended_assets.to_excel(writer, sheet_name=asset_sheet_name, index=False)
        else:
            # Append to existing file
            with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                appended_tml.to_excel(writer, sheet_name=tml_sheet_name, index=False)
                appended_assets.to_excel(writer, sheet_name=asset_sheet_name, index=False)

        # Set column widths using openpyxl
        workbook = load_workbook(output_file)
        for sheet_name in [tml_sheet_name, asset_sheet_name]:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for column in sheet.columns:
                    sheet.column_dimensions[column[0].column_letter].width = 20
        workbook.save(output_file)

        print(f"Saved {records_added} TML records to {tml_sheet_name} and Assets to {asset_sheet_name} in {output_file}")
        return records_added

    @staticmethod
    def create_combined_output(processed_files: list, output_file: str, 
                               template_assets: pd.DataFrame, template_tml: pd.DataFrame,
                               asset_sheet_name: str = "Assets", 
                               tml_sheet_name: str = "TML") -> str:
        """
        Create a combined Excel file with all NEW data from multiple workflow outputs.
        
        Args:
            processed_files: List of paths to processed Excel files
            output_file: Path to save the combined output file
            template_assets: Template Assets data to exclude from combination
            template_tml: Template TML data to exclude from combination
            asset_sheet_name: Name of the Assets sheet (default: "Assets")
            tml_sheet_name: Name of the TML sheet (default: "TML")
            
        Returns:
            Path to the combined output file
            
        Note:
            - Reads Assets and TML sheets from all processed files
            - Excludes template rows (only combines NEW data)
            - Concatenates all Assets data and deduplicates by Equipment ID
            - Concatenates all TML data (preserves all rows)
            - Sets consistent column widths (20) for all columns
        """
        all_assets = []
        all_tml = []
        
        template_assets_len = len(template_assets)
        template_tml_len = len(template_tml)
        
        # Read all Assets and TML sheets from processed files
        for file_path in processed_files:
            if not os.path.exists(file_path):
                print(f"Warning: File not found: {file_path}, skipping")
                continue
            
            try:
                # Read Assets sheet
                assets_df = pd.read_excel(file_path, sheet_name=asset_sheet_name, dtype={"Equipment ID": str})
                
                # Skip template rows (template is at the beginning)
                if len(assets_df) > template_assets_len:
                    new_assets = assets_df.iloc[template_assets_len:].copy()
                    all_assets.append(new_assets)
                    print(f"Read {len(new_assets)} NEW Assets records from {os.path.basename(file_path)}")
                else:
                    print(f"No new Assets records in {os.path.basename(file_path)} (only template)")
                
                # Read TML sheet
                tml_df = pd.read_excel(file_path, sheet_name=tml_sheet_name, dtype={"Equipment ID": str})
                
                # Skip template rows (template is at the beginning)
                if len(tml_df) > template_tml_len:
                    new_tml = tml_df.iloc[template_tml_len:].copy()
                    all_tml.append(new_tml)
                    print(f"Read {len(new_tml)} NEW TML records from {os.path.basename(file_path)}")
                else:
                    print(f"No new TML records in {os.path.basename(file_path)} (only template)")
                    
            except Exception as e:
                print(f"Warning: Error reading {file_path}: {str(e)}, skipping")
                continue
        
        if not all_assets or not all_tml:
            raise ValueError("No valid NEW data found in processed files (only template data)")
        
        # Concatenate all NEW Assets and deduplicate
        combined_assets = pd.concat(all_assets, ignore_index=True).drop_duplicates(subset=["Equipment ID"])
        print(f"Combined Assets: {len(combined_assets)} unique equipment records (NEW data only)")
        
        # Concatenate all NEW TML data
        combined_tml = pd.concat(all_tml, ignore_index=True)
        print(f"Combined TML: {len(combined_tml)} total TML records (NEW data only)")
        
        # Create output directory if it doesn't exist
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        # Save to Excel file (NEW data only, no template)
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            combined_assets.to_excel(writer, sheet_name=asset_sheet_name, index=False)
            combined_tml.to_excel(writer, sheet_name=tml_sheet_name, index=False)
        
        # Set column widths using openpyxl
        workbook = load_workbook(output_file)
        for sheet_name in [asset_sheet_name, tml_sheet_name]:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for column in sheet.columns:
                    sheet.column_dimensions[column[0].column_letter].width = 20
        workbook.save(output_file)
        
        print(f"Saved combined output to {output_file}")
        return output_file

