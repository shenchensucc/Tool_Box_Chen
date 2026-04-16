"""
Inspection Report to APM Dataloader

Maps extracted PDF data (Circuit, CML, readings, date) to Equipment ID using source Excel,
and generates APM TM Data Loader compatible Excel (Measurements sheet).
"""

import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from .excel_reader import _find_canonical_column
from .inspection_report_parser import ExtractedReading


def _get_header_col_map(ws) -> Dict[str, int]:
    """Return {column_name: col_index} from row 1 of *ws*."""
    return {
        str(ws.cell(row=1, column=c).value): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value is not None
    }


def _clear_data_rows(ws) -> None:
    """Delete all rows after the two header rows, preserving formatting."""
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)


def _write_measurements_sheet(output_path: str, df: pd.DataFrame) -> None:
    """Write *df* into the 'Measurements' sheet of an existing workbook at *output_path*.

    Rows 1 and 2 are template headers — they are preserved exactly as-is.
    Data is appended starting at row 3, matched to columns by their row-1 name.
    """
    wb = load_workbook(output_path)

    if "Measurements" in wb.sheetnames:
        ws = wb["Measurements"]
        _clear_data_rows(ws)
        col_map = _get_header_col_map(ws)
    else:
        ws = wb.create_sheet("Measurements", 0)
        # No template headers — write our own in row 1; data starts at row 3 for consistency.
        for col_idx, header in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=header)
        col_map = {h: i + 1 for i, h in enumerate(df.columns)}

    for row_idx, row in enumerate(df.itertuples(index=False), 3):
        for col_name, value in zip(df.columns, row):
            col_idx = col_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)


def _write_assets_sheet(output_path: str, asset_rows: List[Dict]) -> None:
    """Write unique Equipment ID + CMMS System rows into the 'Assets' sheet.

    Rows 1 and 2 are template headers — preserved exactly as-is.
    Data is appended starting at row 3.
    asset_rows: [{"Equipment ID": ..., "CMMS System": ...}, ...]
    """
    wb = load_workbook(output_path)
    if "Assets" not in wb.sheetnames:
        wb.save(output_path)
        return

    ws = wb["Assets"]
    _clear_data_rows(ws)
    col_map = _get_header_col_map(ws)

    equip_col = col_map.get("Equipment ID")
    cmms_col = col_map.get("CMMS System")

    for row_idx, row in enumerate(asset_rows, 3):
        if equip_col:
            ws.cell(row=row_idx, column=equip_col, value=row.get("Equipment ID", ""))
        if cmms_col:
            ws.cell(row=row_idx, column=cmms_col, value=row.get("CMMS System", ""))

    wb.save(output_path)


# Source Excel: Circuit ID -> Equipment ID mapping
# Column aliases for source file
CIRCUIT_ALIASES = ["Circuit ID", "Circuit", "Circuit #", "CircuitID", "Circuit Number", "Sort Field"]
EQUIPMENT_ALIASES = ["Equipment ID", "Equipment", "Equip ID", "EquipmentID"]
CML_GROUP_ALIASES = ["CML Group ID", "TML Group ID", "CMLGroupID", "TMLGroupID"]
CML_ID_ALIASES = ["CML ID", "TML ID", "CML", "TML"]


def _normalize_circuit(s: str) -> str:
    """Normalize circuit ID for matching (strip, collapse spaces)."""
    return " ".join(str(s).strip().split()) if pd.notna(s) else ""


def _find_circuit_column(columns: list) -> Optional[str]:
    """Find the circuit column using all known aliases."""
    # Exact match first (case-insensitive)
    for alias in CIRCUIT_ALIASES:
        for col in columns:
            if str(col).strip().lower() == alias.lower():
                return col
    # Partial match fallback
    for col in columns:
        if any(a.lower() in str(col).lower() for a in ["circuit", "sort field"]):
            return col
    return None


def _load_source_df(source_path: str) -> pd.DataFrame:
    """Load source Excel, trying all sheets until one with a Circuit column is found."""
    xl = pd.ExcelFile(source_path)
    # Try "Source_Data" first for backward compatibility, then all other sheets
    sheet_order = (
        ["Source_Data"] + [s for s in xl.sheet_names if s != "Source_Data"]
        if "Source_Data" in xl.sheet_names
        else xl.sheet_names
    )
    for sheet in sheet_order:
        df = pd.read_excel(source_path, sheet_name=sheet, dtype=str)
        if _find_circuit_column(df.columns.tolist()):
            return df
    raise ValueError(
        f"Source file has no Circuit/Circuit #/Sort Field column. "
        f"Sheets checked: {xl.sheet_names}"
    )


def build_circuit_to_equipment_map(source_path: str, sheet_name: str = "Source_Data") -> Dict[str, str]:
    """
    Read source Excel and build Circuit ID -> Equipment ID mapping.

    Args:
        source_path: Path to source Excel
        sheet_name: Ignored — now auto-detects the first sheet with a Circuit column.

    Returns:
        Dict mapping circuit_id -> equipment_id
    """
    df = _load_source_df(source_path)
    circuit_col = _find_circuit_column(df.columns.tolist())
    equip_col = _find_canonical_column(df.columns.tolist(), "Equipment ID") or next(
        (c for c in df.columns if any(a.lower() in str(c).lower() for a in ["equipment", "equip"])), None
    )

    if not circuit_col or not equip_col:
        raise ValueError(
            f"Source file must have Circuit and Equipment ID columns. Found: {df.columns.tolist()}"
        )

    mapping = {}
    for _, row in df.iterrows():
        circ = _normalize_circuit(row.get(circuit_col, ""))
        equip = str(row.get(equip_col, "")).strip()
        if circ and equip:
            mapping[circ] = equip
    return mapping


def build_source_mapping(source_path: str) -> Tuple[Dict[str, str], Dict[tuple, str]]:
    """
    Read source Excel and build two mappings:
      1. circuit_to_equipment: circuit_id -> equipment_id
      2. circuit_cml_to_group_id: (circuit_id, cml_id) -> cml_group_id

    Looks for columns matching Circuit #/Circuit/Sort Field, Equipment ID, CML Group ID, CML ID.
    Auto-detects the first sheet that has a circuit column.
    """
    df = _load_source_df(source_path)
    cols = df.columns.tolist()

    circuit_col = _find_circuit_column(cols)
    equip_col = _find_canonical_column(cols, "Equipment ID") or next(
        (c for c in cols if any(a.lower() in str(c).lower() for a in ["equipment", "equip"])), None
    )
    cml_group_col = next(
        (c for c in cols if any(a.lower() == str(c).strip().lower() for a in CML_GROUP_ALIASES)), None
    ) or next(
        (c for c in cols if "cml group" in str(c).lower() or "tml group" in str(c).lower()), None
    )
    cml_id_col = next(
        (c for c in cols if str(c).strip().lower() in [a.lower() for a in CML_ID_ALIASES]), None
    )

    circuit_to_equipment: Dict[str, str] = {}
    circuit_cml_to_group_id: Dict[tuple, str] = {}

    for _, row in df.iterrows():
        circ = _normalize_circuit(row.get(circuit_col, "")) if circuit_col else ""
        if not circ:
            continue

        if equip_col:
            equip = str(row.get(equip_col, "")).strip()
            if equip and circ not in circuit_to_equipment:
                circuit_to_equipment[circ] = equip

        if cml_group_col and cml_id_col:
            cml = str(row.get(cml_id_col, "")).strip()
            group_id = str(row.get(cml_group_col, "")).strip()
            if cml and group_id:
                circuit_cml_to_group_id[(circ, cml)] = group_id

    return circuit_to_equipment, circuit_cml_to_group_id


PLACEHOLDER_EQUIPMENT_ID = "Can not Find Equipment ID"
PLACEHOLDER_CML_GROUP_ID = "Can not Find CML Group ID"


def generate_measurements_dataloader(
    readings: List[ExtractedReading],
    circuit_to_equipment: Optional[Dict[str, str]] = None,
    output_path: str = "",
    template_path: Optional[str] = None,
    cmms_system: str = "P1R-100",
    use_placeholder_when_missing: bool = True,
    circuit_cml_to_group_id: Optional[Dict[tuple, str]] = None,
) -> Tuple[int, List[Dict]]:
    """
    Generate APM Measurements dataloader Excel from extracted readings.

    Args:
        readings: List of ExtractedReading from PDFs
        circuit_to_equipment: Circuit ID -> Equipment ID mapping (optional). When None or circuit not found,
            uses PLACEHOLDER_EQUIPMENT_ID if use_placeholder_when_missing else skips the row.
        output_path: Path for output Excel (empty = summary only, no file)
        template_path: Optional TM_Loader template (for structure)
        cmms_system: CMMS System value (default P1R-100)
        use_placeholder_when_missing: When True, use "Need Add Equipment ID" for missing Equipment ID
        circuit_cml_to_group_id: (circuit_id, cml_id) -> CML Group ID mapping from source Excel.
            When provided, the TML Group ID in the dataloader uses the source CML Group ID value.

    Returns:
        (records_count, summary_rows for frontend display)
    """
    summary_rows = []
    rows = []
    circuit_to_equipment = circuit_to_equipment or {}
    circuit_cml_to_group_id = circuit_cml_to_group_id or {}

    for r in readings:
        equip_id = None
        if circuit_to_equipment:
            equip_id = circuit_to_equipment.get(r.circuit_id)
            if not equip_id:
                for circ, eid in circuit_to_equipment.items():
                    if _normalize_circuit(circ) == _normalize_circuit(r.circuit_id):
                        equip_id = eid
                        break
                    if r.circuit_id.startswith(circ) or circ.startswith(r.circuit_id):
                        equip_id = eid
                        break

        if not equip_id:
            if use_placeholder_when_missing:
                equip_id = PLACEHOLDER_EQUIPMENT_ID
            else:
                summary_rows.append({
                    "Circuit": r.circuit_id,
                    "CML": r.cml_id,
                    "Min Reading": r.min_reading,
                    "Date": r.measurement_date,
                    "Comments": "",
                    "Status": "Skipped - no Equipment ID",
                })
                continue

        # Resolve TML Group ID: prefer source CML Group ID, fall back to placeholder
        circ_norm = _normalize_circuit(r.circuit_id)
        tml_group_id = (
            circuit_cml_to_group_id.get((circ_norm, r.cml_id))
            or circuit_cml_to_group_id.get((r.circuit_id, r.cml_id))
            or (PLACEHOLDER_CML_GROUP_ID if circuit_cml_to_group_id else r.circuit_id)
        )

        missing_equip = equip_id == PLACEHOLDER_EQUIPMENT_ID
        missing_group = tml_group_id == PLACEHOLDER_CML_GROUP_ID
        if missing_equip and missing_group:
            status = "Missing Equipment ID and CML Group ID"
        elif missing_equip:
            status = "Missing Equipment ID"
        elif missing_group:
            status = "Missing CML Group ID"
        else:
            status = "OK"

        tml_id = r.cml_id
        measurement_date = r.measurement_date
        if measurement_date and len(measurement_date) == 10:  # YYYY-MM-DD
            measurement_date = f"{measurement_date} 00:00:00"

        rows.append({
            "Equipment ID": equip_id,
            "CMMS System": cmms_system,
            "TML Group ID": tml_group_id,
            "TML ID": tml_id,
            "Readings": str(r.min_reading),
            "Measurement Date": measurement_date or "",
        })

        summary_rows.append({
            "Circuit": r.circuit_id,
            "CML": r.cml_id,
            "Min Reading": r.min_reading,
            "Date": r.measurement_date,
            "Comments": "",
            "Status": status,
        })

    if not rows:
        return 0, summary_rows

    if output_path:
        df = pd.DataFrame(rows)
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        _default_template = (
            Path(__file__).parent.parent / "static" / "templates" / "tml" / "TM_Loader_Template.xlsx"
        )
        _tpl = template_path if (template_path and os.path.exists(template_path)) else (
            str(_default_template) if _default_template.exists() else None
        )

        if _tpl:
            import shutil
            shutil.copy(_tpl, output_path)
        else:
            with pd.ExcelWriter(output_path, engine="openpyxl") as _w:
                pd.DataFrame().to_excel(_w, sheet_name="Measurements", index=False)

        _write_measurements_sheet(output_path, df)

        # Build unique asset rows (one per Equipment ID) for the Assets sheet
        seen: set = set()
        asset_rows: List[Dict] = []
        for dl_row in rows:
            eid = dl_row["Equipment ID"]
            if eid not in seen:
                seen.add(eid)
                asset_rows.append({"Equipment ID": eid, "CMMS System": dl_row["CMMS System"]})
        _write_assets_sheet(output_path, asset_rows)

    return len(rows), summary_rows


def generate_measurements_dataloader_from_rows(
    rows: List[Dict],
    output_path: str = "",
    cmms_system: str = "P1R-100",
    template_path: Optional[str] = None,
    circuit_to_equipment: Optional[Dict[str, str]] = None,
    circuit_cml_to_group_id: Optional[Dict[tuple, str]] = None,
) -> Tuple[int, List[Dict]]:
    """
    Generate APM Measurements dataloader from pre-parsed / user-edited table rows.

    Each row should have: Circuit, CML, Min Reading, Date, Comments.
    Equipment ID and CML Group ID are looked up from source mapping when provided.
    Rows missing Circuit or CML are skipped.
    """
    dl_rows: List[Dict] = []
    summary_rows: List[Dict] = []
    circuit_to_equipment = circuit_to_equipment or {}
    circuit_cml_to_group_id = circuit_cml_to_group_id or {}

    for row in rows:
        circuit = str(row.get("Circuit") or "").strip()
        cml = str(row.get("CML") or "").strip()
        if not circuit or not cml:
            continue

        min_reading = row.get("Min Reading")
        if min_reading is None:
            min_reading = ""
        date_val = str(row.get("Date") or "").strip()
        comments = str(row.get("Comments") or "").strip()

        # Look up Equipment ID from source mapping; fall back to placeholder
        equip_id = None
        if circuit_to_equipment:
            circ_norm = _normalize_circuit(circuit)
            equip_id = circuit_to_equipment.get(circuit) or circuit_to_equipment.get(circ_norm)
            if not equip_id:
                for circ, eid in circuit_to_equipment.items():
                    if _normalize_circuit(circ) == circ_norm:
                        equip_id = eid
                        break
                    if circuit.startswith(circ) or circ.startswith(circuit):
                        equip_id = eid
                        break
        # Backward compat: if the row already has Equipment ID (old format), use it
        if not equip_id:
            equip_id = str(row.get("Equipment ID") or "").strip() or None
        if not equip_id:
            equip_id = PLACEHOLDER_EQUIPMENT_ID

        # Resolve TML Group ID: prefer source CML Group ID, fall back to placeholder
        circ_norm2 = _normalize_circuit(circuit)
        tml_group_id = (
            circuit_cml_to_group_id.get((circ_norm2, cml))
            or circuit_cml_to_group_id.get((circuit, cml))
            or (PLACEHOLDER_CML_GROUP_ID if circuit_cml_to_group_id else circuit)
        )

        missing_equip = equip_id == PLACEHOLDER_EQUIPMENT_ID
        missing_group = tml_group_id == PLACEHOLDER_CML_GROUP_ID
        if missing_equip and missing_group:
            status = "Missing Equipment ID and CML Group ID"
        elif missing_equip:
            status = "Missing Equipment ID"
        elif missing_group:
            status = "Missing CML Group ID"
        else:
            status = "OK"

        measurement_date = date_val
        if measurement_date and len(measurement_date) == 10:
            measurement_date = f"{measurement_date} 00:00:00"

        dl_rows.append({
            "Equipment ID": equip_id,
            "CMMS System": cmms_system,
            "TML Group ID": tml_group_id,
            "TML ID": cml,
            "Readings": str(min_reading),
            "Measurement Date": measurement_date or "",
            "Measurement Comment": comments,
        })
        summary_rows.append({
            "Circuit": circuit,
            "CML": cml,
            "Min Reading": min_reading,
            "Date": date_val,
            "Comments": comments,
            "Status": status,
        })

    if not dl_rows:
        return 0, summary_rows

    if output_path:
        df = pd.DataFrame(dl_rows)
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        # Resolve template: user-supplied → system default → standalone
        _default_template = (
            Path(__file__).parent.parent / "static" / "templates" / "tml" / "TM_Loader_Template.xlsx"
        )
        _tpl = template_path if (template_path and os.path.exists(template_path)) else (
            str(_default_template) if _default_template.exists() else None
        )

        if _tpl:
            import shutil
            shutil.copy(_tpl, output_path)
        else:
            with pd.ExcelWriter(output_path, engine="openpyxl") as _w:
                pd.DataFrame().to_excel(_w, sheet_name="Measurements", index=False)

        _write_measurements_sheet(output_path, df)

        # Build unique asset rows for the Assets sheet
        seen: set = set()
        asset_rows: List[Dict] = []
        for dl_row in dl_rows:
            eid = dl_row["Equipment ID"]
            if eid not in seen:
                seen.add(eid)
                asset_rows.append({"Equipment ID": eid, "CMMS System": dl_row["CMMS System"]})
        _write_assets_sheet(output_path, asset_rows)

    return len(dl_rows), summary_rows
