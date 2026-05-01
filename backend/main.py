import asyncio
import io
import os
import shutil
import time
from pathlib import Path

# Load .env from project root (secrets stay out of git)
try:
    from dotenv import load_dotenv
    _env_path = Path(__file__).resolve().parent.parent / ".env"
    load_dotenv(_env_path)
    if not _env_path.exists():
        load_dotenv()  # Fallback: load from current working directory
except ImportError:
    pass
import tempfile
import zipfile
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

import numpy as np
from backend.logging_config import get_logger, log_params, log_error

logger = get_logger("backend.main")
import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

from backend.models import (
    ChatMessage,
    ChatRequest,
    ChatResponse,
    ColumnStats,
    DeactivateCMLResponse,
    FeatureMapResponse,
    GenerateFromTableRequest,
    HealthResponse,
    HistogramData,
    InspectionReportResponse,
    PreviewResponse,
    ProcessResponse,
    TMLProcessResponse,
)
from backend.tml.deactivate_cml import process_deactivate_cml
from backend.tml.file_handler import FileHandler
from backend.tml.new_cml_helper.pipeline import run_analyze, run_generate, run_refine
from backend.tml.new_cml_helper.schema import (
    AssistantPlan,
    NewCMLAnalyzeResponse,
    NewCMLRefineRequest,
    NewCMLRefineResponse,
)
from backend.tml.tml_batch_runner import TMLBatchError, run_tml_batch
from backend.pipeline.metal_loss import assess_metal_loss_feature, mass_assess_metal_loss
from backend.pipeline.ili_parse import parse_ili_file
from backend.pipeline.ili_reader import (
    identify_ili_columns,
    parse_pasted_ili_text,
    read_ili_data,
)
from backend.pipeline.feature_map_builder import (
    build_feature_map_from_df,
    parse_orientation_to_degrees,
    parse_orientation_to_hours,
    format_orientation_hours,
)
from backend.pipeline.dig_package_reader import build_feature_map_from_dig_package
from backend.pipeline.report_generator import generate_word_report
from backend.pipeline.dig_package import (
    DEFAULT_DIG_PACKAGE_TEMPLATE_FILENAME,
    generate_dig_packages,
    read_default_dig_package_template_bytes,
)
from backend.docs_loader import get_relevant_context
from backend.llm_config import get_chat_base_url, get_api_key, DEFAULT_MODEL
from backend.tools.web_search import web_search
from backend.tools.schemas import WEB_SEARCH_SCHEMA

app = FastAPI(title="Chen's Engineer Toolbox API", version="0.1.0")

# ---------------------------------------------------------------------------
# Dig-package generation progress store
#
# Keyed by job_id (str UUID). Each entry:
#   {"current": int, "total": int, "status": "running"|"done"|"error",
#    "expires_at": float (Unix timestamp)}
#
# Entries are kept for DIG_PACKAGE_PROGRESS_TTL_SECS after completion so a
# final poll from the frontend never hits a 404 due to a race on cleanup.
# ---------------------------------------------------------------------------
_DIG_PACKAGE_PROGRESS: Dict[str, Dict[str, Any]] = {}
DIG_PACKAGE_PROGRESS_TTL_SECS = 300  # 5 minutes


def _cleanup_stale_progress_entries() -> None:
    """Remove expired progress entries (called lazily on each new job start)."""
    now = time.time()
    stale = [k for k, v in _DIG_PACKAGE_PROGRESS.items() if v.get("expires_at", 0) < now]
    for k in stale:
        _DIG_PACKAGE_PROGRESS.pop(k, None)


# PDF.js (Mozilla) for in-app preview — canvas rendering, not the browser PDF plug-in.
_PDFJS_DIR = Path(__file__).resolve().parent / "static" / "pdfjs"
if _PDFJS_DIR.is_dir():
    app.mount(
        "/static/pdfjs",
        StaticFiles(directory=str(_PDFJS_DIR)),
        name="pdfjs",
    )

# CORS middleware for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def log_404_middleware(request, call_next):
    """Log 404 responses with path and method for debugging."""
    response = await call_next(request)
    if response.status_code == 404:
        logger.warning(f"[404] {request.method} {request.url.path} — Route not found. Restart backend if you added new endpoints.")
    return response

MAX_FILE_SIZE = 200 * 1024 * 1024  # 200 MB

# Filesystem-based token store: files live in a dedicated temp dir named by their UUID token.
# This works across multiple uvicorn workers (no shared in-process dict needed).
_TOKEN_DIR = Path(tempfile.gettempdir()) / "toolbox_tokens"
_TOKEN_DIR.mkdir(exist_ok=True)

_TOKEN_TTL_SECONDS = 3600  # clean up tokens older than 1 hour


def _token_path(token: str, suffix: str) -> Path:
    """Return the file path for a given download token."""
    return _TOKEN_DIR / f"{token}{suffix}"


def _store_token(file_path: str) -> str:
    """Register a file for download and return a UUID token.
    The file is moved (or hard-linked) into the token directory so it survives
    even if the caller's temp dir is cleaned up.
    """
    token = str(uuid.uuid4())
    src = Path(file_path)
    dest = _TOKEN_DIR / f"{token}{src.suffix}"
    try:
        src.rename(dest)  # atomic on same filesystem
    except OSError:
        import shutil as _shutil
        _shutil.copy2(src, dest)
    return token


def _resolve_token(token: str) -> Optional[Path]:
    """Return the file path for a token, or None if not found / expired."""
    # Scan for any file whose stem matches the token
    matches = list(_TOKEN_DIR.glob(f"{token}.*"))
    if not matches:
        return None
    p = matches[0]
    if not p.exists():
        return None
    # Honour TTL
    import time
    if time.time() - p.stat().st_mtime > _TOKEN_TTL_SECONDS:
        try:
            p.unlink()
        except OSError:
            pass
        return None
    return p


def _cleanup_old_tokens() -> None:
    """Remove token files older than TTL. Call occasionally from background tasks."""
    import time
    for p in _TOKEN_DIR.iterdir():
        try:
            if time.time() - p.stat().st_mtime > _TOKEN_TTL_SECONDS:
                p.unlink()
        except OSError:
            pass


def validate_excel_file(file: UploadFile) -> None:
    """Validate Excel file type and size"""
    # Validate file type
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Validate file size
    validate_file_size(file)


def validate_new_cml_source_upload(file: UploadFile) -> None:
    """CSV / Excel uploads for New CML Helper analyze/generate."""
    if not file.filename or not file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Each source file must be .csv, .xlsx, or .xls",
        )
    validate_file_size(file)


def validate_pdf_file(file: UploadFile) -> None:
    """Validate PDF file type and size"""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF (.pdf)")
    validate_file_size(file)


def validate_file_size(file: UploadFile) -> None:
    """Validate uploaded file size"""
    file.file.seek(0, 2)  # Seek to end
    file_size = file.file.tell()
    file.file.seek(0)  # Reset to start

    if file_size > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=413, detail=f"File too large. Maximum size is {MAX_FILE_SIZE // 1024 // 1024} MB"
        )


def save_temp_file(upload_file: UploadFile) -> Path:
    """Save uploaded file to temporary location"""
    suffix = Path(upload_file.filename).suffix
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = upload_file.file.read()
        tmp.write(content)
        return Path(tmp.name)



def calculate_stats(series: pd.Series) -> ColumnStats:
    """Calculate statistics for a numeric series"""
    desc = series.describe()
    return ColumnStats(
        count=int(desc["count"]),
        mean=float(desc["mean"]),
        std=float(desc["std"]),
        min=float(desc["min"]),
        max=float(desc["max"]),
        q25=float(desc["25%"]),
        q50=float(desc["50%"]),
        q75=float(desc["75%"]),
    )


def create_histogram(series: pd.Series, column_name: str, bins: int = 30) -> HistogramData:
    """Create histogram data for a numeric series"""
    # Remove NaN values
    clean_series = series.dropna()

    if len(clean_series) == 0:
        return HistogramData(
            column_name=column_name, values=[], bin_edges=[], counts=[]
        )

    counts, bin_edges = np.histogram(clean_series, bins=bins)

    return HistogramData(
        column_name=column_name,
        values=clean_series.tolist(),
        bin_edges=bin_edges.tolist(),
        counts=counts.tolist(),
    )


@app.on_event("startup")
async def startup_log_routes():
    """Log key routes at startup and pre-warm the OCR worker process."""
    ili_routes = [r for r in app.routes if hasattr(r, "path") and "ili" in r.path]
    insp_routes = [r for r in app.routes if hasattr(r, "path") and "inspection-report" in r.path]
    logger.info(f"[startup] ILI routes: {[(r.path, getattr(r, 'methods', '')) for r in ili_routes]}")
    logger.info(f"[startup] Inspection report routes: {[(r.path, getattr(r, 'methods', '')) for r in insp_routes]}")

    logger.info("[startup] HTTP API is ready (/health, /docs).")




@app.get("/")
async def root():
    """Redirect root to API docs"""
    return RedirectResponse(url="/docs", status_code=302)


@app.get("/api/search")
async def api_web_search(q: str, max_results: int = 6):
    """
    Web search endpoint using AI Builders Space MCP (Tavily).
    Query parameter: q (search query)
    """
    from backend.tools.web_search import web_search
    result = web_search(query=q, max_results=max_results)
    return {"query": q, "result": result}


@app.get("/api/chat/models")
async def list_chat_models():
    """List available LLM models for Chat with Chen."""
    from backend.llm_config import LLM_OPTIONS
    return {"models": LLM_OPTIONS, "default": DEFAULT_MODEL}


MAX_AGENT_TURNS = 3


def _to_dict(obj):
    """Convert OpenAI response object to dict for API."""
    if hasattr(obj, "model_dump"):
        return obj.model_dump()
    if hasattr(obj, "dict"):
        return obj.dict()
    return dict(obj) if obj else {}


@app.post("/api/chat", response_model=ChatResponse)
async def chat_completion(request: ChatRequest):
    """
    Chat completion with full agentic loop: tool calls (web_search) executed
    and fed back to the LLM, up to MAX_AGENT_TURNS times.
    """
    import json

    log_params(logger, "chat", {"model": request.model, "msg_count": len(request.messages)})
    logger.info("[Agent] Starting chat completion")

    api_key = get_api_key()
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="AI_BUILDER_TOKEN not configured. Copy .env.example to .env and set your token.",
        )

    # Build messages
    messages: List[dict] = [
        {"role": m.role, "content": m.content}
        for m in request.messages
    ]

    # Inject tool docs context for tool-related questions (last user message)
    last_user = next((m for m in reversed(request.messages) if m.role == "user"), None)
    if last_user and last_user.content:
        ctx = get_relevant_context(last_user.content)
        if ctx:
            system_ctx = (
                "You are 'Chen', a helpful assistant for Chen's Engineer Toolbox. "
                "You have access to the following tool documentation. Use it to answer questions about the tools.\n\n"
                f"{ctx}"
            )
            if messages and messages[0].get("role") == "system":
                messages[0]["content"] = system_ctx + "\n\n" + messages[0]["content"]
            else:
                messages.insert(0, {"role": "system", "content": system_ctx})

    tools = request.tools or [WEB_SEARCH_SCHEMA]
    tool_choice = request.tool_choice or "auto"

    from openai import OpenAI
    client = OpenAI(
        base_url=get_chat_base_url(),
        api_key=api_key,
    )

    msg = None
    turn = 0
    last_tool_calls: List = []

    while turn < MAX_AGENT_TURNS:
        turn += 1
        logger.info(f"[Agent] Turn {turn}/{MAX_AGENT_TURNS}")

        try:
            response = client.chat.completions.create(
                model=request.model,
                messages=messages,
                tools=tools,
                tool_choice=tool_choice,
                max_tokens=2048,
            )
        except Exception as e:
            log_error(logger, "chat", e)
            raise HTTPException(status_code=502, detail=str(e))

        choice = response.choices[0] if response.choices else None
        if not choice:
            raise HTTPException(status_code=502, detail="No completion returned")

        msg = choice.message
        last_tool_calls = getattr(msg, "tool_calls", None) or []

        if not last_tool_calls:
            content = getattr(msg, "content", None) or (msg.get("content") if isinstance(msg, dict) else "") or ""
            logger.info(f"[Agent] Final Answer: '{content[:200]}{'...' if len(content) > 200 else ''}'")
            break

        # Execute tool calls
        assistant_msg = _to_dict(msg)
        messages.append(assistant_msg)

        for tc in last_tool_calls:
            fn = getattr(tc, "function", None) or (tc.get("function") if isinstance(tc, dict) else None)
            if not fn:
                continue
            name = fn.name if hasattr(fn, "name") else fn.get("name")
            args_str = fn.arguments if hasattr(fn, "arguments") else fn.get("arguments", "{}")
            logger.info(f"[Agent] Decided to call tool: '{name}'")

            result = ""
            if name == "web_search":
                try:
                    args = json.loads(args_str) if isinstance(args_str, str) else args_str
                    query = args.get("query", "")
                    result = web_search(query)
                except Exception as e:
                    result = f"Error: {str(e)}"
            else:
                result = f"Unknown tool: {name}"

            tc_id = tc.id if hasattr(tc, "id") else tc.get("id", "call_0")
            messages.append({
                "role": "tool",
                "tool_call_id": tc_id,
                "content": result,
            })
            preview = result[:150] + "..." if len(result) > 150 else result
            logger.info(f"[System] Tool Output: '{preview}'")

    content = getattr(msg, "content", None) or (msg.get("content") if isinstance(msg, dict) else "") or ""
    if not content and turn >= MAX_AGENT_TURNS and last_tool_calls:
        content = "(Max turns reached; could not complete tool chain.)"
        logger.warning("[Agent] Max turns reached with pending tool calls")
    return ChatResponse(
        content=content or "(No response)",
        model=request.model,
        tool_calls=None,
    )


@app.get("/health", response_model=HealthResponse)
async def health_check():
    """Health check endpoint"""
    return HealthResponse(ok=True)


@app.get("/api/tml/download-template/{template_type}")
async def download_template(template_type: str):
    """
    Download blank template files for TML Data Loader
    
    Args:
        template_type: Type of template to download ("source" or "tm_loader")
    
    Returns:
        Excel template file
    """
    # Define template file paths
    template_dir = Path(__file__).parent / "static" / "templates" / "tml"
    
    templates = {
        "source": {
            "path": template_dir / "Source_Data_Template.xlsx",
            "filename": "Source_Data_Template.xlsx"
        },
        "tm_loader": {
            "path": template_dir / "TM_Loader_Template.xlsx",
            "filename": "TM_Loader_Template.xlsx"
        }
    }
    
    if template_type not in templates:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid template type. Must be 'source' or 'tm_loader'"
        )
    
    template_info = templates[template_type]
    template_path = template_info["path"]
    
    if not template_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Template file not found. Please ensure {template_info['filename']} exists in backend/static/templates/tml/"
        )
    
    return FileResponse(
        path=str(template_path),
        filename=template_info["filename"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={template_info['filename']}"}
    )


@app.post("/api/ili/preview", response_model=PreviewResponse)
async def preview_excel(file: UploadFile = File(...)):
    """
    Preview an Excel file and return sheet names, columns, and row counts
    """
    log_params(logger, "ili/preview", {"filename": file.filename, "content_type": file.content_type})
    validate_excel_file(file)
    content = await file.read()
    filename = file.filename

    def _do_preview() -> PreviewResponse:
        suffix = Path(filename).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)
        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            columns: Dict[str, List[str]] = {}
            row_counts: Dict[str, int] = {}
            for sheet_name in sheet_names:
                df = pd.read_excel(tmp_path, sheet_name=sheet_name)
                columns[sheet_name] = df.columns.tolist()
                row_counts[sheet_name] = len(df)
            wb.close()
            logger.info(f"[ili/preview] Processed: {len(sheet_names)} sheets, row_counts={row_counts}")
            return PreviewResponse(
                filename=filename,
                sheet_names=sheet_names,
                columns=columns,
                row_counts=row_counts,
            )
        finally:
            if tmp_path.exists():
                os.unlink(tmp_path)

    try:
        return await asyncio.to_thread(_do_preview)
    except HTTPException:
        raise
    except Exception as e:
        log_error(logger, "ili/preview", e)
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")


@app.post("/api/ili/process", response_model=ProcessResponse)
async def process_ili_data(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    distance_column: str = Form(None),
    depth_column: str = Form(None),
    metal_loss_column: str = Form(None),
):
    """
    Process ILI data from Excel file and return statistics and plot data
    """
    log_params(logger, "ili/process", {
        "filename": file.filename,
        "sheet_name": sheet_name,
        "distance_column": distance_column,
        "depth_column": depth_column,
        "metal_loss_column": metal_loss_column,
    })
    validate_excel_file(file)
    content = await file.read()
    filename = file.filename

    def _do_process() -> ProcessResponse:
        suffix = Path(filename).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)
        try:
            df = pd.read_excel(tmp_path, sheet_name=sheet_name)
            if df.empty:
                raise HTTPException(status_code=400, detail="Sheet is empty")

            ili_cols = identify_ili_columns(df)
            dist_col = distance_column or ili_cols.get("distance")
            dep_col = depth_column or ili_cols.get("depth")
            ml_col = metal_loss_column or ili_cols.get("depth")

            columns_to_analyze = [
                c for c in [dist_col, dep_col, ml_col] if c and c in df.columns
            ] or df.select_dtypes(include=[np.number]).columns.tolist()

            stats: Dict[str, ColumnStats] = {}
            histograms: List[HistogramData] = []
            for col in columns_to_analyze:
                if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
                    series = df[col].dropna()
                    if len(series) > 0:
                        stats[col] = calculate_stats(series)
                        histograms.append(create_histogram(series, col))

            scatter_data = None
            if dist_col and dist_col in df.columns:
                scatter_data = {"x_column": dist_col, "x_values": df[dist_col].tolist()}
                y_data = {}
                if dep_col and dep_col in df.columns:
                    y_data["depth"] = df[dep_col].tolist()
                if ml_col and ml_col in df.columns:
                    y_data["metal_loss"] = df[ml_col].tolist()
                scatter_data["y_data"] = y_data

            logger.info(
                f"[ili/process] Processed: sheet={sheet_name}, total_rows={len(df)}, "
                f"stats_columns={list(stats.keys())}, histograms={len(histograms)}"
            )
            return ProcessResponse(
                filename=filename,
                sheet_name=sheet_name,
                total_rows=len(df),
                stats=stats,
                histograms=histograms,
                scatter_data=scatter_data,
            )
        finally:
            if tmp_path.exists():
                os.unlink(tmp_path)

    try:
        return await asyncio.to_thread(_do_process)
    except HTTPException:
        raise
    except Exception as e:
        log_error(logger, "ili/process", e)
        raise HTTPException(status_code=400, detail=f"Error processing Excel file: {str(e)}")


def _apply_gwd_filter(
    features: List[Dict],
    scatter_data: Optional[Dict],
    gwd_start: Optional[int],
    gwd_end: Optional[int],
    gwd_center: Optional[int],
) -> tuple:
    """
    Filter features by GWD range. Returns (filtered_features, filtered_scatter_data).
    - gwd_start, gwd_end: chainage between start and end GWD
    - gwd_center: ±3 adjacent GWDs (by sorted GWD list order)
    """
    if not scatter_data or not scatter_data.get("girth_welds"):
        return features, scatter_data

    def _gwd_match(a, b):
        try:
            return int(a) == int(b)
        except (TypeError, ValueError):
            return a == b

    girth_welds = scatter_data["girth_welds"]
    gwd_list = [(gw.get("gwd_number"), gw.get("chainage")) for gw in girth_welds if gw.get("gwd_number") is not None and gw.get("chainage") is not None]
    gwd_list.sort(key=lambda x: (x[1], str(x[0])))  # by chainage, then gwd_number

    chainage_min, chainage_max = None, None

    if gwd_center is not None:
        idx = next((i for i, (gn, _) in enumerate(gwd_list) if _gwd_match(gn, gwd_center)), None)
        if idx is not None:
            lo = max(0, idx - 3)
            hi = min(len(gwd_list) - 1, idx + 3)
            chainage_min = gwd_list[lo][1]
            chainage_max = gwd_list[hi][1]
    elif gwd_start is not None or gwd_end is not None:
        start_chainage = next((ch for gn, ch in gwd_list if _gwd_match(gn, gwd_start)), None) if gwd_start is not None else None
        end_chainage = next((ch for gn, ch in gwd_list if _gwd_match(gn, gwd_end)), None) if gwd_end is not None else None
        if start_chainage is not None:
            chainage_min = start_chainage
        if end_chainage is not None:
            chainage_max = end_chainage
        if chainage_min is None and chainage_max is not None:
            chainage_min = min(ch for _, ch in gwd_list)
        if chainage_max is None and chainage_min is not None:
            chainage_max = max(ch for _, ch in gwd_list)

    if chainage_min is None and chainage_max is None:
        return features, scatter_data

    filtered = [f for f in features if (chainage_min is None or f["x"] >= chainage_min) and (chainage_max is None or f["x"] <= chainage_max)]
    new_scatter = {
        "x_column": scatter_data["x_column"],
        "x_values": [f["x"] for f in filtered],
        "y_data": {"depth": [f["y"] for f in filtered], "metal_loss": [f["y"] for f in filtered]},
        "orientation_hours": [f.get("orientation_hours", 6.0) for f in filtered],
    }
    new_girth = [gw for gw in scatter_data.get("girth_welds", []) if (chainage_min is None or gw.get("chainage", 0) >= chainage_min) and (chainage_max is None or gw.get("chainage", float("inf")) <= chainage_max)]
    new_seam = [
        sw for sw in scatter_data.get("seam_welds", [])
        if sw.get("chainage_start") is None
        or ((chainage_min is None or sw.get("chainage_start", 0) >= chainage_min) and (chainage_max is None or sw.get("chainage_end", float("inf")) <= chainage_max))
    ]
    new_scatter["girth_welds"] = new_girth
    new_scatter["seam_welds"] = new_seam
    return filtered, new_scatter


@app.post("/api/ili/parse-paste", response_model=FeatureMapResponse)
async def parse_pasted_ili(pasted_text: str = Form(...)):
    """
    Parse pasted tabular ILI data (e.g. from Excel copy) and return features for
    interactive visualization. Uses ili_reader to auto-detect columns.
    Returns same structure as /api/ili/process (stats, histograms, scatter_data).
    """
    log_params(logger, "ili/parse-paste", {"text_length": len(pasted_text)})
    pasted_text = (pasted_text or "").strip()
    if not pasted_text:
        return FeatureMapResponse(success=False, error="No pasted data provided")

    try:
        df = parse_pasted_ili_text(pasted_text)
        if df.empty or len(df) == 0:
            return FeatureMapResponse(success=False, error="Could not parse pasted data into a table")

        ili_cols = identify_ili_columns(df)
        dist_col = ili_cols.get("distance")
        if not dist_col:
            return FeatureMapResponse(
                success=False,
                error="No distance/chainage column detected. Ensure your data has a column like 'ILI Chainage (m)' or 'Distance'.",
            )

        features, scatter_data, sources = build_feature_map_from_df(df, ili_cols)

        gwd_numbers = sorted({int(f["gwd_number"]) for f in features if isinstance(f.get("gwd_number"), (int, float))})
        logger.info(f"[ili/parse-paste] Parsed {len(features)} features")
        return FeatureMapResponse(
            success=True,
            total_rows=len(features),
            column_mapping={k: v for k, v in ili_cols.items() if v},
            features=features,
            scatter_data=scatter_data,
            sources=sources,
            gwd_numbers=gwd_numbers,
        )
    except Exception as e:
        log_error(logger, "ili/parse-paste", e)
        return FeatureMapResponse(success=False, error=str(e))


@app.post("/api/ili/process-feature-map", response_model=FeatureMapResponse)
async def process_ili_feature_map(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    vendor_format: Optional[str] = Form(None),
    gwd_start: Optional[int] = Form(None),
    gwd_end: Optional[int] = Form(None),
    gwd_center: Optional[int] = Form(None),
):
    """
    Process ILI Excel file and return features for unwrapped pipe visualization.
    Uses auto-identified columns (no manual column selection).

    **Manual sheet mode:** pass ``sheet_name`` (from preview). Omit ``vendor_format``.

    **Auto sheet mode (same as Dig Package ILI parsing):** pass non-empty ``vendor_format``
    (e.g. ``Rosen-MFLA``, ``TDW``). Sheet and header are detected from the workbook; ``sheet_name`` is ignored.

    Optional GWD filter: gwd_start+gwd_end for range, or gwd_center for ±3 adjacent GWDs.
    """
    vf = (vendor_format or "").strip()
    sn = (sheet_name or "").strip()
    log_params(logger, "ili/process-feature-map", {
        "filename": file.filename,
        "sheet_name": sn or None,
        "vendor_format": vf or None,
        "gwd_start": gwd_start,
        "gwd_end": gwd_end,
        "gwd_center": gwd_center,
    })
    validate_excel_file(file)
    content = await file.read()
    filename = file.filename

    def _do_work() -> FeatureMapResponse:
        tmp_path: Optional[Path] = None
        try:
            if vf:
                df, ili_cols, used_sheet = parse_ili_file(content, vf)
                sheet_label = used_sheet or "(auto)"
            else:
                if not sn:
                    return FeatureMapResponse(
                        success=False,
                        error="Provide sheet_name for manual mode, or vendor_format for auto-detect mode.",
                    )
                suffix = Path(filename).suffix
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(content)
                    tmp_path = Path(tmp.name)
                df = pd.read_excel(tmp_path, sheet_name=sn)
                sheet_label = sn
                ili_cols = identify_ili_columns(df)

            if df.empty:
                return FeatureMapResponse(success=False, error="Sheet is empty")
            dist_col = ili_cols.get("distance")
            if not dist_col:
                return FeatureMapResponse(
                    success=False,
                    error="No distance/chainage column detected. Ensure your data has a column like 'ILI Chainage (m)' or 'Distance'.",
                )
            features, scatter_data, sources = build_feature_map_from_df(df, ili_cols)
            gwd_numbers = sorted({int(f["gwd_number"]) for f in features if isinstance(f.get("gwd_number"), (int, float))})
            if gwd_start is not None or gwd_end is not None or gwd_center is not None:
                features, scatter_data = _apply_gwd_filter(
                    features, scatter_data, gwd_start, gwd_end, gwd_center
                )
            logger.info(f"[ili/process-feature-map] Processed {len(features)} features from sheet '{sheet_label}'")
            return FeatureMapResponse(
                success=True,
                total_rows=len(features),
                column_mapping={k: v for k, v in ili_cols.items() if v},
                features=features,
                scatter_data=scatter_data,
                sources=sources,
                gwd_numbers=gwd_numbers,
            )
        finally:
            if tmp_path is not None and tmp_path.exists():
                os.unlink(tmp_path)

    try:
        return await asyncio.to_thread(_do_work)
    except HTTPException:
        raise
    except Exception as e:
        log_error(logger, "ili/process-feature-map", e)
        return FeatureMapResponse(success=False, error=f"{type(e).__name__}: {str(e)}")


@app.post("/api/ili/process-dig-package", response_model=FeatureMapResponse)
async def process_dig_package(file: UploadFile = File(...)):
    """
    Process a dig package Excel file (sectioned format with Feature summary, Joint Summary).
    Extracts ILI features from Feature summary, longseam orientation from Joint Summary.
    Uses 'Distance from TGW (m)' as default x-axis. Supports multiple ILI sources.
    """
    log_params(logger, "ili/process-dig-package", {"filename": file.filename})
    validate_excel_file(file)
    content = await file.read()

    try:
        (features, scatter_data, sources, column_mapping,
         joint_summary_parsed, feature_summary_raw) = await asyncio.to_thread(
            build_feature_map_from_dig_package, content
        )
        gwd_numbers = sorted({int(f["gwd_number"]) for f in features if isinstance(f.get("gwd_number"), (int, float))})
        logger.info(f"[ili/process-dig-package] Parsed {len(features)} features from dig package")
        return FeatureMapResponse(
            success=True,
            total_rows=len(features),
            column_mapping=column_mapping,
            features=features,
            scatter_data=scatter_data,
            sources=sources,
            gwd_numbers=gwd_numbers,
            joint_summary_parsed=joint_summary_parsed if joint_summary_parsed else None,
            feature_summary_raw=feature_summary_raw,
        )
    except ValueError as e:
        return FeatureMapResponse(success=False, error=str(e))
    except Exception as e:
        log_error(logger, "ili/process-dig-package", e)
        return FeatureMapResponse(success=False, error=f"{type(e).__name__}: {str(e)}")


def _convert_excel_to_pdf_win32(xlsx_bytes: bytes) -> bytes:
    """
    Convert Excel bytes → PDF bytes using Excel COM automation (Windows only).
    Must be called from a thread (not the async event loop) because COM is STA.
    Raises RuntimeError on failure.
    """
    import tempfile
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    xlsx_tmp = pdf_tmp = None
    excel = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(xlsx_bytes)
            xlsx_tmp = f.name
        pdf_tmp = xlsx_tmp.replace(".xlsx", ".pdf")

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        wb = excel.Workbooks.Open(
            xlsx_tmp,
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
        )
        # xlTypePDF = 0; xlQualityStandard = 0
        wb.ExportAsFixedFormat(0, pdf_tmp, 0, True)
        wb.Close(False)

        with open(pdf_tmp, "rb") as fh:
            return fh.read()
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
        for p in [xlsx_tmp, pdf_tmp]:
            if p and os.path.exists(p):
                try:
                    os.unlink(p)
                except Exception:
                    pass


@app.post("/api/ili/excel-to-pdf")
async def excel_to_pdf(file: UploadFile = File(...)):
    """
    Convert an Excel workbook to PDF using Excel COM automation (Windows / MS Office required).
    Returns raw PDF bytes with Content-Type application/pdf on success.
    Returns JSON {error: ...} with status 501 if win32com is unavailable,
    or status 500 for any other failure.
    """
    content = await file.read()
    try:
        pdf_bytes = await asyncio.to_thread(_convert_excel_to_pdf_win32, content)
        logger.info(f"[ili/excel-to-pdf] Converted {file.filename} ({len(content)} B) → PDF ({len(pdf_bytes)} B)")
        return Response(content=pdf_bytes, media_type="application/pdf")
    except ImportError:
        logger.warning("[ili/excel-to-pdf] win32com not installed — PDF conversion unavailable")
        return JSONResponse({"error": "win32com_unavailable"}, status_code=501)
    except Exception as e:
        log_error(logger, "ili/excel-to-pdf", e)
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/tml/process", response_model=TMLProcessResponse)
async def process_tml_data(
    source_file: UploadFile = File(...),
    template_file: UploadFile = File(...),
    workflows: str = Form(...),
):
    """
    Process TML data with selected workflows and generate both ZIP and combined outputs
    
    Args:
        source_file: Source Excel file
        template_file: Template Excel file (TM_Loader.xlsx)
        workflows: Comma-separated list of workflow IDs (1-20)
    
    Returns:
        JSON with tokens to download ZIP file and combined file separately
    """
    log_params(logger, "tml/process", {
        "source_filename": source_file.filename,
        "template_filename": template_file.filename,
        "workflows_raw": workflows,
    })
    # Validate file types and sizes
    validate_excel_file(source_file)
    validate_excel_file(template_file)

    # Parse workflow IDs (fast, do before offloading)
    try:
        workflow_ids = [int(w.strip()) for w in workflows.split(",") if w.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid workflow IDs format")

    if not workflow_ids:
        raise HTTPException(status_code=400, detail="No workflows selected")

    # Read file bytes async so the upload stream is consumed before handing off to a thread
    source_content = await source_file.read()
    template_content = await template_file.read()
    source_filename = source_file.filename
    template_filename = template_file.filename

    def _do_tml_process() -> TMLProcessResponse:
        temp_dir = tempfile.mkdtemp()
        temp_source: Optional[Path] = None
        temp_template: Optional[Path] = None

        try:
            # Write file bytes to temp files inside the thread
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(source_filename).suffix) as f:
                f.write(source_content)
                temp_source = Path(f.name)
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(template_filename).suffix) as f:
                f.write(template_content)
                temp_template = Path(f.name)

            try:
                return run_tml_batch(
                    temp_source,
                    temp_template,
                    workflow_ids,
                    Path(temp_dir),
                    _store_token,
                )
            except TMLBatchError as e:
                raise HTTPException(status_code=e.status_code, detail=e.detail)

        finally:
            if temp_source and temp_source.exists():
                try:
                    os.unlink(temp_source)
                except OSError:
                    pass

    try:
        return await asyncio.to_thread(_do_tml_process)
    except HTTPException:
        raise
    except Exception as e:
        log_error(logger, "tml/process", e)
        raise HTTPException(status_code=500, detail=f"Error processing TML data: {str(e)}")


@app.post("/api/tml/new-cml-helper/analyze", response_model=NewCMLAnalyzeResponse)
async def new_cml_helper_analyze(
    files: List[UploadFile] = File(...),
    notes: str = Form(""),
    model: str = Form(DEFAULT_MODEL),
):
    """Upload spreadsheets; returns column profiles plus an AI-assisted mapping/workflow plan."""
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    uploads = []
    for uf in files:
        validate_new_cml_source_upload(uf)
        uploads.append((uf.filename, await uf.read()))

    log_params(logger, "tml/new-cml-helper/analyze", {"n_files": len(uploads), "model": model})

    try:
        return await asyncio.to_thread(run_analyze, uploads, notes, model)
    except Exception as e:
        log_error(logger, "tml/new-cml-helper/analyze", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tml/new-cml-helper/refine", response_model=NewCMLRefineResponse)
async def new_cml_helper_refine(body: NewCMLRefineRequest):
    """Apply user answers to questions / constants and recompute validation."""
    try:
        return await asyncio.to_thread(run_refine, body)
    except Exception as e:
        log_error(logger, "tml/new-cml-helper/refine", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tml/new-cml-helper/generate", response_model=TMLProcessResponse)
async def new_cml_helper_generate(
    template_file: UploadFile = File(...),
    source_files: List[UploadFile] = File(...),
    plan_json: str = Form(...),
    workflows: str = Form(""),
):
    """Assemble Source_Data from uploads + confirmed plan, then run TML batch."""
    validate_excel_file(template_file)
    if not source_files:
        raise HTTPException(status_code=400, detail="No source files uploaded")

    template_content = await template_file.read()
    template_filename = template_file.filename or "template.xlsx"

    uploads_dict: Dict[str, bytes] = {}
    for uf in source_files:
        validate_new_cml_source_upload(uf)
        uploads_dict[uf.filename] = await uf.read()

    try:
        plan = AssistantPlan.model_validate_json(plan_json)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid plan_json: {e}") from e

    if workflows.strip():
        try:
            workflow_ids = [int(w.strip()) for w in workflows.split(",") if w.strip()]
        except ValueError as e:
            raise HTTPException(status_code=400, detail="Invalid workflows format") from e
        if not workflow_ids:
            raise HTTPException(status_code=400, detail="No workflows selected")
    else:
        workflow_ids = list(plan.recommended_workflows)

    log_params(logger, "tml/new-cml-helper/generate", {
        "template_filename": template_filename,
        "n_sources": len(uploads_dict),
        "workflows": workflow_ids,
    })

    def _do_gen() -> TMLProcessResponse:
        try:
            return run_generate(
                uploads_dict,
                plan,
                workflow_ids,
                template_content,
                template_filename,
                _store_token,
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        except TMLBatchError as e:
            raise HTTPException(status_code=e.status_code, detail=e.detail) from e

    try:
        return await asyncio.to_thread(_do_gen)
    except HTTPException:
        raise
    except Exception as e:
        log_error(logger, "tml/new-cml-helper/generate", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tml/deactivate-cml", response_model=DeactivateCMLResponse)
async def deactivate_cml(
    source_file: UploadFile = File(...),
    template_file: Optional[UploadFile] = File(default=None),
):
    """
    De-active CML: Generate a dataloader that deactivates all CMLs in the uploaded source sheet.
    
    - Source file: Required. Must have sheet "Source_Data" with Equipment ID, CML Group ID, sub-CML ID.
    - Template file: Optional. If not provided, uses default TM_Loader_Template.xlsx from system.
    
    Output: {source_filename}_deactive.xlsx with Status Indicator = "Inactive" for all CMLs.
    """
    tpl_name = (template_file.filename if template_file and template_file.filename else "default")
    logger.info(f"[tml/deactivate-cml] Request received: source={source_file.filename}, template={tpl_name}")
    log_params(logger, "tml/deactivate-cml", {"source_filename": source_file.filename})
    validate_excel_file(source_file)

    # Read file bytes async before handing off to thread
    source_content = await source_file.read()
    source_filename = source_file.filename

    template_content: Optional[bytes] = None
    template_filename: Optional[str] = None
    if template_file is not None and template_file.filename:
        validate_excel_file(template_file)
        template_content = await template_file.read()
        template_filename = template_file.filename

    default_template = Path(__file__).parent / "static" / "templates" / "tml" / "TM_Loader_Template.xlsx"

    if template_content is None and not default_template.exists():
        raise HTTPException(
            status_code=400,
            detail="No template file provided and default TM_Loader_Template.xlsx not found in backend/static/templates/tml/"
        )

    def _do_deactivate() -> DeactivateCMLResponse:
        temp_dir = tempfile.mkdtemp()
        temp_source_path: Optional[Path] = None

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(source_filename).suffix) as f:
                f.write(source_content)
                temp_source_path = Path(f.name)

            if template_content is not None:
                with tempfile.NamedTemporaryFile(delete=False, suffix=Path(template_filename).suffix) as f:
                    f.write(template_content)
                    temp_template_path: Path = Path(f.name)
            else:
                temp_template_path = default_template

            source_stem = Path(source_filename).stem
            output_filename = f"{source_stem}_deactive.xlsx"
            output_path = Path(temp_dir) / output_filename

            records_count, result_path, sheet_used = process_deactivate_cml(
                source_path=str(temp_source_path),
                template_path=str(temp_template_path),
                output_path=str(output_path),
            )

            if records_count == 0 or result_path is None:
                raise HTTPException(
                    status_code=400,
                    detail="No records to process. Ensure source file has a sheet with Equipment ID, CML Group ID, sub-CML ID columns (or their aliases)."
                )

            download_token = _store_token(str(result_path))

            logger.info(
                f"[tml/deactivate-cml] Deactivated {records_count} CMLs, sheet='{sheet_used}', output={output_filename}"
            )

            return DeactivateCMLResponse(
                success=True,
                message=f"Successfully deactivated {records_count} CML(s)",
                download_token=download_token,
                records_count=records_count,
                output_filename=output_filename,
                sheet_used=sheet_used,
            )
        finally:
            if temp_source_path and temp_source_path.exists():
                try:
                    os.unlink(temp_source_path)
                except OSError:
                    pass

    try:
        return await asyncio.to_thread(_do_deactivate)
    except HTTPException:
        raise
    except ValueError as e:
        logger.warning(f"[tml/deactivate-cml] ValueError: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except FileNotFoundError as e:
        logger.warning(f"[tml/deactivate-cml] FileNotFoundError: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log_error(logger, "tml/deactivate-cml", e)
        import traceback
        tb = traceback.format_exc()
        logger.error(f"[tml/deactivate-cml] Traceback: {tb}")
        raise HTTPException(
            status_code=500,
            detail=f"Error: {type(e).__name__}: {str(e)}. Check server logs for full traceback."
        )


# ---------------------------------------------------------------------------
# Inspection report OCR — Azure Document Intelligence via thread pool
# ---------------------------------------------------------------------------
# OCR requests run in a thread pool so they don't block the async event loop.
# _ocr_busy ensures only one parse is in-flight at a time.
_ocr_busy = False


async def _run_ocr(temp_pdfs: List[Path], source_filenames: List[str]):
    """Run parse_inspection_report_pdfs in a thread so it doesn't block the event loop."""
    from backend.tml.inspection_report_parser import parse_inspection_report_pdfs

    def _parse():
        return parse_inspection_report_pdfs(temp_pdfs, source_filenames)

    return await asyncio.to_thread(_parse)


@app.post("/api/tml/inspection-report/read", response_model=InspectionReportResponse)
async def read_inspection_reports(
    pdf_files: List[UploadFile] = File(..., description="UT inspection report PDFs"),
):
    """
    Read and summarize UT inspection report PDFs. No source Excel or dataloader.
    Returns extracted Circuit, CML, Min Reading, Date for user verification.
    """
    global _ocr_busy
    logger.info(f"[inspection-report/read] Request received, pdf_count={len(pdf_files)}")
    log_params(logger, "tml/inspection-report/read", {"pdf_count": len(pdf_files)})

    if _ocr_busy:
        raise HTTPException(
            status_code=503,
            detail="PDF parsing is already in progress. Please wait for it to finish and try again.",
        )

    for pf in pdf_files:
        validate_pdf_file(pf)

    pdf_data = [(await pf.read(), pf.filename) for pf in pdf_files]

    temp_pdfs: List[Path] = []
    _ocr_busy = True
    try:
        for pdf_bytes, _ in pdf_data:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
                f.write(pdf_bytes)
                temp_pdfs.append(Path(f.name))

        source_filenames = [name for _, name in pdf_data]

        readings = await _run_ocr(temp_pdfs, source_filenames)

        if not readings:
            return InspectionReportResponse(
                success=False,
                message="No data extracted from PDFs. Check report format or try OCR.",
                error="No Circuit, CML, or readings found in uploaded PDFs.",
            )

        # Dataloader summary generation is lightweight; run in thread pool.
        def _make_summary():
            from backend.tml.inspection_dataloader import generate_measurements_dataloader
            return generate_measurements_dataloader(
                readings,
                circuit_to_equipment={},
                output_path="",
                use_placeholder_when_missing=True,
            )

        records_count, summary = await asyncio.to_thread(_make_summary)

        logger.info(f"[tml/inspection-report/read] Read {len(pdf_files)} PDFs, {len(summary)} CML(s)")

        return InspectionReportResponse(
            success=True,
            message=f"Read {len(pdf_files)} PDF(s), extracted **{len(summary)}** CML(s). Use Generate Dataloader to create Excel.",
            records_count=records_count,
            summary=summary,
        )
    except HTTPException:
        raise
    except RuntimeError as e:
        # OCR worker crashed — backend is still alive
        logger.error(f"[inspection-report/read] OCR worker error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        log_error(logger, "tml/inspection-report/read", e)
        import traceback
        logger.error(f"[inspection-report/read] Traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"{type(e).__name__}: {str(e)}. Check backend logs for full traceback.",
        )
    finally:
        _ocr_busy = False
        for p in temp_pdfs:
            if p.exists():
                try:
                    os.unlink(p)
                except OSError:
                    pass


@app.post("/api/tml/inspection-report", response_model=InspectionReportResponse)
async def process_inspection_reports(
    pdf_files: List[UploadFile] = File(..., description="UT inspection report PDFs"),
    source_file: Optional[UploadFile] = File(default=None, description="Optional: Excel with Circuit ID and Equipment ID"),
    template_file: Optional[UploadFile] = File(default=None),
):
    """
    Parse PDFs and generate APM Measurements dataloader. Source Excel optional;
    when missing, Equipment ID = "Need Add Equipment ID" (incomplete dataloader).
    """
    global _ocr_busy
    log_params(logger, "tml/inspection-report", {
        "source_filename": source_file.filename if source_file else None,
        "pdf_count": len(pdf_files),
    })

    if _ocr_busy:
        raise HTTPException(
            status_code=503,
            detail="PDF parsing is already in progress. Please wait for it to finish and try again.",
        )

    for pf in pdf_files:
        validate_pdf_file(pf)
    if source_file:
        validate_excel_file(source_file)

    # Read all file bytes async before spawning any threads/processes
    pdf_bytes_list: List[tuple[bytes, str]] = [
        (await pf.read(), pf.filename) for pf in pdf_files
    ]
    source_content: Optional[bytes] = await source_file.read() if source_file else None
    source_filename_str: Optional[str] = source_file.filename if source_file else None

    template_content: Optional[bytes] = None
    template_filename_str: Optional[str] = None
    if template_file and template_file.filename:
        validate_excel_file(template_file)
        template_content = await template_file.read()
        template_filename_str = template_file.filename

    # Write PDFs to temp files (quick — stays in the async handler)
    temp_pdfs: List[Path] = []
    temp_source_path: Optional[Path] = None
    template_path_str: Optional[str] = None

    for pdf_bytes, _ in pdf_bytes_list:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
            f.write(pdf_bytes)
            temp_pdfs.append(Path(f.name))

    if source_content is not None:
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(source_filename_str).suffix) as f:
            f.write(source_content)
            temp_source_path = Path(f.name)

    if template_content is not None:
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(template_filename_str).suffix) as f:
            f.write(template_content)
            template_path_str = f.name

    source_filenames = [name for _, name in pdf_bytes_list]
    _ocr_busy = True
    try:
        # ── Phase 1: OCR ──────────────────────────────────────────────────────
        readings = await _run_ocr(temp_pdfs, source_filenames)

        if not readings:
            return InspectionReportResponse(
                success=False,
                message="No data extracted from PDFs. Check report format or try OCR.",
                error="No Circuit, CML, or readings found in uploaded PDFs.",
            )

        # ── Phase 2: Dataloader generation — lightweight; runs in thread ───
        def _generate_dataloader() -> InspectionReportResponse:
            from backend.tml.inspection_dataloader import (
                build_source_mapping,
                generate_measurements_dataloader,
            )

            circuit_to_equipment = {}
            circuit_cml_to_group_id = {}
            if temp_source_path:
                circuit_to_equipment, circuit_cml_to_group_id = build_source_mapping(str(temp_source_path))

            temp_dir = tempfile.mkdtemp()
            output_path = Path(temp_dir) / "Inspection_Report_Dataloader.xlsx"
            records_count, summary = generate_measurements_dataloader(
                readings,
                circuit_to_equipment,
                str(output_path),
                template_path=template_path_str,
                use_placeholder_when_missing=True,
                circuit_cml_to_group_id=circuit_cml_to_group_id,
            )
            if records_count == 0:
                return InspectionReportResponse(
                    success=True,
                    message="No records to write.",
                    summary=summary,
                    records_count=0,
                )

            download_token = _store_token(str(output_path))
            missing_equip = sum(1 for s in summary if "Equipment ID" in s.get("Status", ""))
            missing_group = sum(1 for s in summary if "CML Group ID" in s.get("Status", ""))
            notes = []
            if missing_equip:
                notes.append(f"{missing_equip} record(s) could not find Equipment ID — check dataloader before APM upload.")
            if missing_group:
                notes.append(f"{missing_group} record(s) could not find CML Group ID — check dataloader before APM upload.")
            logger.info(
                f"[tml/inspection-report] Processed {len(pdf_bytes_list)} PDFs, {records_count} records"
            )
            return InspectionReportResponse(
                success=True,
                message=f"Generated dataloader with {records_count} record(s) from extract reading table successfully.",
                download_token=download_token,
                output_filename="Inspection_Report_Dataloader.xlsx",
                records_count=records_count,
                summary=summary,
                notes=notes,
            )

        return await asyncio.to_thread(_generate_dataloader)

    except HTTPException:
        raise
    except RuntimeError as e:
        logger.error(f"[inspection-report] OCR worker error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log_error(logger, "tml/inspection-report", e)
        import traceback
        logger.error(f"[inspection-report] Traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"{type(e).__name__}: {str(e)}. Check backend logs for full traceback.",
        )
    finally:
        _ocr_busy = False
        for p in temp_pdfs:
            if p.exists():
                try:
                    os.unlink(p)
                except OSError:
                    pass
        if temp_source_path and temp_source_path.exists():
            try:
                os.unlink(temp_source_path)
            except OSError:
                pass


@app.post("/api/tml/inspection-report/generate-from-table", response_model=InspectionReportResponse)
async def generate_inspection_dataloader_from_table(
    rows_json: str = Form(...),
    cmms_system: str = Form("P1R-100"),
    template_file: Optional[UploadFile] = File(None),
    source_file: Optional[UploadFile] = File(None),
):
    """
    Generate APM dataloader from pre-parsed / user-edited table rows.

    Accepts multipart form: rows_json (JSON-encoded list), cmms_system, optional template_file,
    optional source_file (Excel with Circuit #, Equipment ID, CML Group ID columns).
    Falls back to system TM_Loader_Template.xlsx when no template is uploaded.
    """
    import json as _json
    try:
        rows = _json.loads(rows_json)
    except Exception:
        raise HTTPException(status_code=400, detail="rows_json is not valid JSON.")

    if not rows:
        return InspectionReportResponse(
            success=False,
            message="No rows provided.",
            error="Empty rows list.",
        )

    temp_dir = tempfile.mkdtemp()
    output_path = Path(temp_dir) / "Inspection_Report_Dataloader.xlsx"
    temp_template_path: Optional[str] = None
    temp_source_path: Optional[str] = None

    if template_file and template_file.filename:
        validate_excel_file(template_file)
        tpl_bytes = await template_file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(template_file.filename).suffix) as f:
            f.write(tpl_bytes)
            temp_template_path = f.name

    if source_file and source_file.filename:
        validate_excel_file(source_file)
        src_bytes = await source_file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(source_file.filename).suffix) as f:
            f.write(src_bytes)
            temp_source_path = f.name

    try:
        from backend.tml.inspection_dataloader import (
            build_source_mapping,
            generate_measurements_dataloader_from_rows,
        )

        circuit_to_equipment = {}
        circuit_cml_to_group_id = {}
        if temp_source_path:
            circuit_to_equipment, circuit_cml_to_group_id = build_source_mapping(temp_source_path)

        records_count, summary = generate_measurements_dataloader_from_rows(
            rows,
            str(output_path),
            cmms_system=cmms_system,
            template_path=temp_template_path,
            circuit_to_equipment=circuit_to_equipment,
            circuit_cml_to_group_id=circuit_cml_to_group_id,
        )

        if records_count == 0:
            return InspectionReportResponse(
                success=True,
                message="No valid records to write. Check that Circuit and CML columns are filled.",
                summary=summary,
                records_count=0,
            )

        download_token = _store_token(str(output_path))

        missing_equip = sum(1 for s in summary if "Equipment ID" in s.get("Status", ""))
        missing_group = sum(1 for s in summary if "CML Group ID" in s.get("Status", ""))
        notes = []
        if missing_equip:
            notes.append(f"{missing_equip} record(s) could not find Equipment ID — check dataloader before APM upload.")
        if missing_group:
            notes.append(f"{missing_group} record(s) could not find CML Group ID — check dataloader before APM upload.")

        logger.info(f"[tml/inspection-report/generate-from-table] {records_count} records")
        return InspectionReportResponse(
            success=True,
            message=f"Generated dataloader with {records_count} record(s) from extract reading table successfully.",
            download_token=download_token,
            output_filename="Inspection_Report_Dataloader.xlsx",
            records_count=records_count,
            summary=summary,
            notes=notes,
        )
    except Exception as e:
        log_error(logger, "tml/inspection-report/generate-from-table", e)
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)}")
    finally:
        if temp_template_path:
            try:
                os.unlink(temp_template_path)
            except OSError:
                pass
        if temp_source_path:
            try:
                os.unlink(temp_source_path)
            except OSError:
                pass


@app.get("/api/tml/download/{file_token}")
async def download_tml_file(file_token: str):
    """
    Download TML output file using a file token
    
    Args:
        file_token: Unique token for the file (returned from /api/tml/process)
    
    Returns:
        ZIP or Excel file based on the token
    """
    resolved = _resolve_token(file_token)
    if resolved is None:
        logger.warning(f"[tml/download] Token not found or expired: {file_token[:8]}...")
        raise HTTPException(
            status_code=404,
            detail=f"File not found. Token invalid or expired (prefix={file_token[:8]}...)."
        )

    file_path = str(resolved)
    file_name = resolved.name
    media_type = (
        "application/zip"
        if file_path.endswith(".zip")
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={file_name}"},
    )


@app.post("/api/pipeline/metal-loss/assess")
async def assess_metal_loss(
    do: float = Form(...),
    tp: float = Form(...),
    YS: float = Form(...),
    TS: float = Form(...),
    dimp_org_percent: float = Form(...),
    Limp_org: float = Form(...),
    date_ILI: str = Form(...),
    ILI_dimp_tolerance: float = Form(...),
    ILI_Limp_tolerance: float = Form(...),
    CR_low: float = Form(...),
    CR_ave: float = Form(...),
    CR_high: float = Form(...),
    month_CR: int = Form(...),
    feature_ID: str = Form(""),
    vendor_ILI: str = Form(""),
    CR_Limp: float = Form(0.0)
):
    """
    Assess metal loss feature and return calculated results.
    
    Returns:
        JSON with assessment results including depth/pressure arrays
    """
    log_params(logger, "metal-loss/assess", {
        "do": do, "tp": tp, "YS": YS, "TS": TS,
        "dimp_org_percent": dimp_org_percent, "Limp_org": Limp_org,
        "date_ILI": date_ILI, "ILI_dimp_tolerance": ILI_dimp_tolerance,
        "ILI_Limp_tolerance": ILI_Limp_tolerance,
        "CR_low": CR_low, "CR_ave": CR_ave, "CR_high": CR_high,
        "month_CR": month_CR, "feature_ID": feature_ID,
        "vendor_ILI": vendor_ILI, "CR_Limp": CR_Limp,
    })
    try:
        results = assess_metal_loss_feature(
            do=do,
            tp=tp,
            YS=YS,
            TS=TS,
            dimp_org_percent=dimp_org_percent,
            Limp_org=Limp_org,
            date_ILI=date_ILI,
            ILI_dimp_tolerance=ILI_dimp_tolerance,
            ILI_Limp_tolerance=ILI_Limp_tolerance,
            CR_low=CR_low,
            CR_ave=CR_ave,
            CR_high=CR_high,
            month_CR=month_CR,
            feature_ID=feature_ID,
            vendor_ILI=vendor_ILI,
            CR_Limp=CR_Limp
        )
        logger.info(f"[metal-loss/assess] Success for feature_ID={feature_ID}")
        return results
    except Exception as e:
        log_error(logger, "metal-loss/assess", e)
        raise HTTPException(status_code=400, detail=f"Error in assessment: {str(e)}")


@app.post("/api/pipeline/metal-loss/mass-assess")
async def mass_assess_metal_loss_endpoint(
    file: UploadFile = File(...),
    do: float = Form(...),
    tp: float = Form(...),
    YS: float = Form(...),
    TS: float = Form(...),
    depth_tolerance: float = Form(...),
    length_tolerance: float = Form(...),
    depth_cr: float = Form(4.0),
    length_cr: float = Form(25.0),
    start_year: int = Form(...)
):
    """
    Mass assess metal loss features from an Excel file.
    """
    log_params(logger, "metal-loss/mass-assess", {
        "filename": file.filename,
        "do": do, "tp": tp, "YS": YS, "TS": TS,
        "depth_tolerance": depth_tolerance, "length_tolerance": length_tolerance,
        "depth_cr": depth_cr, "length_cr": length_cr, "start_year": start_year,
    })
    validate_excel_file(file)
    content = await file.read()
    filename = file.filename

    def _do_mass_assess() -> str:
        suffix = Path(filename).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)
        try:
            df = read_ili_data(tmp_path)
            logger.info(f"[metal-loss/mass-assess] Read {len(df)} rows from Excel")
            df_result = mass_assess_metal_loss(
                df=df, do=do, tp=tp, YS=YS, TS=TS,
                depth_tolerance=depth_tolerance, length_tolerance=length_tolerance,
                depth_cr=depth_cr, length_cr=length_cr, start_year=start_year,
            )
            logger.info(f"[metal-loss/mass-assess] Processed {len(df_result)} rows, output columns={list(df_result.columns)}")
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as out:
                df_result.to_excel(out.name, index=False)
                return out.name
        finally:
            if tmp_path.exists():
                os.unlink(tmp_path)

    try:
        out_path = await asyncio.to_thread(_do_mass_assess)
        stamp = datetime.now().strftime("%Y%m%d")
        return FileResponse(
            path=out_path,
            filename=f"Mass_Metal_Loss_Assessment_{stamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Mass_Metal_Loss_Assessment_{stamp}.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error in mass assessment: {str(e)}")


@app.post("/api/pipeline/metal-loss/export-word")
async def export_word_report(
    assessment_results: str = Form(...),
    depth_growth_chart: UploadFile = File(...),
    sop_decay_chart: UploadFile = File(...),
    sop_cutoff_chart: UploadFile = File(...)
):
    """
    Generate and download Word document report.
    
    Parameters:
        assessment_results: JSON string of assessment results
        depth_growth_chart: PNG image of depth growth chart
        sop_decay_chart: PNG image of SOP decay chart
        sop_cutoff_chart: PNG image of SOP cutoff chart
    
    Returns:
        Word document (.docx) file
    """
    log_params(logger, "metal-loss/export-word", {
        "assessment_results_len": len(assessment_results) if assessment_results else 0,
        "depth_growth_chart": depth_growth_chart.filename,
        "sop_decay_chart": sop_decay_chart.filename,
        "sop_cutoff_chart": sop_cutoff_chart.filename,
    })
    try:
        import json
        
        # Parse assessment results
        results = json.loads(assessment_results)
        
        # Read chart images
        chart_images = {
            'depth_growth': await depth_growth_chart.read(),
            'sop_decay': await sop_decay_chart.read(),
            'sop_cutoff': await sop_cutoff_chart.read()
        }
        
        # Generate Word document (run in thread pool to avoid blocking async event loop)
        doc_bytes = await asyncio.to_thread(generate_word_report, results, chart_images)
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as tmp:
            tmp.write(doc_bytes)
            tmp_path = tmp.name
        
        logger.info("[metal-loss/export-word] Word report generated successfully")
        # Return file
        return FileResponse(
            path=tmp_path,
            filename=f"Metal_Loss_Assessment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx",
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename=Metal_Loss_Assessment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"}
        )
    
    except json.JSONDecodeError as e:
        log_error(logger, "metal-loss/export-word (JSON decode)", e)
        raise HTTPException(status_code=400, detail="Invalid assessment results JSON")
    except Exception as e:
        log_error(logger, "metal-loss/export-word", e)
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")


@app.post("/api/pipeline/dig-package/preview-mdl")
async def preview_mdl_endpoint(mdl_file: UploadFile = File(...)):
    """
    Parse an MDL file and return the list of valid Dig IDs without generating any packages.
    Used by the frontend to show a quick preview after upload so the user knows the file
    was read correctly before starting the full (slow) generation run.
    """
    log_params(logger, "dig-package/preview-mdl", {"filename": mdl_file.filename})
    try:
        validate_excel_file(mdl_file)
        content = await mdl_file.read()

        from backend.pipeline.dig_package import parse_mdl_file, extract_dig_ids
        mdl_df, col_map = await asyncio.to_thread(parse_mdl_file, content)
        dig_ids = extract_dig_ids(mdl_df, col_map)
        return JSONResponse({"dig_ids": [str(d) for d in dig_ids], "count": len(dig_ids)})
    except HTTPException:
        raise
    except Exception as e:
        log_error(logger, "dig-package/preview-mdl", e)
        raise HTTPException(status_code=500, detail=f"Error previewing MDL: {str(e)}")


@app.get("/api/pipeline/dig-package/blank-template-zip")
async def dig_package_blank_template_zip():
    """
    Return a small ZIP containing only the bundled Dig Package Excel template + a README.
    Use this to verify the browser can download from the API without running MDL/ILI generation
    (which may time out during PDF conversion).
    """
    log_params(logger, "dig-package/blank-template-zip", {})
    try:
        raw = await asyncio.to_thread(read_default_dig_package_template_bytes)
    except FileNotFoundError as e:
        raise HTTPException(status_code=503, detail=str(e)) from e
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(DEFAULT_DIG_PACKAGE_TEMPLATE_FILENAME, raw)
        zf.writestr(
            "README_BLANK.txt",
            "Blank Dig Package template (bundled on server)\n"
            "==============================================\n"
            "This archive is the raw .xlsx template only — no MDL or ILI data were applied.\n"
            "If you can open this file, download + template wiring are working.\n"
            "For populated packages, use Generate in the app; enable Skip PDF if Excel→PDF hangs.\n",
        )
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=Dig_Package_BLANK_TEMPLATE_{ts}.zip"},
    )


@app.get("/api/pipeline/dig-package/progress/{job_id}")
async def get_dig_package_progress(job_id: str):
    """
    Return progress for an in-flight (or recently completed) dig-package generation job.
    Returns 404 for unknown or expired job IDs.

    Fields:
    - ``phase``: e.g. ``receiving_upload``, ``parse_mdl``, ``parse_ili``, ``building_zip``
    - ``message``: human-readable detail (large ILI files can take many minutes to parse)
    """
    entry = _DIG_PACKAGE_PROGRESS.get(job_id)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"No progress entry found for job_id='{job_id}'")
    current = entry.get("current", 0)
    total = entry.get("total", 1)
    phase = entry.get("phase", "")
    # During ILI parse, current/total are file index / file count; during dig loop, dig index / dig count.
    pct = round(current / total * 100) if total > 0 else 0
    return JSONResponse({
        "job_id": job_id,
        "current": current,
        "total": total,
        "pct": pct,
        "status": entry.get("status", "running"),
        "phase": phase,
        "message": entry.get("message", ""),
    })


@app.post("/api/pipeline/dig-package/generate")
async def generate_dig_package_endpoint(
    mdl_file: UploadFile = File(...),
    ili_files: List[UploadFile] = File(default=[]),
    template_file: Optional[UploadFile] = File(default=None),
    revision: str = Form("0"),
    ili_formats: str = Form(""),
    job_id: str = Form(""),
    max_digs: str = Form(""),
    skip_pdf: str = Form(""),
    skip_ili: str = Form(""),
    include_debug: str = Form(""),
):
    """
    Generate dig packages from MDL, multiple ILI data files, and an optional template.
    If no template is uploaded, the bundled ``2026 Dig Package Template.xlsx`` under
    ``backend/static/templates/dig_package/`` is used (must exist on the server).
    If job_id is provided the endpoint writes per-dig progress to the progress store
    so the frontend can poll GET /api/pipeline/dig-package/progress/{job_id}.
    Optional ``max_digs`` (positive integer form field) limits generation to the first N
    dig IDs in MDL order — use ``1`` for a single-package smoke test.
    Optional ``skip_pdf`` (``true`` / ``1``) skips Excel→PDF conversion (often slow on Windows COM).
    Optional ``skip_ili`` (``true`` / ``1``) skips ILI workbook parsing — MDL-only packages (empty feature table).
    Optional ``include_debug`` (``true`` / ``1``) adds ``Dig_Package_Debug.json`` to the ZIP (column map + per-dig MDL values).
    When ``skip_ili`` is set, ILI files may be omitted (empty list); ``ili_formats`` is ignored.
    Returns the ZIP directly as a streaming response (no temp file written to disk).
    """
    tpl_label = (
        template_file.filename
        if template_file and template_file.filename
        else "(default bundled template)"
    )
    log_params(logger, "dig-package/generate", {
        "mdl_filename": mdl_file.filename,
        "ili_count": len(ili_files),
        "ili_filenames": [f.filename for f in ili_files],
        "template_filename": tpl_label,
        "revision": revision,
        "ili_formats": ili_formats,
        "job_id": job_id or "(none)",
        "max_digs": max_digs or "(all)",
        "skip_pdf": skip_pdf or "(false)",
        "skip_ili": skip_ili or "(false)",
        "include_debug": include_debug or "(false)",
    })
    # Resolve early so except blocks can always reference it.
    effective_job_id = job_id.strip() if job_id else None
    try:
        validate_excel_file(mdl_file)
        skip_ili_flag = str(skip_ili).strip().lower() in ("1", "true", "yes", "on")
        include_debug_flag = str(include_debug).strip().lower() in ("1", "true", "yes", "on")
        if not skip_ili_flag:
            for ili_file in ili_files:
                validate_excel_file(ili_file)
        elif ili_files:
            for ili_file in ili_files:
                validate_excel_file(ili_file)

        # Register job before reading the body so GET /progress/{job_id} never 404s while multipart uploads.
        if effective_job_id:
            _cleanup_stale_progress_entries()
            _DIG_PACKAGE_PROGRESS[effective_job_id] = {
                "current": 0,
                "total": 0,
                "status": "running",
                "phase": "receiving_upload",
                "message": "Reading uploaded MDL and ILI files (this can take a while for large uploads)…",
                "expires_at": 0,
            }

        mdl_content = await mdl_file.read()
        if template_file and template_file.filename:
            validate_excel_file(template_file)
            template_content = await template_file.read()
        else:
            try:
                template_content = read_default_dig_package_template_bytes()
            except FileNotFoundError as e:
                raise HTTPException(
                    status_code=503,
                    detail=str(e),
                ) from e
        ili_contents = [await f.read() for f in ili_files] if ili_files else []

        if skip_ili_flag:
            # MDL-only: ignore any ili_formats from the client (no ILI files read).
            formats_list: List[str] = []
        else:
            if not ili_files:
                raise HTTPException(
                    status_code=400,
                    detail="Upload at least one ILI file, or enable skip_ili (MDL-only packages).",
                )
            formats_list = (
                [fmt.strip() for fmt in ili_formats.split(",") if fmt.strip()]
                if ili_formats
                else ["Rosen-MFLA"] * len(ili_files)
            )
            if len(formats_list) != len(ili_files):
                raise HTTPException(
                    status_code=400,
                    detail=f"ILI format count ({len(formats_list)}) does not match uploaded ILI file count ({len(ili_files)}).",
                )

        max_digs_int: Optional[int] = None
        if max_digs and str(max_digs).strip():
            try:
                max_digs_int = int(str(max_digs).strip())
            except ValueError as e:
                raise HTTPException(status_code=400, detail="max_digs must be a positive integer") from e
            if max_digs_int < 1:
                raise HTTPException(status_code=400, detail="max_digs must be >= 1")

        skip_pdf_flag = str(skip_pdf).strip().lower() in ("1", "true", "yes", "on")

        logger.info(
            f"[dig-package/generate] {len(ili_contents)} ILI files, formats={formats_list}, "
            f"max_digs={max_digs_int}, skip_pdf={skip_pdf_flag}, skip_ili={skip_ili_flag}, "
            f"include_debug={include_debug_flag}"
        )

        def _progress_callback(
            current: int,
            total: int,
            *,
            phase: str = "",
            message: str = "",
        ) -> None:
            if effective_job_id:
                entry: Dict[str, Any] = {
                    "current": current,
                    "total": total,
                    "status": "running",
                    "expires_at": 0,
                }
                if phase:
                    entry["phase"] = phase
                if message:
                    entry["message"] = message
                _DIG_PACKAGE_PROGRESS[effective_job_id] = entry

        zip_buffer = await asyncio.to_thread(
            lambda: generate_dig_packages(
                mdl_content=mdl_content,
                ili_contents=ili_contents,
                template_content=template_content,
                revision=revision,
                ili_formats=formats_list,
                progress_callback=_progress_callback,
                max_digs=max_digs_int,
                skip_pdf=skip_pdf_flag,
                skip_ili=skip_ili_flag,
                include_debug=include_debug_flag,
            )
        )

        # Mark job done with TTL so a final frontend poll still succeeds.
        if effective_job_id:
            _DIG_PACKAGE_PROGRESS[effective_job_id]["status"] = "done"
            _DIG_PACKAGE_PROGRESS[effective_job_id]["expires_at"] = (
                time.time() + DIG_PACKAGE_PROGRESS_TTL_SECS
            )

        logger.info("[dig-package/generate] Dig packages generated successfully")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Dig_Packages_R{revision}_{timestamp}.zip"
        # Stream BytesIO directly — no temp file written to disk.
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        if effective_job_id:
            _DIG_PACKAGE_PROGRESS[effective_job_id] = {
                "current": 0,
                "total": 0,
                "status": "error",
                "phase": "error",
                "message": "Request failed (HTTP validation or client error).",
                "expires_at": time.time() + DIG_PACKAGE_PROGRESS_TTL_SECS,
            }
        raise
    except Exception as e:
        log_error(logger, "dig-package/generate", e)
        if effective_job_id:
            _DIG_PACKAGE_PROGRESS[effective_job_id] = {
                "current": 0,
                "total": 0,
                "status": "error",
                "phase": "error",
                "message": str(e)[:500],
                "expires_at": time.time() + DIG_PACKAGE_PROGRESS_TTL_SECS,
            }
        raise HTTPException(
            status_code=500,
            detail=(
                f"Error generating dig packages: {str(e)} "
                f"(failure log on server: backend/logs/dig_package_last_failure.log)"
            ),
        )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8000, reload=True) 