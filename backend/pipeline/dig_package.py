"""
Dig Package Generation Module

This module handles the generation of dig package Excel and PDF files from:
- MDL (Master Dig List)
- ILI (In-Line Inspection) data
- Template Excel file

It matches features, populates templates, and generates individual dig packages.
"""

import io
import json
import traceback
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from backend.logging_config import get_logger
from backend.pipeline.dig_package_layout import (
    AnchorNotFoundError,
    OPTIONAL_EXCAVATION_LAYOUT_BLOCKS,
    OPTIONAL_LAYOUT_VALUE_FIELDS,
    _writable_cell,
    load_layout_manifest,
    resolve_excavation_block_top_left,
    resolve_feature_table_data_start,
    resolve_field_cell,
)
from backend.pipeline.ili_parse import ILIParseTimeoutError, parse_ili_file_with_timeout
from backend.pipeline.ili_reader import find_column_names, read_excel_with_detected_header

logger = get_logger("backend.pipeline.dig_package")

# Overwritten on each failed run; share this path when debugging HTTP 500 from dig-package/generate.
DIG_PACKAGE_LAST_FAILURE_LOG = (
    Path(__file__).resolve().parent.parent / "logs" / "dig_package_last_failure.log"
)


def _persist_dig_package_failure_log(lines: List[str]) -> None:
    try:
        DIG_PACKAGE_LAST_FAILURE_LOG.parent.mkdir(parents=True, exist_ok=True)
        DIG_PACKAGE_LAST_FAILURE_LOG.write_text(
            "Last dig package generation failure (overwritten each failed run).\n"
            "Share this file when reporting issues.\n\n"
            + "\n".join(lines)
            + "\n",
            encoding="utf-8",
        )
        logger.error("Wrote failure details to %s", DIG_PACKAGE_LAST_FAILURE_LOG)
    except OSError as err:
        logger.error("Could not write dig package failure log: %s", err)


def _append_failure_line(failure_lines: List[str], where: str, detail: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    failure_lines.append(f"[{ts}] {where}: {detail}")


def _write_dig_package_debug_json(
    temp_path: Path,
    mdl_df: pd.DataFrame,
    mdl_col_map: Dict[str, str],
    dig_ids_all: List[Any],
    dig_ids: List[Any],
    revision: str,
    max_digs: Optional[int],
    skip_pdf: bool,
    skip_ili: bool,
    ili_data_parsed: List[Dict[str, Any]],
) -> None:
    """Write ``Dig_Package_Debug.json`` into the ZIP staging folder (development visibility)."""
    dig_id_col = mdl_col_map.get("dig_id")
    per_dig: List[Dict[str, Any]] = []
    for dig_id in dig_ids:
        mdl_features = _mdl_rows_for_dig_id(mdl_df, dig_id_col, dig_id)
        if mdl_features.empty:
            continue
        row = mdl_features.iloc[0]
        vals: Dict[str, Optional[str]] = {}
        for k, col in mdl_col_map.items():
            if col in row.index:
                v = row[col]
                vals[k] = None if pd.isna(v) else str(v)
        per_dig.append({"dig_id": str(dig_id), "mdl_values_by_key": vals})
    ili_sources = [
        {
            "format": x["format"],
            "rows": len(x["df"]),
            "sheet_columns_mapped": list((x["col_map"] or {}).keys()),
        }
        for x in ili_data_parsed
    ]
    payload = {
        "mdl_col_map": dict(mdl_col_map),
        "dig_ids_in_mdl": [str(x) for x in dig_ids_all],
        "dig_ids_requested": [str(x) for x in dig_ids],
        "revision": revision,
        "max_digs": max_digs,
        "options": {"skip_pdf": skip_pdf, "skip_ili": skip_ili},
        "per_dig_mdl": per_dig,
        "ili_sources_parsed": ili_sources,
    }
    (temp_path / "Dig_Package_Debug.json").write_text(
        json.dumps(payload, indent=2, default=str),
        encoding="utf-8",
    )


# Bundled 2026 Dig Package Template (used when the client does not upload a template).
DEFAULT_DIG_PACKAGE_TEMPLATE_FILENAME = "2026 Dig Package Template.xlsx"


def default_dig_package_template_path() -> Path:
    """`backend/static/templates/dig_package/2026 Dig Package Template.xlsx`."""
    return (
        Path(__file__).resolve().parent.parent
        / "static"
        / "templates"
        / "dig_package"
        / DEFAULT_DIG_PACKAGE_TEMPLATE_FILENAME
    )


def read_default_dig_package_template_bytes() -> bytes:
    """
    Load the default Dig Package Excel template from disk.

    Raises:
        FileNotFoundError: If the file is missing — copy the 2026 template into
            ``backend/static/templates/dig_package/`` or pass an uploaded template.
    """
    path = default_dig_package_template_path()
    if not path.is_file():
        raise FileNotFoundError(
            f"Default dig package template not found at {path}. "
            f"Place {DEFAULT_DIG_PACKAGE_TEMPLATE_FILENAME} in that folder, or upload a template."
        )
    return path.read_bytes()


# Default assessment range used as fallback when GW-count filter cannot be applied.
DEFAULT_ASSESSMENT_RANGE_M = 30.0

# Girth weld feature type keywords (case-insensitive substring match).
GWD_FEATURE_TYPE_KEYWORDS = ["girth weld", "gwd", "gw", "weld"]


# ============================================================================
# Keyword Definitions for Column Matching
# ============================================================================

# MDL columns use a different schema from source ILI files.
MDL_COLUMN_KEYWORDS = {
    # Prefer "Dig ID" (numeric Integrity IDs like 6000); "Dig Name" is the long package label for filenames.
    "dig_id": ["Dig ID", "NEW Dig Name", "Excavation ID"],
    "dig_name": ["Dig Name", "DigName", "Dig Package Name"],
    # "Target Feature ID" in PNG Integrity MDL; also accept plain "Feature ID" / "ID#" variants.
    "feature_id": ["Target Feature ID", "Feature ID", "ILI Feature ID", "ID#", "Feature Identifier", "Target ID"],
    "pipeline_name": ["Pipeline Name", "Pipeline_Name", "PipelineName", "Line Name"],
    # PNG MDL uses "Pipe NPS" (nominal pipe size) where other formats may use OD in mm.
    "pipe_od": ["Pipe NPS", "NPS", "Pipe OD", "Pipe_OD", "PipeOD", "OD (mm)", "Pipe Diameter"],
    "pipe_nwt": ["Pipe NWT", "Pipe_NWT", "PipeNWT", "Nominal Wall Thickness (mm)", "Wall Thickness"],
    "mop": ["MOP", "MAOP", "Maximum Operating Pressure"],
    "sep": ["SEP", "Safe Excavation Pressure"],
    # PNG MDL uses "TGW Lat (deg)" / "TGW Long (deg)" — "Lat" and "Lon" substring still matches.
    "latitude": ["TGW Lat", "Latitude", "Lat"],
    "longitude": ["TGW Long", "Longitude", "Lon", "Long"],
    "milepost": ["Milepost", "MP", "Mile Post"],
    "pipe_year": ["Pipe Year", "Year", "Installation Year"],
    "pipe_grade": ["Pipe Grade", "Grade", "Material Grade"],
    # PNG MDL uses "Originating ILI" for the ILI vendor / run identifier.
    "ili_run_name": ["Originating ILI", "ILI Run Name", "ILI Run", "Run Name", "ILI Source", "ILI Vendor"],
    # "ILI Time" in PNG MDL is the ILI run date, used as accuracy/timing reference.
    "ili_run_accuracy": ["ILI Time", "ILI Run Date", "ILI Run Accuracy", "Run Accuracy", "ILI Accuracy"],
    "upstream_agm": ["Upstream AGM", "US AGM", "US_AGM"],
    "downstream_agm": ["Downstream AGM", "DS AGM", "DS_AGM"],
    # PNG MDL uses "Total Assessment Length (m)" — "Assessment Length" substring still matches.
    "assessment_length": ["Total Assessment Length", "Assessment Length", "Assess Length"],
    # PNG MDL: "Start Assessment to TGW (m)" / "End Assessment to TGW (m)"
    "start_assessment": ["Start Assessment", "Assessment Start", "US Assessment"],
    "end_assessment": ["End Assessment", "Assessment End", "DS Assessment"],
    "exposure_length": ["Exposure Length", "Expose Length"],
    "start_exposure": ["Start Exposure", "Exposure Start", "US Exposure"],
    "end_exposure": ["End Exposure", "Exposure End", "DS Exposure"],
    "target_girth_weld": ["Target Girth Weld (TGW)", "Target Girth Weld", "Target GW", "TGW", "Target Joint"],
    "length": ["Feature Length", "Length", "Length (mm)", "Length (in)"],
    "width": ["Feature Width", "Width", "Width (mm)", "Width (in)"],
    # Revision as stored in MDL (e.g. 0 for initial issuance). Used as fallback when
    # the user does not supply a revision override.
    "dig_revision": ["Dig Package Revision", "Revision", "Rev"],
    # Additional MDL columns read for display/context but not written to template cells.
    "feature_type": ["Feature Type", "Anomaly Type"],
    "anomaly_description": ["Anomaly Description", "Feature Description", "Description"],
    "class_location": ["Class Location", "Class"],
    "required_completion_date": ["Required Completion Date", "Completion Date"],
    # PNG / Integrity MDL — populates template Issue Date when present (else =TODAY()).
    "issue_date": [
        "Dig Issuance / Revision / Cancellation Date",
        "Dig Issuance",
        "Issuance / Revision / Cancellation Date",
        "Dig Issuance Date",
        "Package Issuance Date",
        "Issuance Date",
    ],
}

# Worksheet name keywords (MDL)
MDL_WORKSHEET_KEYWORDS = [
    "Dig Notification Log",
    "Features&Dig",
    "Features",
    "Dig",
]


# ============================================================================
# Utility Functions
# ============================================================================


def get_cell_from_named_range(workbook, range_name: str):
    """
    Get cell from Excel named range.
    
    Args:
        workbook: openpyxl workbook object
        range_name: Name of the defined range
        
    Returns:
        Cell object if found, None otherwise
    """
    try:
        if range_name not in workbook.defined_names:
            return None
        
        named_range = workbook.defined_names[range_name]
        destinations = list(named_range.destinations)
        
        if not destinations:
            return None
        
        sheet_name, cell_address = destinations[0]
        sheet = workbook[sheet_name]
        return sheet[cell_address]
    except Exception as e:
        logger.debug(f"Error getting named range '{range_name}': {e}")
        return None


def is_valid_dig_id(dig_id: Union[str, int, float, None]) -> bool:
    """
    Accept Integrity-style numeric dig IDs (e.g. 6000) or legacy IDs containing 'GW'.
    """
    if dig_id is None or (isinstance(dig_id, float) and pd.isna(dig_id)):
        return False
    s = str(dig_id).strip()
    if not s or s.lower() in {"-", "nan", "none"}:
        return False
    if "GW" in s.upper():
        return True
    # PNG Integrity program: short numeric Dig ID (column "Dig ID")
    try:
        n = float(s.replace(",", ""))
        if n.is_integer() and 1000 <= abs(n) <= 999999:
            return True
    except (ValueError, TypeError, AttributeError):
        pass
    return False


def _sanitize_package_filename_base(name: str) -> str:
    """Safe single segment for Excel/PDF output stem (PNG-style dig package names)."""
    out = "".join(c if c not in '<>:"/\\|?*' else "_" for c in str(name).strip())
    return out.strip(" ._") or "dig_package"


def _mdl_rows_for_dig_id(mdl_df: pd.DataFrame, dig_id_col: str, dig_id: Any) -> pd.DataFrame:
    """Match MDL rows where Dig ID may be int, float, or string (e.g. 6000 vs 6000.0)."""
    col = mdl_df[dig_id_col]
    tnum = pd.to_numeric(pd.Series([dig_id]), errors="coerce").iloc[0]
    cnum = pd.to_numeric(col, errors="coerce")
    if pd.notna(tnum):
        mask = cnum == tnum
        if bool(mask.any()):
            return mdl_df[mask]
    ds = col.astype(str).str.strip()
    return mdl_df[ds == str(dig_id).strip()]


def _parse_single_ili_file(
    content: bytes,
    v_format: str,
    file_index: int,
    record_failure: Callable[[str, str], None],
) -> Optional[Dict[str, Any]]:
    """Parse one ILI file; return dict or None on failure (timeout or parse error)."""
    try:
        df, col_map, sheet = parse_ili_file_with_timeout(content, v_format)
        logger.info(f"ILI file {file_index + 1} ({v_format}): sheet={sheet}, shape={df.shape}")
        return {"df": df, "col_map": col_map, "format": v_format}
    except ILIParseTimeoutError as e:
        logger.error(
            f"ILI file {file_index + 1} ({v_format}) timed out — excluded from all dig packages: {e}"
        )
        record_failure("parse_ili", f"file {file_index + 1} ({v_format}) TIMEOUT: {e}")
        return None
    except Exception as e:
        logger.error(
            f"Error parsing ILI file {file_index + 1} ({v_format}): {type(e).__name__}: {e} — "
            "this source will be excluded from all dig packages"
        )
        record_failure("parse_ili", f"file {file_index + 1} ({v_format}): {type(e).__name__}: {e}")
        return None


def package_output_stem(mdl_row: pd.Series, mdl_col_map: Dict[str, str], dig_id: Any) -> str:
    """Excel/PDF base name: prefer **Dig Name** (e.g. ID6000_R1R2_..._ML), else dig id."""
    name_col = mdl_col_map.get("dig_name")
    if name_col and name_col in mdl_row.index:
        val = mdl_row[name_col]
        if pd.notna(val) and str(val).strip() not in ("", "-"):
            return _sanitize_package_filename_base(str(val).strip())
    return _sanitize_package_filename_base(dig_id)


def _normalize_match_value(value: Any) -> str:
    """Normalize IDs and labels so matching is resilient to whitespace and numeric formatting."""
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)) if float(value).is_integer() else f"{float(value):.6f}".rstrip("0").rstrip(".")
    text = str(value).strip()
    try:
        numeric_value = float(text.replace(",", ""))
        return str(int(numeric_value)) if numeric_value.is_integer() else f"{numeric_value:.6f}".rstrip("0").rstrip(".")
    except ValueError:
        return " ".join(text.upper().split())


def _split_multivalue_cell(value: Any) -> List[str]:
    """
    Split a MDL cell that may contain multiple values separated by newlines.

    PNG Integrity MDL encodes multiple target features per dig in a single cell,
    one entry per line — e.g. "2982520000000\\n26042200000000\\n40000167".
    Returns a deduplicated list of stripped, non-empty, non-dash values.
    Single-value cells return a one-element list.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    raw = str(value)
    parts = [p.strip() for p in raw.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    seen: set = set()
    result = []
    for p in parts:
        if p and p not in ("-", "N/A", "n/a") and p not in seen:
            seen.add(p)
            result.append(p)
    return result


def _coerce_numeric(value: Any) -> Optional[float]:
    numeric = pd.to_numeric(value, errors="coerce")
    return float(numeric) if pd.notna(numeric) else None


def _dimensions_match(mdl_value: Any, ili_value: Any) -> bool:
    """
    Match numeric dimensions directly or via mm/in conversion.
    This catches common MDL-vs-vendor unit differences without overfitting.
    """
    mdl_num = _coerce_numeric(mdl_value)
    ili_num = _coerce_numeric(ili_value)
    if mdl_num is None or ili_num is None:
        return False

    direct_match = abs(mdl_num - ili_num) <= max(0.01, 0.01 * max(abs(mdl_num), abs(ili_num)))
    if direct_match:
        return True

    mm_to_in = mdl_num / 25.4
    in_to_mm = mdl_num * 25.4
    # Inch-scale compare (mdl mm vs ili in); do not inflate tolerance with in_to_mm (mm×25.4) or
    # unrelated values match (e.g. 12.7 mm vs 0.75 in).
    inch_tol = max(0.02, 0.02 * max(abs(ili_num), abs(mm_to_in)))
    if abs(mm_to_in - ili_num) <= inch_tol:
        return True
    mm_tol = max(0.01, 0.01 * max(abs(ili_num), abs(in_to_mm)))
    return abs(in_to_mm - ili_num) <= mm_tol


def _get_mdl_value(row: pd.Series, col_map: Dict[str, str], col_name: str) -> Any:
    """
    Safely retrieve a value from an MDL row using the column mapping.
    Returns '-' when the column is unmapped or the cell is NaN/empty.
    Used by all populate_* functions to avoid duplicating this lookup pattern.
    """
    col = col_map.get(col_name)
    if col and col in row.index:
        val = row[col]
        return val if pd.notna(val) else "-"
    return "-"


def _first_line_mdl_value(row: pd.Series, col_map: Dict[str, str], col_name: str) -> Any:
    """Like :func:`_get_mdl_value` but for multi-line MDL cells, use the first non-empty line."""
    v = _get_mdl_value(row, col_map, col_name)
    if not isinstance(v, str) or v in ("-", ""):
        return v
    if "\n" not in v:
        return v.strip()
    for line in v.splitlines():
        t = line.strip()
        if t:
            return t
    return "-"


def _issue_date_cell_value(mdl_row: pd.Series, mdl_col_map: Dict[str, str]) -> Optional[Any]:
    """
    MDL-driven Issue Date for the template cell: column ``issue_date`` (e.g. Dig Issuance / …).
    Returns ``None`` when unmapped or empty — caller may use ``=TODAY()`` instead.
    """
    col = mdl_col_map.get("issue_date")
    if not col or col not in mdl_row.index:
        return None
    raw = mdl_row[col]
    if raw is None or pd.isna(raw):
        return None
    ts = pd.to_datetime(raw, errors="coerce")
    if not pd.isna(ts):
        return ts.to_pydatetime().replace(tzinfo=None)
    s = str(raw).strip()
    return s if s else None


def _is_girth_weld_row(feature_type_val: Any) -> bool:
    """Return True when a feature type string looks like a girth weld entry."""
    if feature_type_val is None:
        return False
    text = str(feature_type_val).lower().strip()
    return any(kw in text for kw in GWD_FEATURE_TYPE_KEYWORDS)


# ============================================================================
# MDL Parsing Functions
# ============================================================================

def parse_mdl_file(file_content: bytes) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Parse MDL (Master Dig List) Excel file.
    
    Args:
        file_content: Bytes content of Excel file
        
    Returns:
        Tuple of (DataFrame with mapped columns, column mapping dict)
    """
    df, column_mapping, sheet_name, header_row = read_excel_with_detected_header(
        file_content=file_content,
        keyword_map=MDL_COLUMN_KEYWORDS,
        sheet_keywords=MDL_WORKSHEET_KEYWORDS,
        min_matches=3,
    )
    logger.info(
        f"parse_mdl_file: sheet='{sheet_name}', header_row={header_row}, mapped={list(column_mapping.keys())}"
    )
    return df, column_mapping


def extract_dig_ids(mdl_df: pd.DataFrame, column_mapping: Dict[str, str]) -> List[str]:
    """
    Extract unique valid Dig IDs from MDL.
    
    Args:
        mdl_df: MDL DataFrame
        column_mapping: Column name mapping
        
    Returns:
        List of unique valid Dig IDs
    """
    dig_id_col = column_mapping.get("dig_id")
    if not dig_id_col:
        return []
    
    dig_ids = mdl_df[dig_id_col].dropna().unique().tolist()
    valid_dig_ids = [dig_id for dig_id in dig_ids if is_valid_dig_id(dig_id)]
    
    return sorted(valid_dig_ids)


# ============================================================================
# Feature Matching Functions
# ============================================================================

def match_features_by_id(mdl_feature_id: Any, ili_df: pd.DataFrame, ili_col_map: Dict[str, str]) -> pd.DataFrame:
    """
    Match features by Feature ID.
    
    Args:
        mdl_feature_id: Feature ID from MDL
        ili_df: ILI DataFrame
        ili_col_map: ILI column mapping
        
    Returns:
        Matched rows from ILI DataFrame
    """
    feat_id_col = ili_col_map.get("feature_id")
    if not feat_id_col:
        return pd.DataFrame()

    mdl_key = _normalize_match_value(mdl_feature_id)
    matched = ili_df[ili_df[feat_id_col].apply(_normalize_match_value) == mdl_key]
    return matched


def match_features_by_dimensions(mdl_length: float, mdl_width: float, 
                                 ili_df: pd.DataFrame, ili_col_map: Dict[str, str]) -> pd.DataFrame:
    """
    Match features by dimensions (length and width).
    
    Args:
        mdl_length: Feature length from MDL
        mdl_width: Feature width from MDL
        ili_df: ILI DataFrame
        ili_col_map: ILI column mapping
        
    Returns:
        Matched rows from ILI DataFrame
    """
    length_col = ili_col_map.get("length")
    width_col = ili_col_map.get("width")
    
    if not length_col or not width_col:
        return pd.DataFrame()
    
    if pd.isna(mdl_length) or pd.isna(mdl_width):
        return pd.DataFrame()

    matched = ili_df[
        ili_df[length_col].apply(lambda x: _dimensions_match(mdl_length, x)) &
        ili_df[width_col].apply(lambda x: _dimensions_match(mdl_width, x))
    ]
    return matched


def get_target_feature_indices(mdl_features: pd.DataFrame, ili_df: pd.DataFrame,
                               mdl_col_map: Dict[str, str], ili_col_map: Dict[str, str]) -> List[int]:
    """
    Get indices of target features in ILI DataFrame.

    Pre-builds a normalized-ID lookup dict from the ILI feature_id column so
    that each MDL row lookup is O(1) rather than O(N) over the ILI frame.

    Args:
        mdl_features: MDL features for current dig ID
        ili_df: ILI DataFrame
        mdl_col_map: MDL column mapping
        ili_col_map: ILI column mapping

    Returns:
        List of ILI DataFrame indices that are target features
    """
    target_indices: List[int] = []

    feat_id_col_mdl = mdl_col_map.get("feature_id")
    feat_id_col_ili = ili_col_map.get("feature_id")
    length_col_mdl = mdl_col_map.get("length")
    width_col_mdl = mdl_col_map.get("width")

    # Build lookup: normalized_id → list of ILI row indices (O(N) once).
    ili_id_lookup: Dict[str, List[int]] = {}
    if feat_id_col_ili and feat_id_col_ili in ili_df.columns:
        for ili_idx, val in ili_df[feat_id_col_ili].items():
            key = _normalize_match_value(val)
            if key:
                ili_id_lookup.setdefault(key, []).append(int(ili_idx))

    for _, mdl_row in mdl_features.iterrows():
        # Try Feature ID matching first using the pre-built lookup.
        # PNG Integrity MDL stores multiple feature IDs as newline-delimited strings in
        # a single cell — split before matching so each ID is looked up individually.
        id_matched = False
        if feat_id_col_mdl and feat_id_col_mdl in mdl_row.index:
            raw_feat_id = mdl_row[feat_id_col_mdl]
            individual_ids = _split_multivalue_cell(raw_feat_id)
            for feat_id in individual_ids:
                key = _normalize_match_value(feat_id)
                if key and key in ili_id_lookup:
                    target_indices.extend(ili_id_lookup[key])
                    id_matched = True

        if id_matched:
            continue

        # Fallback to dimension matching when no IDs resolved.
        # MDL length/width cells may also be multi-value; use the first numeric entry.
        if length_col_mdl and width_col_mdl:
            raw_length = mdl_row.get(length_col_mdl)
            raw_width = mdl_row.get(width_col_mdl)
            # Take the first numeric value from potentially multi-line cells.
            length_parts = _split_multivalue_cell(raw_length)
            width_parts = _split_multivalue_cell(raw_width)
            for lp, wp in zip(length_parts, width_parts):
                try:
                    length = float(str(lp).replace(",", ""))
                    width_str = str(wp).strip()
                    if width_str.upper() == "N/A":
                        continue
                    width = float(width_str.replace(",", ""))
                    matched = match_features_by_dimensions(length, width, ili_df, ili_col_map)
                    if not matched.empty:
                        target_indices.extend(matched.index.tolist())
                except (ValueError, TypeError):
                    continue

    return list(set(target_indices))


# ============================================================================
# Template Population Functions
# ============================================================================

def populate_single_value_fields(
    wb,
    mdl_row: pd.Series,
    mdl_col_map: Dict[str, str],
    revision: str,
    excavation_num: int,
    layout_manifest: Dict[str, Any],
    *,
    failure_lines: Optional[List[str]] = None,
    dig_context: str = "",
):
    """
    Populate single-value fields using anchor + offset from ``layout_manifest`` (see
    ``dig_package_layout.json``). Raises ``LayoutError`` / ``AnchorNotFoundError`` if an anchor is missing,
    except optional fields (``mop``, ``sep``, ``milepost``, ILI run/AGM, ``excavation_count``,
    ``issue_date``, etc.): logged and skipped if the template has no matching label.
    When the Issue Date anchor exists, the cell is filled from MDL ``issue_date`` (e.g. column
    *Dig Issuance / Revision / Cancellation Date*) if non-empty; otherwise ``=TODAY()``.
    """
    def _dig_display() -> str:
        """PNG-style packages label the dig using Dig Name (e.g. ID6000_…_ML), not the numeric Dig ID."""
        name_col = mdl_col_map.get("dig_name")
        if name_col and name_col in mdl_row.index:
            val = mdl_row[name_col]
            if pd.notna(val) and str(val).strip() not in ("", "-"):
                return str(val).strip()
        return _get_mdl_value(mdl_row, mdl_col_map, "dig_id")

    mdl_revision = _get_mdl_value(mdl_row, mdl_col_map, "dig_revision")
    effective_revision = str(mdl_revision) if mdl_revision not in ("-", None, "") else str(revision)

    field_values = {
        "dig_display": _dig_display(),
        "revision": effective_revision,
        "pipeline_name": _get_mdl_value(mdl_row, mdl_col_map, "pipeline_name"),
        "target_girth_weld": _first_line_mdl_value(mdl_row, mdl_col_map, "target_girth_weld"),
        "pipe_od": _get_mdl_value(mdl_row, mdl_col_map, "pipe_od"),
        "pipe_nwt": _get_mdl_value(mdl_row, mdl_col_map, "pipe_nwt"),
        "mop": _get_mdl_value(mdl_row, mdl_col_map, "mop"),
        "sep": _get_mdl_value(mdl_row, mdl_col_map, "sep"),
        "latitude": _get_mdl_value(mdl_row, mdl_col_map, "latitude"),
        "longitude": _get_mdl_value(mdl_row, mdl_col_map, "longitude"),
        "milepost": _get_mdl_value(mdl_row, mdl_col_map, "milepost"),
        "pipe_year": _get_mdl_value(mdl_row, mdl_col_map, "pipe_year"),
        "pipe_grade": _get_mdl_value(mdl_row, mdl_col_map, "pipe_grade"),
        "ili_run_name": _get_mdl_value(mdl_row, mdl_col_map, "ili_run_name"),
        "ili_run_accuracy": _get_mdl_value(mdl_row, mdl_col_map, "ili_run_accuracy"),
        "upstream_agm": _get_mdl_value(mdl_row, mdl_col_map, "upstream_agm"),
        "downstream_agm": _get_mdl_value(mdl_row, mdl_col_map, "downstream_agm"),
        "excavation_count": excavation_num,
    }

    ctx = dig_context.strip() or f"excavation #{excavation_num}"

    for field_id, value in field_values.items():
        try:
            cell = resolve_field_cell(wb, layout_manifest, field_id, mdl_col_map=mdl_col_map)
            cell.value = value
        except AnchorNotFoundError as e:
            if field_id in OPTIONAL_LAYOUT_VALUE_FIELDS:
                logger.warning(
                    "Dig Package template has no anchor for %r; skipping cell (%s)",
                    field_id,
                    e,
                )
                continue
            if failure_lines is not None:
                _append_failure_line(
                    failure_lines,
                    f"populate_single_value_fields ({ctx})",
                    f"field {field_id!r}: {e}",
                )
            raise

    try:
        issue_cell = resolve_field_cell(wb, layout_manifest, "issue_date", mdl_col_map=mdl_col_map)
        mdl_issue = _issue_date_cell_value(mdl_row, mdl_col_map)
        issue_cell.value = mdl_issue if mdl_issue is not None else "=TODAY()"
    except AnchorNotFoundError as e:
        logger.warning(
            "Dig Package template has no anchor for 'issue_date'; skipping (%s)",
            e,
        )


def populate_excavation_summary(
    wb,
    mdl_row: pd.Series,
    mdl_col_map: Dict[str, str],
    excavation_num: int,
    layout_manifest: Dict[str, Any],
    *,
    failure_lines: Optional[List[str]] = None,
):
    """Populate excavation and exposure blocks using layout manifest row_deltas.

    If a block anchor is missing, the block is skipped. Optional blocks (see
    ``OPTIONAL_EXCAVATION_LAYOUT_BLOCKS``) do not append to ``failure_lines``.
    """

    def _write_block(block_key: str) -> None:
        block = layout_manifest.get(block_key) or {}
        rows_spec = block.get("rows") or []
        try:
            ws, ar, vc = resolve_excavation_block_top_left(wb, layout_manifest, block_key)
        except AnchorNotFoundError as e:
            logger.warning(
                "Dig Package template has no anchor for layout block %r; skipping (%s)",
                block_key,
                e,
            )
            if (
                failure_lines is not None
                and block_key not in OPTIONAL_EXCAVATION_LAYOUT_BLOCKS
            ):
                _append_failure_line(
                    failure_lines,
                    "populate_excavation_summary",
                    f"block {block_key!r} skipped (no anchor): {e}",
                )
            return
        for spec in rows_spec:
            rd = int(spec.get("row_delta", 0))
            kind = spec.get("kind")
            mdl_k = spec.get("mdl_key")
            if kind == "excavation_label":
                val = f"Excavation #{excavation_num}"
            elif mdl_k:
                val = _get_mdl_value(mdl_row, mdl_col_map, mdl_k)
            else:
                val = "-"
            _writable_cell(ws, ar + rd, vc).value = val

    _write_block("excavation_summary")
    _write_block("exposure_summary")


def populate_feature_table(
    wb,
    ili_datasets: List[Dict[str, Any]],
    excavation_num: int,
    layout_manifest: Dict[str, Any],
    mdl_col_map: Optional[Dict[str, str]] = None,
):
    """
    Populate feature tables for multiple ILI datasets with dynamic rows.
    First data row = anchor row for \"Feature ID\" + data_start_row_offset from manifest.

    When ``ili_datasets`` is empty, writes a single placeholder row so MDL-only packages still
    show that ILI merge is pending or yielded no rows in range.
    """
    try:
        ws, current_row = resolve_feature_table_data_start(wb, layout_manifest, mdl_col_map=mdl_col_map)
    except Exception as e:
        logger.error(f"Feature table layout not resolved: {e}")
        return

    if not ili_datasets:
        try:
            ws.insert_rows(current_row)
            note = _writable_cell(ws, current_row, 1)
            note.value = (
                "ILI anomaly listing: no vendor data merged yet, or no rows in assessment range "
                "(MDL description fields above are populated from the dig notification)."
            )
            note.font = Font(italic=True, color="666666")
            # Do not call merge_cells here — the template may already define merges in this
            # row; merging again or writing to MergedCell placeholders raises read-only errors.
        except Exception as e:
            logger.error(f"Feature table empty placeholder failed: {e}")
        return

    # Helper for ILI values
    def get_ili_value(row, col_map, col_name):
        col = col_map.get(col_name)
        if col and col in row.index:
            val = row[col]
            if pd.notna(val):
                if isinstance(val, (int, float)):
                    return val if val >= 0 else 0.0
                return val
        return "-"

    # Loop through each ILI dataset
    for ds_idx, dataset in enumerate(ili_datasets):
        ili_df = dataset["df"]
        ili_col_map = dataset["col_map"]
        target_indices = dataset["target_indices"]
        target_gw_chainage = dataset["target_gw_chainage"]
        vendor_format = dataset["format"]

        # Add header for different datasets if more than one
        if len(ili_datasets) > 1:
            ws.insert_rows(current_row)
            header_cell = _writable_cell(ws, current_row, 1)
            header_cell.value = f"--- ILI DATA SOURCE: {vendor_format} ---"
            header_cell.font = Font(bold=True, size=12, color="0000FF")
            current_row += 1

        # Populate rows for this dataset
        for idx, (ili_idx, ili_row) in enumerate(ili_df.iterrows()):
            # Insert row
            ws.insert_rows(current_row)
            
            # Check if this is a target feature
            is_target = ili_idx in target_indices
            
            # Values
            feat_id = get_ili_value(ili_row, ili_col_map, "feature_id")
            chainage = get_ili_value(ili_row, ili_col_map, "distance")
            
            # Calculate distance from TGW
            dist_from_tgw = "-"
            if isinstance(chainage, (int, float)) and isinstance(target_gw_chainage, (int, float)):
                dist_from_tgw = chainage - target_gw_chainage
            
            # Set values (use _writable_cell: template merges make many coords MergedCell)
            _writable_cell(ws, current_row, 1).value = str(feat_id)
            _writable_cell(ws, current_row, 2).value = excavation_num
            _writable_cell(ws, current_row, 3).value = get_ili_value(ili_row, ili_col_map, "feature_type")
            _writable_cell(ws, current_row, 4).value = get_ili_value(ili_row, ili_col_map, "feature_desc")
            _writable_cell(ws, current_row, 5).value = get_ili_value(ili_row, ili_col_map, "depth")
            _writable_cell(ws, current_row, 6).value = get_ili_value(ili_row, ili_col_map, "length")
            _writable_cell(ws, current_row, 7).value = get_ili_value(ili_row, ili_col_map, "width")
            _writable_cell(ws, current_row, 8).value = get_ili_value(ili_row, ili_col_map, "orientation")
            _writable_cell(ws, current_row, 9).value = chainage
            _writable_cell(ws, current_row, 10).value = dist_from_tgw
            
            # Apply formatting
            if is_target:
                for col_idx in range(1, 11):
                    cell = _writable_cell(ws, current_row, col_idx)
                    cell.font = Font(bold=True, color="FF0000")
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            current_row += 1
        
        # Add spacing between datasets
        if ds_idx < len(ili_datasets) - 1:
            ws.insert_rows(current_row)
            current_row += 1


# ============================================================================
# PDF Conversion Functions
# ============================================================================

def convert_excel_to_pdf(excel_path: str, pdf_path: str) -> bool:
    """
    Convert Excel file to PDF.
    
    Args:
        excel_path: Path to Excel file
        pdf_path: Path to output PDF file
        
    Returns:
        True if successful, False otherwise
    """
    excel = None
    workbook = None
    try:
        # Try Windows COM automation (best quality)
        import win32com.client

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(str(Path(excel_path).absolute()))
        workbook.ExportAsFixedFormat(0, str(Path(pdf_path).absolute()))
        return True
    except ImportError:
        logger.warning(f"win32com not available. PDF conversion skipped for {excel_path}")
        return False
    except Exception as e:
        logger.error(f"Error converting Excel to PDF: {type(e).__name__}: {e}")
        return False
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


# ============================================================================
# Main Generation Function
# ============================================================================

def get_target_gw_chainage(mdl_row: pd.Series, ili_df: pd.DataFrame, 
                          mdl_col_map: Dict[str, str], ili_col_map: Dict[str, str]) -> Optional[float]:
    """
    Get chainage of Target Girth Weld.
    
    Args:
        mdl_row: MDL row for current dig
        ili_df: ILI DataFrame
        mdl_col_map: MDL column mapping
        ili_col_map: ILI column mapping
        
    Returns:
        Chainage value or None if not found
    """
    # Get TGW ID from MDL
    tgw_col = mdl_col_map.get("target_girth_weld")
    if not tgw_col or tgw_col not in mdl_row.index:
        return None
        
    tgw_id = mdl_row[tgw_col]
    if pd.isna(tgw_id):
        return None
        
    tgw_id_str = _normalize_match_value(tgw_id)

    # Columns to search in ILI
    search_columns = [
        ili_col_map.get("joint_number"),
        ili_col_map.get("feature_id"),
        ili_col_map.get("feature_desc"),
        find_column_names(ili_df, ["Girth Weld", "Girth Weld No", "GWD", "GW"]),
    ]

    chainage_col = ili_col_map.get("distance")
    if not chainage_col or chainage_col not in ili_df.columns:
        return None

    for col_name in search_columns:
        if not col_name or col_name not in ili_df.columns:
            continue

        matches = ili_df[ili_df[col_name].apply(_normalize_match_value) == tgw_id_str]
        if not matches.empty:
            chainage = _coerce_numeric(matches.iloc[0][chainage_col])
            if chainage is not None:
                return chainage

    return None


def filter_ili_by_gw_count(
    ili_df: pd.DataFrame,
    target_gw_chainage: float,
    ili_col_map: Dict[str, str],
    n_each_side: int = 3,
) -> pd.DataFrame:
    """
    Filter ILI data to N girth welds upstream and downstream of the Target Girth Weld.

    Walks the sorted list of girth-weld rows (identified by feature_type keywords),
    finds the TGW position, and uses the chainage of the N-th GWD on each side as
    the range boundary. Clamps to however many GWDs exist if fewer than N are
    available on either side, and logs a warning in that case.

    Falls back to DEFAULT_ASSESSMENT_RANGE_M on each side when:
    - the chainage column is absent, or
    - no girth-weld rows at all are found in the ILI data.

    Args:
        ili_df: Full ILI DataFrame for one vendor source.
        target_gw_chainage: Chainage of the Target Girth Weld (from get_target_gw_chainage).
        ili_col_map: ILI column name mapping.
        n_each_side: Number of GWDs to include on each side (default 3).

    Returns:
        Filtered copy of ili_df.
    """
    chainage_col = ili_col_map.get("distance")
    if not chainage_col or chainage_col not in ili_df.columns:
        logger.warning("filter_ili_by_gw_count: chainage column missing, returning full dataset")
        return ili_df

    chainage_numeric = pd.to_numeric(ili_df[chainage_col], errors="coerce")

    # Identify girth-weld rows using feature_type column.
    feature_type_col = ili_col_map.get("feature_type")
    gwd_chainages: List[float] = []

    if feature_type_col and feature_type_col in ili_df.columns:
        gwd_mask = ili_df[feature_type_col].apply(_is_girth_weld_row)
        gwd_chainages_raw = chainage_numeric[gwd_mask].dropna().sort_values().tolist()
        gwd_chainages = [float(c) for c in gwd_chainages_raw]

    if not gwd_chainages:
        logger.warning(
            "filter_ili_by_gw_count: no girth-weld rows found, "
            f"falling back to ±{DEFAULT_ASSESSMENT_RANGE_M}m window"
        )
        min_ch = target_gw_chainage - DEFAULT_ASSESSMENT_RANGE_M
        max_ch = target_gw_chainage + DEFAULT_ASSESSMENT_RANGE_M
        mask = (chainage_numeric >= min_ch) & (chainage_numeric <= max_ch)
        return ili_df[mask].copy()

    # Find TGW position: closest GWD chainage to the target.
    tgw_idx = min(range(len(gwd_chainages)), key=lambda i: abs(gwd_chainages[i] - target_gw_chainage))

    upstream_gwds = gwd_chainages[:tgw_idx]   # sorted ascending, all < TGW
    downstream_gwds = gwd_chainages[tgw_idx + 1:]  # all > TGW

    # Take up to n_each_side from each direction; clamp and warn if fewer available.
    if len(upstream_gwds) < n_each_side:
        logger.warning(
            f"filter_ili_by_gw_count: only {len(upstream_gwds)} GWDs upstream "
            f"(requested {n_each_side}); clamping to available"
        )
    if len(downstream_gwds) < n_each_side:
        logger.warning(
            f"filter_ili_by_gw_count: only {len(downstream_gwds)} GWDs downstream "
            f"(requested {n_each_side}); clamping to available"
        )

    n_up = min(n_each_side, len(upstream_gwds)) if upstream_gwds else 0
    n_dn = min(n_each_side, len(downstream_gwds)) if downstream_gwds else 0
    us_boundary = upstream_gwds[-n_up] if n_up else (target_gw_chainage - DEFAULT_ASSESSMENT_RANGE_M)
    ds_boundary = downstream_gwds[n_dn - 1] if n_dn else (target_gw_chainage + DEFAULT_ASSESSMENT_RANGE_M)

    # Add a small buffer beyond the boundary GWDs so those GWD rows are included.
    buffer = 0.5
    min_ch = us_boundary - buffer
    max_ch = ds_boundary + buffer

    mask = (chainage_numeric >= min_ch) & (chainage_numeric <= max_ch)
    filtered = ili_df[mask].copy()
    logger.info(
        f"filter_ili_by_gw_count: TGW@{target_gw_chainage:.1f}m, "
        f"range [{min_ch:.1f}, {max_ch:.1f}]m, {len(filtered)} rows retained"
    )
    return filtered


# Interleave ILI parsing with per-dig Excel writes when dig count is modest (MDL-first, then each ILI).
# Large programs keep the batch path: parse all ILI files first, then all digs (fewer workbook rebuilds).
INTERLEAVED_ILI_MERGE_MAX_DIGS = 32


def _collect_ili_datasets_for_dig(
    mdl_first_row: pd.Series,
    mdl_features: pd.DataFrame,
    mdl_col_map: Dict[str, str],
    ili_data_parsed: List[Dict[str, Any]],
    *,
    dig_id: Any = None,
) -> List[Dict[str, Any]]:
    """Build ILI dataset list for one dig from already-parsed ILI sources (possibly a subset)."""
    ili_datasets_for_dig: List[Dict[str, Any]] = []
    for ili_item in ili_data_parsed:
        df = ili_item["df"]
        col_map = ili_item["col_map"]
        v_format = ili_item["format"]

        target_gw_chainage = get_target_gw_chainage(mdl_first_row, df, mdl_col_map, col_map)

        if target_gw_chainage is not None:
            df_filtered = filter_ili_by_gw_count(df, target_gw_chainage, col_map)
        else:
            did = f"{dig_id}" if dig_id is not None else "?"
            logger.warning(
                f"Dig {did} ({v_format}): TGW not found in ILI — "
                f"falling back to ±{DEFAULT_ASSESSMENT_RANGE_M}m window"
            )
            chainage_col = col_map.get("distance")
            if chainage_col and chainage_col in df.columns:
                df_filtered = df.copy()
            else:
                df_filtered = df.copy()
            target_gw_chainage = 0.0

        target_indices = get_target_feature_indices(mdl_features, df_filtered, mdl_col_map, col_map)

        if not df_filtered.empty:
            ili_datasets_for_dig.append({
                "df": df_filtered,
                "col_map": col_map,
                "target_indices": target_indices,
                "target_gw_chainage": target_gw_chainage,
                "format": v_format,
            })

    return ili_datasets_for_dig


def generate_dig_packages(
    mdl_content: bytes,
    ili_contents: List[bytes],
    template_content: bytes,
    revision: str,
    ili_formats: List[str],
    progress_callback: Optional[Callable[..., None]] = None,
    layout_manifest: Optional[Dict[str, Any]] = None,
    max_digs: Optional[int] = None,
    skip_pdf: bool = False,
    skip_ili: bool = False,
    include_debug: bool = False,
):
    """
    Generate dig packages from MDL, multiple ILI datasets, and template.

    Cell positions come from ``dig_package_layout.json`` (anchor + offset), not Excel defined names.

    On failure, details are written to ``backend/logs/dig_package_last_failure.log`` (overwritten each run).

    Args:
        mdl_content: MDL Excel file content
        ili_contents: List of ILI Excel file contents
        template_content: Template Excel file content
        revision: Revision identifier
        ili_formats: List of vendor formats corresponding to ili_contents
        progress_callback: Optional ``callable(current, total, *, phase=..., message=...)``.
            Phases include ``parse_mdl``, ``parse_ili``, ``mdl_template`` (MDL-only Excel fields),
            ``merge_ili`` (refresh after each ILI when dig count ≤ 32), and ``dig_generation``.
            Programs with more than 32 digs use a batch path: all ILI files are parsed first, then digs.
        layout_manifest: Optional pre-loaded layout dict (tests); default loads bundled JSON.
        max_digs: If set, only the first N dig IDs (MDL order) are processed — for smoke tests and dev.
        skip_pdf: If True, skip Excel→PDF conversion (often slow / blocking on Windows COM). Excel only.
        skip_ili: If True, do not read or parse ILI workbooks — MDL-only dig packages (empty feature table).
            Use while developing template/layout without waiting on large ILI files.
        include_debug: If True, add ``Dig_Package_Debug.json`` to the ZIP with ``mdl_col_map`` and per-dig
            MDL values the pipeline resolved.

    Each ILI workbook parse is limited by env ``DIG_PACKAGE_ILI_PARSE_TIMEOUT_SEC`` (default 300 seconds;
    ``0`` disables). On timeout that file is skipped like a parse error.

    Returns:
        BytesIO object containing the ZIP file
    """
    if skip_ili:
        ili_contents = []
        ili_formats = []
    elif len(ili_contents) != len(ili_formats):
        raise ValueError(
            f"ILI file count ({len(ili_contents)}) does not match format count ({len(ili_formats)})."
        )

    failure_lines: List[str] = []

    def record_failure(where: str, detail: str) -> None:
        _append_failure_line(failure_lines, where, detail)
        logger.warning("dig_package %s", failure_lines[-1])

    try:
        return _generate_dig_packages_impl(
            mdl_content,
            ili_contents,
            template_content,
            revision,
            ili_formats,
            progress_callback=progress_callback,
            layout_manifest=layout_manifest,
            max_digs=max_digs,
            skip_pdf=skip_pdf,
            skip_ili=skip_ili,
            include_debug=include_debug,
            failure_lines=failure_lines,
            record_failure=record_failure,
        )
    except Exception as e:
        record_failure("fatal", f"{type(e).__name__}: {e}")
        failure_lines.extend(traceback.format_exc().splitlines())
        _persist_dig_package_failure_log(failure_lines)
        raise


def _generate_dig_packages_impl(
    mdl_content: bytes,
    ili_contents: List[bytes],
    template_content: bytes,
    revision: str,
    ili_formats: List[str],
    progress_callback: Optional[Callable[..., None]] = None,
    layout_manifest: Optional[Dict[str, Any]] = None,
    max_digs: Optional[int] = None,
    skip_pdf: bool = False,
    skip_ili: bool = False,
    include_debug: bool = False,
    *,
    failure_lines: List[str],
    record_failure: Callable[[str, str], None],
):
    """Internal implementation; see :func:`generate_dig_packages`."""

    logger.info(
        f"generate_dig_packages: Parsing MDL, ILI files={len(ili_contents)}, formats={ili_formats}, "
        f"skip_ili={skip_ili}"
    )
    if progress_callback:
        progress_callback(
            0,
            0,
            phase="parse_mdl",
            message="Parsing Master Dig List (MDL)…",
        )
    mdl_df, mdl_col_map = parse_mdl_file(mdl_content)
    logger.debug(f"MDL columns mapped: {list(mdl_col_map.keys())}")

    dig_ids_all = extract_dig_ids(mdl_df, mdl_col_map)
    logger.info(
        f"Extracted {len(dig_ids_all)} dig IDs: {dig_ids_all[:5]}..." if len(dig_ids_all) > 5
        else f"Extracted dig IDs: {dig_ids_all}"
    )
    if not dig_ids_all:
        raise ValueError("No valid Dig IDs found in MDL file")

    dig_ids: List[Any] = list(dig_ids_all)
    if max_digs is not None and max_digs > 0:
        dig_ids = dig_ids[:max_digs]
        logger.info(
            f"max_digs={max_digs}: processing {len(dig_ids)} dig ID(s): {dig_ids} "
            f"(of {len(dig_ids_all)} in MDL)"
        )

    resolved_layout = layout_manifest if layout_manifest is not None else load_layout_manifest()

    # Serialize template once; re-parse from bytes per dig (safe, avoids deepcopy pitfalls).
    template_buf = io.BytesIO()
    _template_wb_for_serialization = load_workbook(io.BytesIO(template_content))
    _template_wb_for_serialization.save(template_buf)
    cached_template_bytes = template_buf.getvalue()
    logger.debug(f"Template serialized to {len(cached_template_bytes)} bytes for per-dig reuse")

    use_interleaved = len(dig_ids) <= INTERLEAVED_ILI_MERGE_MAX_DIGS
    if use_interleaved:
        logger.info(
            f"generate_dig_packages: interleaved ILI merge (MDL-first, refresh after each ILI) — "
            f"{len(dig_ids)} dig(s), {len(ili_contents)} ILI file(s)"
        )
    else:
        logger.info(
            f"generate_dig_packages: batch ILI parse (all ILI then all digs) — "
            f"{len(dig_ids)} dig(s) > interleave threshold {INTERLEAVED_ILI_MERGE_MAX_DIGS}"
        )

    ili_data_parsed: List[Dict[str, Any]] = []
    failed_ili_formats: List[str] = []

    if not use_interleaved and not skip_ili:
        # Large programs: parse every ILI workbook first, then run the dig loop once per dig.
        n_ili = len(ili_contents)
        for i, (content, v_format) in enumerate(zip(ili_contents, ili_formats)):
            if progress_callback and n_ili > 0:
                progress_callback(
                    i,
                    n_ili,
                    phase="parse_ili",
                    message=(
                        f"Parsing ILI file {i + 1}/{n_ili} ({v_format}) — "
                        "large workbooks can take several minutes per file"
                    ),
                )
            parsed = _parse_single_ili_file(content, v_format, i, record_failure)
            if parsed is not None:
                ili_data_parsed.append(parsed)
            else:
                failed_ili_formats.append(v_format)

        if not ili_data_parsed:
            raise ValueError("No ILI files could be parsed successfully")
    elif not use_interleaved and skip_ili:
        logger.info("skip_ili=True: skipping ILI parse (batch path)")

    import tempfile as _tempfile
    with _tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)

        if use_interleaved:
            summary = {
                "revision": revision,
                "dig_ids_in_mdl": [str(d) for d in dig_ids_all],
                "dig_ids_requested": [str(d) for d in dig_ids],
                "max_digs": max_digs,
                "skip_pdf": skip_pdf,
                "skip_ili": skip_ili,
                "include_debug": include_debug,
                "generated": [],
                "skipped": [],
                "ili_files": [],
                "ili_files_failed": [],
            }

            if progress_callback:
                progress_callback(
                    0,
                    0,
                    phase="mdl_template",
                    message=f"Writing MDL-only fields (notification / excavation) for {len(dig_ids)} dig(s)…",
                )
            dig_id_col = mdl_col_map.get("dig_id")
            for excavation_num, dig_id in enumerate(dig_ids, start=1):
                mdl_features = _mdl_rows_for_dig_id(mdl_df, dig_id_col, dig_id)
                if mdl_features.empty:
                    continue
                mdl_first_row = mdl_features.iloc[0]
                wb = load_workbook(io.BytesIO(cached_template_bytes))
                populate_single_value_fields(
                    wb,
                    mdl_first_row,
                    mdl_col_map,
                    revision,
                    excavation_num,
                    resolved_layout,
                    failure_lines=failure_lines,
                    dig_context=f"dig_id={dig_id} (mdl-only)",
                )
                populate_excavation_summary(
                    wb,
                    mdl_first_row,
                    mdl_col_map,
                    excavation_num,
                    resolved_layout,
                    failure_lines=failure_lines,
                )
                populate_feature_table(wb, [], excavation_num, resolved_layout, mdl_col_map=mdl_col_map)
                out_stem = package_output_stem(mdl_first_row, mdl_col_map, dig_id)
                excel_filename = f"{out_stem}_DP_R{revision}.xlsx"
                wb.save(str(temp_path / excel_filename))

            if not skip_ili:
                ili_data_parsed = []
                failed_ili_formats = []
                n_ili = len(ili_contents)
                for i, (content, v_format) in enumerate(zip(ili_contents, ili_formats)):
                    if progress_callback and n_ili > 0:
                        progress_callback(
                            i,
                            n_ili,
                            phase="parse_ili",
                            message=(
                                f"Parsing ILI file {i + 1}/{n_ili} ({v_format}) — "
                                "large workbooks can take several minutes per file"
                            ),
                        )
                    parsed = _parse_single_ili_file(content, v_format, i, record_failure)
                    if parsed is not None:
                        ili_data_parsed.append(parsed)
                    else:
                        failed_ili_formats.append(v_format)

                    if progress_callback and n_ili > 0:
                        progress_callback(
                            i + 1,
                            n_ili,
                            phase="merge_ili",
                            message=(
                                f"Merging ILI {i + 1}/{n_ili} ({v_format}) into dig package(s) — "
                                f"{len(ili_data_parsed)} vendor source(s) loaded"
                            ),
                        )
                    for excavation_num, dig_id in enumerate(dig_ids, start=1):
                        mdl_features = _mdl_rows_for_dig_id(mdl_df, dig_id_col, dig_id)
                        if mdl_features.empty:
                            continue
                        mdl_first_row = mdl_features.iloc[0]
                        ili_datasets_for_dig = _collect_ili_datasets_for_dig(
                            mdl_first_row, mdl_features, mdl_col_map, ili_data_parsed, dig_id=dig_id
                        )
                        wb = load_workbook(io.BytesIO(cached_template_bytes))
                        populate_single_value_fields(
                            wb,
                            mdl_first_row,
                            mdl_col_map,
                            revision,
                            excavation_num,
                            resolved_layout,
                            failure_lines=failure_lines,
                            dig_context=f"dig_id={dig_id}",
                        )
                        populate_excavation_summary(
                            wb,
                            mdl_first_row,
                            mdl_col_map,
                            excavation_num,
                            resolved_layout,
                            failure_lines=failure_lines,
                        )
                        populate_feature_table(
                            wb, ili_datasets_for_dig, excavation_num, resolved_layout, mdl_col_map=mdl_col_map
                        )
                        out_stem = package_output_stem(mdl_first_row, mdl_col_map, dig_id)
                        excel_filename = f"{out_stem}_DP_R{revision}.xlsx"
                        wb.save(str(temp_path / excel_filename))

                if not ili_data_parsed:
                    raise ValueError("No ILI files could be parsed successfully")

                summary["ili_files"] = [
                    {"format": item["format"], "rows": len(item["df"])} for item in ili_data_parsed
                ]
                summary["ili_files_failed"] = failed_ili_formats
            else:
                ili_data_parsed = []
                failed_ili_formats = []
                summary["ili_files"] = []
                summary["ili_files_failed"] = []
                logger.info(
                    "skip_ili=True: interleaved path — keeping MDL-only workbooks, no ILI parse/merge"
                )

            total = len(dig_ids)
            if progress_callback and total > 0:
                progress_callback(
                    0,
                    total,
                    phase="dig_generation",
                    message=f"Finalizing PDF and summary for {total} dig package(s)…",
                )

            for excavation_num, dig_id in enumerate(dig_ids, start=1):
                dig_id_col = mdl_col_map.get("dig_id")
                mdl_features = _mdl_rows_for_dig_id(mdl_df, dig_id_col, dig_id)
                if mdl_features.empty:
                    summary["skipped"].append({
                        "dig_id": str(dig_id),
                        "reason": "No MDL rows matched dig ID after parsing",
                    })
                    record_failure(
                        "skipped_dig",
                        f"dig_id={dig_id}: no MDL rows matched after parsing",
                    )
                    if progress_callback:
                        progress_callback(
                            excavation_num,
                            total,
                            phase="dig_generation",
                            message=f"Skipped dig {excavation_num}/{total} (no MDL rows)",
                        )
                    continue

                mdl_first_row = mdl_features.iloc[0]
                ili_datasets_for_dig = _collect_ili_datasets_for_dig(
                    mdl_first_row, mdl_features, mdl_col_map, ili_data_parsed, dig_id=dig_id
                )
                out_stem = package_output_stem(mdl_first_row, mdl_col_map, dig_id)
                excel_filename = f"{out_stem}_DP_R{revision}.xlsx"
                excel_path = temp_path / excel_filename

                if not ili_datasets_for_dig and not skip_ili:
                    try:
                        excel_path.unlink()
                    except OSError:
                        pass
                    summary["skipped"].append({
                        "dig_id": str(dig_id),
                        "reason": "No ILI rows found in assessment range",
                    })
                    record_failure(
                        "skipped_dig",
                        f"dig_id={dig_id}: no ILI rows in assessment range",
                    )
                    if progress_callback:
                        progress_callback(
                            excavation_num,
                            total,
                            phase="dig_generation",
                            message=f"Skipped dig {excavation_num}/{total} (no ILI in range)",
                        )
                    continue

                if progress_callback:
                    progress_callback(
                        excavation_num,
                        total,
                        phase="dig_generation",
                        message=f"PDF/summary for dig {excavation_num}/{total} ({dig_id})",
                    )

                pdf_generated = False
                pdf_filename = f"{out_stem}_DP_R{revision}.pdf"
                pdf_path = temp_path / pdf_filename
                if not skip_pdf:
                    pdf_generated = convert_excel_to_pdf(str(excel_path), str(pdf_path))
                else:
                    logger.info(f"skip_pdf=True: omitting PDF for {excel_filename}")
                matched_count = (
                    sum(len(ds["target_indices"]) for ds in ili_datasets_for_dig)
                    if ili_datasets_for_dig
                    else 0
                )
                summary["generated"].append({
                    "dig_id": str(dig_id),
                    "dig_name": str(_get_mdl_value(mdl_first_row, mdl_col_map, "dig_name")),
                    "excavation_num": excavation_num,
                    "excel_file": excel_filename,
                    "pdf_file": pdf_filename if pdf_generated else None,
                    "pdf_generated": pdf_generated,
                    "ili_dataset_count": len(ili_datasets_for_dig),
                    "features_matched": matched_count,
                })

        else:
            summary = {
                "revision": revision,
                "dig_ids_in_mdl": [str(d) for d in dig_ids_all],
                "dig_ids_requested": [str(d) for d in dig_ids],
                "max_digs": max_digs,
                "skip_pdf": skip_pdf,
                "skip_ili": skip_ili,
                "include_debug": include_debug,
                "generated": [],
                "skipped": [],
                "ili_files": [{"format": item["format"], "rows": len(item["df"])} for item in ili_data_parsed],
                "ili_files_failed": failed_ili_formats,
            }

            total = len(dig_ids)
            if progress_callback and total > 0:
                progress_callback(
                    0,
                    total,
                    phase="dig_generation",
                    message=f"Building {total} dig package(s)…",
                )

            # Process each dig ID
            for excavation_num, dig_id in enumerate(dig_ids, start=1):
                dig_id_col = mdl_col_map.get("dig_id")
                mdl_features = _mdl_rows_for_dig_id(mdl_df, dig_id_col, dig_id)
                if mdl_features.empty:
                    summary["skipped"].append({"dig_id": str(dig_id), "reason": "No MDL rows matched dig ID after parsing"})
                    record_failure(
                        "skipped_dig",
                        f"dig_id={dig_id}: no MDL rows matched after parsing",
                    )
                    if progress_callback:
                        progress_callback(
                            excavation_num,
                            total,
                            phase="dig_generation",
                            message=f"Skipped dig {excavation_num}/{total} (no MDL rows)",
                        )
                    continue

                mdl_first_row = mdl_features.iloc[0]
                ili_datasets_for_dig = _collect_ili_datasets_for_dig(
                    mdl_first_row, mdl_features, mdl_col_map, ili_data_parsed, dig_id=dig_id
                )

                if not ili_datasets_for_dig and not skip_ili:
                    summary["skipped"].append({"dig_id": str(dig_id), "reason": "No ILI rows found in assessment range"})
                    record_failure(
                        "skipped_dig",
                        f"dig_id={dig_id}: no ILI rows in assessment range",
                    )
                    if progress_callback:
                        progress_callback(
                            excavation_num,
                            total,
                            phase="dig_generation",
                            message=f"Skipped dig {excavation_num}/{total} (no ILI in range)",
                        )
                    continue

                # Load per-dig workbook from cached bytes (safe serialise-once pattern).
                wb = load_workbook(io.BytesIO(cached_template_bytes))
                populate_single_value_fields(
                    wb,
                    mdl_first_row,
                    mdl_col_map,
                    revision,
                    excavation_num,
                    resolved_layout,
                    failure_lines=failure_lines,
                    dig_context=f"dig_id={dig_id}",
                )
                populate_excavation_summary(
                    wb,
                    mdl_first_row,
                    mdl_col_map,
                    excavation_num,
                    resolved_layout,
                    failure_lines=failure_lines,
                )
                populate_feature_table(
                    wb, ili_datasets_for_dig, excavation_num, resolved_layout, mdl_col_map=mdl_col_map
                )

                out_stem = package_output_stem(mdl_first_row, mdl_col_map, dig_id)
                excel_filename = f"{out_stem}_DP_R{revision}.xlsx"
                excel_path = temp_path / excel_filename
                wb.save(str(excel_path))

                if progress_callback:
                    progress_callback(
                        excavation_num,
                        total,
                        phase="dig_generation",
                        message=f"Finished Excel for dig {excavation_num}/{total} ({dig_id})",
                    )

                pdf_generated = False
                pdf_filename = f"{out_stem}_DP_R{revision}.pdf"
                pdf_path = temp_path / pdf_filename
                if not skip_pdf:
                    pdf_generated = convert_excel_to_pdf(str(excel_path), str(pdf_path))
                else:
                    logger.info(f"skip_pdf=True: omitting PDF for {excel_filename}")
                matched_count = (
                    sum(len(ds["target_indices"]) for ds in ili_datasets_for_dig)
                    if ili_datasets_for_dig
                    else 0
                )
                summary["generated"].append({
                    "dig_id": str(dig_id),
                    "dig_name": str(_get_mdl_value(mdl_first_row, mdl_col_map, "dig_name")),
                    "excavation_num": excavation_num,
                    "excel_file": excel_filename,
                    "pdf_file": pdf_filename if pdf_generated else None,
                    "pdf_generated": pdf_generated,
                    "ili_dataset_count": len(ili_datasets_for_dig),
                    "features_matched": matched_count,
                })

        if not summary["generated"]:
            raise ValueError("No dig packages were generated. Check MDL mapping and ILI matching results.")

        fail_log_path = temp_path / "Dig_Package_Generation_Failures.log"
        fail_header = (
            "Non-fatal issues (optional field skips, ILI parse failures, skipped digs).\n"
            "After an HTTP 500, see backend/logs/dig_package_last_failure.log on the server for the traceback.\n"
            "---\n"
        )
        fail_log_path.write_text(
            fail_header
            + ("\n".join(failure_lines) + "\n" if failure_lines else "(none)\n"),
            encoding="utf-8",
        )

        summary_path = temp_path / f"Dig_Package_Generation_Summary_R{revision}.json"
        summary_path.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")

        if include_debug:
            _write_dig_package_debug_json(
                temp_path,
                mdl_df,
                mdl_col_map,
                dig_ids_all,
                dig_ids,
                revision,
                max_digs,
                skip_pdf,
                skip_ili,
                ili_data_parsed,
            )

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for file_path in temp_path.glob("*"):
                if file_path.is_file():
                    zip_file.write(file_path, file_path.name)

        zip_buffer.seek(0)
        file_count = len([f for f in temp_path.glob("*") if f.is_file()])
        logger.info(f"generate_dig_packages: Created ZIP with {file_count} files for {len(dig_ids)} dig IDs")
        return zip_buffer

