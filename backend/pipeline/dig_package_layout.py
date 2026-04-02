"""
Dig package Excel layout: resolve value cells by anchor text + offset (no defined names required).

Manifest: backend/static/templates/dig_package/dig_package_layout.json
Override: set env DIG_PACKAGE_LAYOUT_JSON to an absolute path.
"""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.logging_config import get_logger

logger = get_logger("backend.pipeline.dig_package_layout")

SCAN_MAX_ROW = 300
SCAN_MAX_COL = 40


class LayoutError(Exception):
    """Template layout could not be resolved."""


class AnchorNotFoundError(LayoutError):
    def __init__(self, message: str, *, sheet: str = "", anchor_text: str = ""):
        super().__init__(message)
        self.sheet = sheet
        self.anchor_text = anchor_text


def default_layout_manifest_path() -> Path:
    return (
        Path(__file__).resolve().parent.parent
        / "static"
        / "templates"
        / "dig_package"
        / "dig_package_layout.json"
    )


def load_layout_manifest(path: Optional[Union[Path, str]] = None) -> Dict[str, Any]:
    """Load JSON layout manifest. Env DIG_PACKAGE_LAYOUT_JSON overrides default path."""
    env_path = os.environ.get("DIG_PACKAGE_LAYOUT_JSON", "").strip()
    p = path or (Path(env_path) if env_path else default_layout_manifest_path())
    p = Path(p)
    if not p.is_file():
        raise FileNotFoundError(
            f"Dig package layout manifest not found: {p}. "
            "Add dig_package_layout.json under backend/static/templates/dig_package/ "
            "or set DIG_PACKAGE_LAYOUT_JSON."
        )
    data = json.loads(p.read_text(encoding="utf-8"))
    if not isinstance(data, dict) or "fields" not in data:
        raise LayoutError("Invalid layout manifest: expected a JSON object with 'fields'.")
    return data


def _writable_cell(ws: Worksheet, row: int, column: int) -> Cell:
    """Resolve MergedCell placeholders to the merge anchor (writable Cell)."""
    cell = ws.cell(row=row, column=column)
    if not isinstance(cell, MergedCell):
        return cell  # type: ignore[return-value]
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= column <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)  # type: ignore[return-value]
    return cell  # type: ignore[return-value]


def _anchor_merge_max_col(ws: Worksheet, row: int, col: int) -> int:
    """
    Rightmost column index of the merged range containing (row, col), or ``col`` if unmerged.

    Layout offsets are defined as ``N`` columns to the right of the **label merge** (not the
    anchor cell alone). Otherwise ``anchor_col + 1`` lands inside the same merge as the label
    and :func:`_writable_cell` snaps to the merge top-left, overwriting the label text.
    """
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return int(rng.max_col)
    return col


def _normalize_label(s: str) -> str:
    return " ".join(str(s).strip().split())


def _cell_matches(cell_value: Any, text: str, match: str) -> bool:
    if cell_value is None:
        return False
    sv = str(cell_value).strip()
    if not sv:
        return False
    t = text.strip()
    if match == "equals":
        return _normalize_label(sv).lower() == _normalize_label(t).lower()
    if match == "contains":
        return t.lower() in sv.lower()
    if match == "startswith":
        return sv.lower().startswith(t.lower())
    return False


def get_target_sheet(wb: Workbook, sheet_names: List[str]) -> Worksheet:
    for name in sheet_names:
        if name in wb.sheetnames:
            return wb[name]
    raise LayoutError(
        f"No worksheet found in {sheet_names!r}. Workbook has: {wb.sheetnames!r}"
    )


def find_anchor_cell(
    ws: Worksheet,
    text: str,
    match: str = "equals",
) -> Tuple[int, int]:
    """
    Scan the sheet top-to-bottom, left-to-right; return (row, col) of the first match.
    """
    for row in range(1, SCAN_MAX_ROW + 1):
        for col in range(1, SCAN_MAX_COL + 1):
            cell = ws.cell(row, col)
            if _cell_matches(cell.value, text, match):
                return row, col
    raise AnchorNotFoundError(
        f"Anchor not found: {match!r} {text!r} on sheet {ws.title!r}",
        sheet=ws.title,
        anchor_text=text,
    )


def _dedupe_anchor_candidates(pairs: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
    seen: set = set()
    out: List[Tuple[str, str]] = []
    for text, match in pairs:
        key = (text.strip().lower(), match)
        if key in seen:
            continue
        seen.add(key)
        out.append((text, match))
    return out


# Maps layout manifest ``field_id`` (single-value fields) to MDL logical column keys.
# Values are the dig-notification column *header* strings from ``mdl_col_map`` — often the same
# text as the Dig Package template row labels, so they work as automatic anchor fallbacks.
# Extra label phrases tried after JSON fallbacks (MDL may omit column; template wording varies).
LAYOUT_FIELD_EXTRA_SYNONYMS: Dict[str, List[str]] = {
    "sep": [
        "Safe Excavation Pressure",
        "Safe Excavation",
        "Excavation Pressure",
        "Min Excavation",
    ],
    "mop": [
        "MAOP",
        "Maximum Operating Pressure",
        "Operating Pressure",
        "Max Operating",
    ],
    "pipe_od": [
        "Outside Diameter",
        "OD (mm)",
    ],
    "milepost": [
        "Nearest ROW MP",
        "ROW MP",
        "Station",
        "Chainage",
        "KP",
        "Distance",
        "Mile Post",
        "Location",
    ],
    "target_girth_weld": [
        "Target Girth Weld (TGW)",
        "Target Joint",
        "Target GW",
    ],
    "ili_run_name": [
        "ILI Run",
        "Inspection Run",
        "Tool Run",
        "ILI Tool",
        "Originating",
    ],
    "ili_run_accuracy": [
        "ILI Date",
        "Run Date",
        "Inspection Date",
        "ILI Year",
    ],
    "upstream_agm": [
        "Upstream",
        "US AGM",
    ],
    "downstream_agm": [
        "Downstream",
        "DS AGM",
    ],
    "issue_date": [
        "Date of Issue",
        "Package Issue Date",
    ],
}

# Single-value fields allowed to skip if no cell anchor exists (MDL/template may omit the row).
OPTIONAL_LAYOUT_VALUE_FIELDS = frozenset({
    "mop",
    "sep",
    "milepost",
    "ili_run_name",
    "ili_run_accuracy",
    "upstream_agm",
    "downstream_agm",
    "excavation_count",  # many current templates omit "Number of Excavations"
    "issue_date",  # templates may omit Issue Date / use different wording
})

# Layout blocks: if anchor missing, skip silently (no user-facing failure log).
OPTIONAL_EXCAVATION_LAYOUT_BLOCKS = frozenset({
    "exposure_summary",
})

LAYOUT_FIELD_ID_TO_MDL_KEY: Dict[str, str] = {
    "dig_display": "dig_name",
    "revision": "dig_revision",
    "pipeline_name": "pipeline_name",
    "pipe_od": "pipe_od",
    "pipe_nwt": "pipe_nwt",
    "mop": "mop",
    "sep": "sep",
    "latitude": "latitude",
    "longitude": "longitude",
    "milepost": "milepost",
    "pipe_year": "pipe_year",
    "pipe_grade": "pipe_grade",
    "ili_run_name": "ili_run_name",
    "ili_run_accuracy": "ili_run_accuracy",
    "upstream_agm": "upstream_agm",
    "downstream_agm": "downstream_agm",
    "issue_date": "issue_date",
    "target_girth_weld": "target_girth_weld",
}


def mdl_header_hint_for_layout_field(
    field_id: str,
    mdl_col_map: Optional[Dict[str, str]],
) -> Optional[str]:
    """Return the MDL sheet column header for this layout field, if known."""
    if not mdl_col_map:
        return None
    mdl_key = LAYOUT_FIELD_ID_TO_MDL_KEY.get(field_id)
    if not mdl_key:
        return None
    return mdl_col_map.get(mdl_key)


def _anchor_candidate_tuples(
    spec: Dict[str, Any],
    manifest: Dict[str, Any],
    mdl_header_hint: Optional[str] = None,
    field_id: Optional[str] = None,
) -> List[Tuple[str, str]]:
    """
    Build (text, match) pairs: primary anchor, then MDL column header hint (if any),
    then JSON ``anchor_fallbacks``, then ``LAYOUT_FIELD_EXTRA_SYNONYMS`` for ``field_id``.
    MDL hints use the dig-notification file wording.
    """
    out: List[Tuple[str, str]] = []
    primary = spec.get("anchor") or {}
    t = (primary.get("text") or "").strip()
    m = primary.get("match") or manifest.get("match_default") or "equals"
    if t:
        out.append((t, m))
    hint = (mdl_header_hint or "").strip()
    if hint:
        out.append((hint, "contains"))
    default_m = manifest.get("match_default") or "equals"
    for fb in spec.get("anchor_fallbacks") or []:
        ft = (fb.get("text") or "").strip()
        fm = fb.get("match") or default_m
        if ft:
            out.append((ft, fm))
    if field_id:
        for syn in LAYOUT_FIELD_EXTRA_SYNONYMS.get(field_id, ()):
            st = syn.strip()
            if st:
                out.append((st, "contains"))
    return _dedupe_anchor_candidates(out)


def find_anchor_cell_candidates(
    ws: Worksheet,
    candidates: List[Tuple[str, str]],
    *,
    context: str = "",
) -> Tuple[int, int]:
    """Try each (text, match) until one resolves; templates vary in label wording."""
    if not candidates:
        raise AnchorNotFoundError(
            f"{context}No anchor candidates on sheet {ws.title!r}",
            sheet=ws.title,
            anchor_text="",
        )
    tried: List[str] = []
    for text, match in candidates:
        try:
            return find_anchor_cell(ws, text, match)
        except AnchorNotFoundError:
            tried.append(f"{match!r} {text!r}")
            continue
    raise AnchorNotFoundError(
        f"{context}Anchor not found on sheet {ws.title!r}; tried: {', '.join(tried)}",
        sheet=ws.title,
        anchor_text=candidates[0][0],
    )


def resolve_field_cell(
    wb: Workbook,
    manifest: Dict[str, Any],
    field_id: str,
    mdl_col_map: Optional[Dict[str, str]] = None,
) -> Cell:
    """
    Return the openpyxl Cell where a single-value field should be written.

    If ``mdl_col_map`` is provided (from the parsed MDL), the dig-notification **column header**
    for that field is tried right after the manifest primary anchor — usually matches the
    template label and avoids one-off JSON edits per project.
    """
    fields = manifest.get("fields") or {}
    if field_id not in fields:
        raise LayoutError(f"Unknown layout field_id: {field_id!r}")
    spec = fields[field_id]
    ws = get_target_sheet(wb, manifest.get("sheet_names") or ["Dig Package"])
    hint = mdl_header_hint_for_layout_field(field_id, mdl_col_map)
    candidates = _anchor_candidate_tuples(
        spec, manifest, mdl_header_hint=hint, field_id=field_id
    )
    if not candidates:
        raise LayoutError(f"Layout field {field_id!r} has no anchor text")

    ar, ac = find_anchor_cell_candidates(
        ws, candidates, context=f"field {field_id!r}: "
    )
    off = spec.get("offset") or {"row": 0, "col": 1}
    dr = int(off.get("row", 0))
    dc = int(off.get("col", 1))
    tr = ar + dr
    label_right = _anchor_merge_max_col(ws, ar, ac)
    tc = label_right + dc
    return _writable_cell(ws, tr, tc)


def resolve_excavation_block_top_left(
    wb: Workbook,
    manifest: Dict[str, Any],
    block_key: str,
) -> Tuple[Worksheet, int, int]:
    """
    Return (worksheet, anchor_row, value_column) for a vertical value stack.
    Value for row_delta N is ws.cell(anchor_row + N, value_column).
    """
    block = manifest.get(block_key) or {}
    v_off = int(block.get("value_column_offset_from_anchor", 1))
    ws = get_target_sheet(wb, manifest.get("sheet_names") or ["Dig Package"])
    candidates = _anchor_candidate_tuples(block, manifest, mdl_header_hint=None)
    if not candidates:
        raise LayoutError(f"Layout block {block_key!r} has no anchor text")
    ar, ac = find_anchor_cell_candidates(
        ws, candidates, context=f"{block_key}: "
    )
    label_right = _anchor_merge_max_col(ws, ar, ac)
    return ws, ar, label_right + v_off


def resolve_feature_table_data_start(
    wb: Workbook,
    manifest: Dict[str, Any],
    mdl_col_map: Optional[Dict[str, str]] = None,
) -> Tuple[Worksheet, int]:
    """First data row and worksheet for feature table (anchor row + offset)."""
    block = manifest.get("feature_table") or {}
    if not (block.get("anchor") or {}).get("text"):
        block = {**block, "anchor": {**(block.get("anchor") or {}), "text": "Feature ID", "match": "equals"}}
    row_off = int(block.get("data_start_row_offset", 2))
    ws = get_target_sheet(wb, manifest.get("sheet_names") or ["Dig Package"])
    ft_hint = mdl_col_map.get("feature_id") if mdl_col_map else None
    candidates = _anchor_candidate_tuples(block, manifest, mdl_header_hint=ft_hint)
    if not candidates:
        candidates = [("Feature ID", "equals")]
    ar, _ac = find_anchor_cell_candidates(ws, candidates, context="feature_table: ")
    return ws, ar + row_off


def verify_layout_against_workbook(
    wb: Workbook,
    manifest: Dict[str, Any],
    mdl_col_map: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    """
    Try to resolve every field; return list of dicts with ok / field_id / error.
    Used by CLI dev tool and tests.
    Optional ``mdl_col_map`` enables MDL header hints (same as generation).
    """
    results: List[Dict[str, Any]] = []
    for fid in (manifest.get("fields") or {}):
        try:
            resolve_field_cell(wb, manifest, fid, mdl_col_map=mdl_col_map)
            results.append({"field_id": fid, "ok": True, "error": ""})
        except AnchorNotFoundError as e:
            if fid in OPTIONAL_LAYOUT_VALUE_FIELDS:
                results.append({"field_id": fid, "ok": True, "error": ""})
            else:
                results.append({"field_id": fid, "ok": False, "error": str(e)})
        except LayoutError as e:
            results.append({"field_id": fid, "ok": False, "error": str(e)})
    for block_name in ("excavation_summary", "exposure_summary"):
        if block_name not in manifest:
            continue
        try:
            resolve_excavation_block_top_left(wb, manifest, block_name)
            results.append({"field_id": f"[{block_name}]", "ok": True, "error": ""})
        except (LayoutError, AnchorNotFoundError) as e:
            results.append({"field_id": f"[{block_name}]", "ok": False, "error": str(e)})
    if "feature_table" in manifest:
        try:
            resolve_feature_table_data_start(wb, manifest, mdl_col_map=mdl_col_map)
            results.append({"field_id": "[feature_table]", "ok": True, "error": ""})
        except (LayoutError, AnchorNotFoundError) as e:
            results.append({"field_id": "[feature_table]", "ok": False, "error": str(e)})
    return results
