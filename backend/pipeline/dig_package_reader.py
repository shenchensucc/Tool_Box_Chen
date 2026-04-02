"""
Dig Package Reader for ILI Visual Tool

Parses dig package Excel files that are visually broken down by sections with headers.
- Extracts ILI feature details from "Feature summary" section (typically bottom)
- Extracts longseam orientation from "Joint Summary" section
- Supports multiple ILI sources
- Uses "Distance from TGW (m)" as default x-axis

Shared with dig_package module: uses ili_reader.identify_ili_columns and similar column mapping.
"""

import io
import re
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from backend.logging_config import get_logger
from backend.pipeline.ili_reader import identify_ili_columns, find_column_names
from backend.pipeline.ili_reader import COLUMN_KEYWORDS
from backend.pipeline.feature_map_builder import (
    build_feature_map_from_df,
    format_orientation_hours,
    parse_orientation_to_hours,
)

logger = get_logger("backend.pipeline.dig_package_reader")

# Section header keywords (case-insensitive) to locate data blocks
FEATURE_SUMMARY_KEYWORDS = ["feature summary", "feature summary table", "ili feature summary"]
JOINT_SUMMARY_KEYWORDS = ["joint summary", "joint summary table", "girth weld summary"]

# Prefer "Distance from TGW (m)" as default x-axis for dig packages
DIG_PACKAGE_DISTANCE_KEYWORDS = [
    "Distance from TGW (m)",
    "Distance from TGW",
    "distance from tgw",
    "ILI Chainage (m)",
    "Chainage",
    "Odometer",
    "Log Dist.",
]


def _normalize_text(val: Any) -> str:
    """Normalize whitespace so merged-cell line breaks still match keywords."""
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val).replace("\xa0", " ")).strip()


def _get_effective_cell_value(ws, row_idx: int, col_idx: int):
    """
    Return the displayed value for a cell.

    For merged cells, Excel stores the content only in the top-left cell. This
    helper lets us read the same value from the rest of the merged range so the
    parsed table better matches what the user sees in Excel.
    """
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value is not None:
        return cell.value

    coord = cell.coordinate
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None


def _find_section_start_row(ws, keywords: List[str]) -> Optional[int]:
    """
    Find the first row where a section header matches any keyword.
    Section headers are typically in column A or merged cells.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=500), start=1):
        for cell in row[:5]:  # Check first 5 columns
            val = _get_effective_cell_value(ws, row_idx, cell.column)
            if val is None:
                continue
            val_str = _normalize_text(val).lower()
            for kw in keywords:
                if kw.lower() in val_str or val_str in kw.lower():
                    return row_idx
    return None


def _find_header_row_after(ws, start_row: int) -> Optional[int]:
    """
    Find the row containing column headers (first row with multiple non-empty cells).
    Typically the row immediately after section header or a few rows down.
    """
    for row_idx in range(start_row + 1, min(start_row + 20, ws.max_row + 1)):
        vals = [_normalize_text(_get_effective_cell_value(ws, row_idx, col_idx)) for col_idx in range(1, 21)]
        non_empty = [v for v in vals if v]
        unique_non_empty = list(dict.fromkeys(non_empty))
        # Reject merged title rows like "Feature Summary" repeated across columns.
        if len(non_empty) >= 3 and len(unique_non_empty) >= 3:
            return row_idx
    return None


def _make_unique_headers(headers: List[str]) -> List[str]:
    """Ensure DataFrame headers are unique so UI/Arrow rendering does not fail."""
    counts: Dict[str, int] = {}
    unique_headers: List[str] = []
    for idx, header in enumerate(headers):
        base = header or f"_col{idx}"
        current = counts.get(base, 0)
        if current == 0:
            unique_headers.append(base)
        else:
            unique_headers.append(f"{base}__{current + 1}")
        counts[base] = current + 1
    return unique_headers


def _read_section_as_dataframe(ws, header_row: int) -> pd.DataFrame:
    """Read data from header_row to end of contiguous data."""
    headers = [_get_effective_cell_value(ws, header_row, col_idx) for col_idx in range(1, ws.max_column + 1)]
    # Clean headers
    headers = [_normalize_text(h) if h is not None else f"_col{i}" for i, h in enumerate(headers)]
    headers = _make_unique_headers(headers)
    data = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        vals = [_get_effective_cell_value(ws, row_idx, col_idx) for col_idx in range(1, len(headers) + 1)]
        if not any(_normalize_text(v) for v in vals):
            # Empty row - stop if we've seen data, or continue for sparse tables
            if data:
                break
            continue
        data.append(vals)
    # Pad rows to match header length
    n_cols = len(headers)
    for i, row in enumerate(data):
        if len(row) < n_cols:
            data[i] = row + [None] * (n_cols - len(row))
        elif len(row) > n_cols:
            data[i] = row[:n_cols]
    return pd.DataFrame(data, columns=headers)


def _clean_joint_source_name(label: Any) -> str:
    source_name = _normalize_text(label)
    for part in [
        "Long Seam Orientation",
        "Longseam Orientation",
        "Seam Orientation",
        "Long Seam",
        "Longseam",
        "Orientation",
        "hh:mm",
        "Girth Weld No.",
        "Girth Weld No",
    ]:
        source_name = source_name.replace(part, " ")
    source_name = source_name.replace("(", " ").replace(")", " ")
    return re.sub(r"\s+", " ", source_name).strip(" -:")


def _is_unit_like_parenthetical(text: str) -> bool:
    """Return True for metric/unit qualifiers, not source labels."""
    norm = _normalize_text(text).lower()
    if not norm:
        return True
    if norm in {"hh:mm", "m", "mm", "in", "ft", "deg", "degree", "degrees", "clock", "clock position"}:
        return True
    # Short unit tokens only — avoid classifying multi-word source names (e.g. "alpha tool") as units.
    return bool(re.fullmatch(r"[a-z%:/.\- ]{1,6}", norm))


def _clean_joint_metric_label(label: Any) -> str:
    metric = _normalize_text(label)
    for item in re.findall(r"\(([^)]+)\)", str(label or "")):
        if not _is_unit_like_parenthetical(item):
            metric = metric.replace(f"({item})", " ")
    metric = re.sub(r"\s+", " ", metric).strip(" -:")
    return metric


def _extract_sources_from_block_label(label: Any) -> List[str]:
    """Pull source names from merged labels using structure, not vendor-specific keywords."""
    raw = str(label or "")
    candidates: List[str] = []

    for item in re.findall(r"\(([^)]+)\)", raw):
        norm = _normalize_text(item)
        if norm and not _is_unit_like_parenthetical(norm):
            candidates.append(norm)

    if not candidates:
        lines = [line for line in re.split(r"[\r\n]+", raw) if _normalize_text(line)]
        # In newline-based merged labels, the first line is usually the metric
        # and the following lines are the sources.
        for line in lines[1:]:
            norm = _normalize_text(line).strip("()")
            if norm and not _is_unit_like_parenthetical(norm):
                candidates.append(norm)

    seen = set()
    ordered = []
    for item in candidates:
        key = item.lower()
        if key not in seen:
            seen.add(key)
            ordered.append(item)
    return ordered


def _is_longseam_block_label(label: Any) -> bool:
    low = _normalize_text(label).lower()
    return ("long seam" in low or "longseam" in low or "seam orientation" in low) and "joint length" not in low


def _parse_header_token_for_gwd(token: str) -> Optional[int]:
    """Parse a header cell token that may include openpyxl duplicate suffix (e.g. 30900__2)."""
    t = str(token).strip().replace(",", "")
    t = re.sub(r"__\d+$", "", t)
    try:
        v = int(float(t))
        if 100 < v < 99999:
            return v
    except (ValueError, TypeError):
        pass
    return None


def _gwd_int_from_text(text: Any) -> Optional[int]:
    """Extract a plausible girth-weld number (4–5 digits) from a header string."""
    s = _normalize_text(text)
    for m in re.finditer(r"\b(\d{4,5})\b", s.replace(",", "")):
        v = int(m.group(1))
        if 100 < v < 99999:
            return v
    return None


def _discover_gwd_columns_from_headers(joint_df: pd.DataFrame) -> Dict[int, List[int]]:
    """
    Map each GWD number to ordered column indices (left-to-right).

    Supports one column per GWD (classic matrix) and multiple adjacent columns per GWD
    when Excel renames duplicates (e.g. 30900 and 30900__2 for two ILI runs).
    """
    gwd_to_cols: Dict[int, List[int]] = {}
    for col_idx, col_name in enumerate(joint_df.columns):
        col_str = " ".join(str(c) for c in (col_name if isinstance(col_name, tuple) else [col_name]))
        gwd_val: Optional[int] = None
        for part in col_str.replace(",", " ").split():
            gwd = _parse_header_token_for_gwd(part)
            if gwd is not None:
                gwd_val = gwd
                break
        if gwd_val is None:
            gwd_val = _gwd_int_from_text(col_str)
        if gwd_val is None:
            continue
        lst = gwd_to_cols.setdefault(gwd_val, [])
        if col_idx not in lst:
            lst.append(col_idx)
    return gwd_to_cols


def _discover_gwd_columns_from_first_numeric_row(joint_df: pd.DataFrame) -> Dict[int, List[int]]:
    """When headers lack GWD ids, use the first row where cells look like GWD numbers."""
    gwd_to_cols: Dict[int, List[int]] = {}
    for _, row in joint_df.iterrows():
        for col_idx, val in enumerate(row):
            try:
                gwd = int(float(str(val).replace(",", "")))
            except (ValueError, TypeError):
                continue
            if 100 < gwd < 99999:
                lst = gwd_to_cols.setdefault(gwd, [])
                if col_idx not in lst:
                    lst.append(col_idx)
        if sum(len(v) for v in gwd_to_cols.values()) >= 2:
            break
        gwd_to_cols.clear()
    return gwd_to_cols


def _reshape_joint_summary_dataframe(joint_df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize Joint Summary for merged-label layouts into explicit Metric/Source columns.

    Example input rows:
    - "Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)", same again, 11:28, 12:40, ...
    - "Long Seam Orientation (hh:mm) (2022 Rosen) (2025 TDW)", same again, 11:38, 12:56, ...

    Output rows:
    - "Long Seam Orientation (hh:mm)", "2022 Rosen", 11:28, 12:40, ...
    - "Long Seam Orientation (hh:mm)", "2025 TDW", 11:38, 12:56, ...
    """
    if joint_df is None or joint_df.empty or len(joint_df.columns) < 3:
        return joint_df

    df = joint_df.copy()
    cols = list(df.columns)
    first_col = cols[0]
    second_col = cols[1] if len(cols) > 1 else None

    # Common merged-header pattern: first two columns are both "Girth Weld No."
    first_norm = _normalize_text(first_col).lower()
    second_norm = _normalize_text(second_col).lower() if second_col is not None else ""
    if first_norm.startswith("girth weld") and second_norm.startswith("girth weld"):
        cols[0] = "Metric"
        cols[1] = "Source"
        df.columns = cols
    elif second_col is not None and second_norm in {"source", "ili source"}:
        cols[0] = "Metric"
        cols[1] = "Source"
        df.columns = cols
    elif first_norm.startswith("girth weld") and second_col is not None:
        gwd_hdr = _parse_header_token_for_gwd(str(second_col))
        if gwd_hdr is not None and not second_norm.startswith("girth"):
            # Single leading label column + GWD columns (4570, 4580, …); merged block labels in col0.
            prev_raw_norm: Optional[str] = None
            sub_i = 0
            new_rows: List[List[Any]] = []
            for idx in df.index:
                raw = df.at[idx, first_col]
                raw_norm = _normalize_text(raw)
                if prev_raw_norm is not None and raw_norm == prev_raw_norm:
                    sub_i += 1
                else:
                    sub_i = 0
                prev_raw_norm = raw_norm
                sources = _extract_sources_from_block_label(raw)
                cleaned_metric = _clean_joint_metric_label(raw)
                if sources:
                    src = sources[sub_i] if sub_i < len(sources) else sources[-1]
                    row_out = [cleaned_metric, _clean_joint_source_name(src)]
                else:
                    row_out = [cleaned_metric, ""]
                for c in cols[1:]:
                    row_out.append(df.at[idx, c])
                new_rows.append(row_out)
            out = pd.DataFrame(new_rows, columns=["Metric", "Source"] + cols[1:])
            return out
        return df
    else:
        return df

    row_index_by_metric: Dict[str, int] = {}
    metric_col = df.columns[0]
    source_col = df.columns[1]

    for idx in df.index:
        raw_metric = df.at[idx, metric_col]
        raw_source = df.at[idx, source_col]
        metric_text = _normalize_text(raw_metric)
        source_text = _normalize_text(raw_source)

        if not metric_text:
            continue

        sources = _extract_sources_from_block_label(metric_text)
        cleaned_metric = _clean_joint_metric_label(metric_text)

        if sources:
            metric_occurrence = row_index_by_metric.get(metric_text, 0)
            assigned_source = sources[metric_occurrence] if metric_occurrence < len(sources) else sources[-1]
            row_index_by_metric[metric_text] = metric_occurrence + 1
            df.at[idx, metric_col] = cleaned_metric
            if not source_text or source_text == metric_text or source_text == cleaned_metric:
                df.at[idx, source_col] = assigned_source
        elif source_text:
            df.at[idx, metric_col] = cleaned_metric
            df.at[idx, source_col] = _clean_joint_source_name(source_text)

    return df


def parse_dig_package_excel(file_content: bytes) -> Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame], Dict[str, Any]]:
    """
    Parse a dig package Excel file with section headers.

    Returns:
        (feature_summary_df, joint_summary_df, metadata)
        - feature_summary_df: ILI features from "Feature summary" section
        - joint_summary_df: Joint/girth weld data from "Joint Summary" (for longseam orientation)
        - metadata: dict with section info, sheet used, etc.
    """
    wb = load_workbook(io.BytesIO(file_content), data_only=True, read_only=False)
    metadata = {"sheets_searched": [], "feature_section_found": False, "joint_section_found": False}

    feature_df = None
    joint_df = None

    for sheet_name in wb.sheetnames:
        metadata["sheets_searched"].append(sheet_name)
        ws = wb[sheet_name]

        # Look for Feature summary section
        if feature_df is None:
            start_row = _find_section_start_row(ws, FEATURE_SUMMARY_KEYWORDS)
            if start_row is not None:
                header_row = _find_header_row_after(ws, start_row)
                if header_row is not None:
                    feature_df = _read_section_as_dataframe(ws, header_row)
                    if not feature_df.empty:
                        metadata["feature_section_found"] = True
                        metadata["feature_sheet"] = sheet_name
                        metadata["feature_header_row"] = header_row

        # Look for Joint Summary section (for longseam orientation)
        if joint_df is None:
            start_row = _find_section_start_row(ws, JOINT_SUMMARY_KEYWORDS)
            if start_row is not None:
                header_row = _find_header_row_after(ws, start_row)
                if header_row is not None:
                    joint_df = _read_section_as_dataframe(ws, header_row)
                    joint_df = _reshape_joint_summary_dataframe(joint_df)
                    if not joint_df.empty:
                        metadata["joint_section_found"] = True
                        metadata["joint_sheet"] = sheet_name
                        metadata["joint_header_row"] = header_row

        if feature_df is not None and joint_df is not None:
            break

    wb.close()
    return feature_df, joint_df, metadata


def _orientation_cell_accepted(val: Any, hours: Optional[float]) -> bool:
    """True if the cell looks like a clock longseam reading (not joint length metres)."""
    if hours is None:
        return False
    val_text = _normalize_text(val)
    is_clock_like = ":" in val_text and any(ch.isdigit() for ch in val_text)
    return bool(is_clock_like or 0 <= hours <= 12.0)


def _parse_joint_summary_matrix(
    joint_df: pd.DataFrame,
    parse_orientation_to_hours,
    logger,
) -> Optional[Dict[str, Any]]:
    """
    Parse Joint Summary when structure is: GWDs as columns, rows per ILI source.
    E.g. | Girth Weld No. | 3150 | 3160 | 3170 | 3180 | ...
         | 2022 Rosen | 01:36 | 07:42 | 11:50 | 11:30 | ...
         | 2025 TDW | 01:30 | 07:42 | 11:38 | 11:30 | ...

    Also supports PNG-style packages with two adjacent columns per GWD (duplicate headers
    like 30900 / 30900__2) and a single longseam row whose merged label lists both ILI sources.

    Value at column GWD_i = D/S longseam at that GWD. For span [GWD_i, GWD_{i+1}], use value at GWD_i.
    """
    if joint_df.empty or len(joint_df.columns) < 2:
        return None

    gwd_to_cols = _discover_gwd_columns_from_headers(joint_df)
    if len(gwd_to_cols) < 2:
        gwd_to_cols = _discover_gwd_columns_from_first_numeric_row(joint_df)

    if len(gwd_to_cols) < 2:
        logger.debug("[Joint Summary] Matrix: could not find GWD columns")
        return None

    # Identify the Target column / target GWD for positional anchoring.
    target_gwd_from_header: Optional[int] = None
    for col_idx_h, col_name_h in enumerate(joint_df.columns):
        hn = _normalize_text(col_name_h).strip().lower()
        if hn == "target" or hn.startswith("target "):
            tg = _gwd_int_from_text(col_name_h)
            if tg is not None:
                target_gwd_from_header = tg
                break
            for gwd_num, col_list in gwd_to_cols.items():
                if col_idx_h in col_list:
                    target_gwd_from_header = gwd_num
                    break
            break
    if target_gwd_from_header is None:
        for col_name_h in joint_df.columns:
            hn = _normalize_text(col_name_h).lower()
            if "target" in hn and "gwd" in hn:
                target_gwd_from_header = _gwd_int_from_text(col_name_h)
                if target_gwd_from_header is not None:
                    break

    joint_lengths_raw: Dict[int, List[float]] = {}

    by_source: Dict[str, Dict[int, float]] = {}
    parsed = []
    seen_sources = set()
    block_row_index: Dict[str, int] = {}
    source_col_idx = None
    metric_col_idx = 0
    if len(joint_df.columns) >= 2:
        second_header = _normalize_text(joint_df.columns[1]).lower()
        first_header = _normalize_text(joint_df.columns[0]).lower()
        if second_header == "source":
            source_col_idx = 1
        elif first_header == "metric" and second_header.startswith("girth weld"):
            source_col_idx = 1

    col_widths = {len(v) for v in gwd_to_cols.values()}
    max_col_width = max(col_widths) if col_widths else 0

    for idx, row in joint_df.iterrows():
        first_cell = _normalize_text(row.iloc[metric_col_idx] if len(row) else "")
        second_cell = _normalize_text(row.iloc[1]) if len(row) > 1 else ""
        source_cell = _normalize_text(row.iloc[source_col_idx]) if source_col_idx is not None and source_col_idx < len(row) else ""
        first_lower = first_cell.lower()
        second_lower = second_cell.lower()
        label_lower = first_lower or second_lower
        logger.debug(
            f"[Joint Summary] Row {idx}: col0={repr(first_cell[:60] if first_cell else '')}, "
            f"col1={repr(second_cell[:40] if second_cell else '')}"
        )
        if "girth weld" in label_lower or ("joint" in label_lower and "no" in label_lower):
            continue
        if "joint length" in label_lower:
            for gwd_j, col_list in gwd_to_cols.items():
                for col_j in col_list:
                    if col_j >= len(row):
                        continue
                    raw_l = row.iloc[col_j]
                    if raw_l is None:
                        continue
                    try:
                        lm = float(str(raw_l).replace(",", ""))
                        if 0.5 < lm < 200.0:
                            joint_lengths_raw.setdefault(gwd_j, []).append(lm)
                    except (ValueError, TypeError):
                        pass
            continue

        block_label = first_cell if _is_longseam_block_label(first_cell) else (
            second_cell if _is_longseam_block_label(second_cell) else ""
        )
        block_sources = _extract_sources_from_block_label(block_label) if block_label else []
        interleaved = (
            max_col_width > 1
            and block_label
            and len(block_sources) == max_col_width
            and col_widths == {max_col_width}
        )

        if interleaved:
            consumed_block = False
            for si, raw_src in enumerate(block_sources):
                source_name = _clean_joint_source_name(raw_src) or raw_src
                gwd_to_hours_one: Dict[int, float] = {}
                for gwd in sorted(gwd_to_cols.keys()):
                    cols = gwd_to_cols[gwd]
                    if si >= len(cols):
                        continue
                    col_idx = cols[si]
                    if col_idx >= len(row):
                        continue
                    val = row.iloc[col_idx]
                    hours = parse_orientation_to_hours(val) if val is not None else None
                    if _orientation_cell_accepted(val, hours):
                        gwd_to_hours_one[gwd] = hours  # type: ignore[assignment]
                if not gwd_to_hours_one:
                    continue
                consumed_block = True
                eff_name = source_name
                if eff_name in seen_sources:
                    eff_name = f"{eff_name} #{idx}"
                seen_sources.add(eff_name)
                by_source[eff_name] = gwd_to_hours_one
                for gwd, hours in gwd_to_hours_one.items():
                    raw_idx = gwd_to_cols[gwd][si]
                    raw_val = row.iloc[raw_idx] if raw_idx < len(row) else None
                    parsed.append({
                        "GWD": gwd,
                        "Source": eff_name,
                        "Long Seam (D/S)": str(raw_val) if raw_val is not None else None,
                        "Longseam (hours)": round(hours, 2),
                    })
                logger.debug(f"[Joint Summary] Matrix (interleaved): source '{eff_name}' -> {len(gwd_to_hours_one)} GWDs")
            if consumed_block:
                block_row_index[block_label] = len(block_sources)
                continue

        gwd_to_hours: Dict[int, float] = {}
        has_orientation_values = False
        for gwd, col_list in gwd_to_cols.items():
            col_idx = col_list[0]
            if col_idx >= len(row):
                continue
            val = row.iloc[col_idx]
            hours = parse_orientation_to_hours(val) if val is not None else None
            if _orientation_cell_accepted(val, hours):
                has_orientation_values = True
                gwd_to_hours[gwd] = hours  # type: ignore[assignment]
        if not gwd_to_hours or not has_orientation_values:
            continue

        if source_cell:
            source_name = _clean_joint_source_name(source_cell) or f"Source_{idx}"
        elif first_cell and not block_label and "joint length" not in first_lower:
            source_name = _clean_joint_source_name(first_cell) or f"Source_{idx}"
        elif second_cell and not block_label and "joint length" not in second_lower:
            source_name = _clean_joint_source_name(second_cell) or f"Source_{idx}"
        elif block_label:
            block_sources_stacked = _extract_sources_from_block_label(block_label)
            current_idx = block_row_index.get(block_label, 0)
            source_name = (
                block_sources_stacked[current_idx]
                if current_idx < len(block_sources_stacked)
                else (block_sources_stacked[-1] if block_sources_stacked else f"Source_{idx}")
            )
            block_row_index[block_label] = current_idx + 1
        else:
            source_name = _clean_joint_source_name(first_cell or second_cell) or f"Source_{idx}"

        if source_name in seen_sources:
            source_name = f"{source_name} #{idx}"
        seen_sources.add(source_name)

        for gwd, hours in gwd_to_hours.items():
            raw_idx = gwd_to_cols[gwd][0]
            raw_val = row.iloc[raw_idx] if raw_idx < len(row) else None
            parsed.append({
                "GWD": gwd,
                "Source": source_name,
                "Long Seam (D/S)": str(raw_val) if raw_val is not None else None,
                "Longseam (hours)": round(hours, 2),
            })

        by_source[source_name] = gwd_to_hours
        logger.debug(f"[Joint Summary] Matrix: source '{source_name}' -> {len(gwd_to_hours)} GWDs")

    if not by_source:
        return None

    joint_lengths: Dict[int, float] = {
        gwd: round(sum(vals) / len(vals), 2)
        for gwd, vals in joint_lengths_raw.items()
    }

    return {
        "by_source": by_source,
        "gwd_by_source": by_source,
        "parsed": parsed,
        "gwd_to_chainage": {},
        "gwd_order": sorted(gwd_to_cols.keys()),
        "target_gwd_from_header": target_gwd_from_header,
        "joint_lengths": joint_lengths,
    }


def _get_dig_package_distance_column(df: pd.DataFrame) -> Optional[str]:
    """Prefer 'Distance from TGW (m)' for dig packages, fallback to standard distance columns."""
    for kw in DIG_PACKAGE_DISTANCE_KEYWORDS:
        found = find_column_names(df, [kw])
        if found:
            return found
    return find_column_names(df, COLUMN_KEYWORDS["distance"])


SOURCE_TOKEN_STOPWORDS = {
    "ili", "tool", "inspection", "source", "vendor", "feature", "summary",
    "joint", "long", "seam", "orientation", "orient", "clock",
}


def _source_tokens(value: Any) -> List[str]:
    text = _normalize_text(value).lower()
    if not text:
        return []
    return [tok for tok in re.findall(r"[a-z0-9]+", text) if tok and tok not in SOURCE_TOKEN_STOPWORDS]


def _match_joint_source_name(feature_source: str, available_joint_sources: List[str]) -> Optional[str]:
    """Match a Feature Summary source to the closest Joint Summary source."""
    if not available_joint_sources:
        return None
    if len(available_joint_sources) == 1:
        return available_joint_sources[0]

    fs_norm = _normalize_text(feature_source).lower()
    fs_tokens = set(_source_tokens(feature_source))
    best_source = None
    best_score = (-1, -1, -1)

    for joint_source in available_joint_sources:
        js_norm = _normalize_text(joint_source).lower()
        js_tokens = set(_source_tokens(joint_source))
        subset = 1 if js_tokens and js_tokens.issubset(fs_tokens) else 0
        overlap = len(fs_tokens & js_tokens)
        contains = 1 if (js_norm and js_norm in fs_norm) or (fs_norm and fs_norm in js_norm) else 0
        score = (subset, overlap, contains)
        if score > best_score:
            best_score = score
            best_source = joint_source

    if best_source and best_score > (0, 0, 0):
        return best_source
    return available_joint_sources[0]


def _normalize_gwd_number(value: Any) -> Optional[int]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _get_seam_hours_for_gwd(seam_map: Dict[int, float], gwd_number: Optional[int]) -> Optional[float]:
    if gwd_number is None or not seam_map:
        return None
    if gwd_number in seam_map:
        return seam_map[gwd_number]
    nearest = min(seam_map.keys(), key=lambda key: abs(key - gwd_number), default=None)
    if nearest is not None and abs(nearest - gwd_number) <= 15:
        return seam_map[nearest]
    return None


def _segment_joint_length_m(joint_lengths: Dict[int, float], g_left: int, g_right: int) -> Optional[float]:
    """Metres for the pipe joint between g_left and g_right (TGW layout).

    Same convention as longseam columns: the **upstream** GWD (g_left) column carries the
    value for the span to the next weld. Prefer g_left, then g_right for odd templates that
    key lengths on the downstream weld only.
    """
    for key in (g_left, g_right):
        v = joint_lengths.get(key)
        if v is not None:
            try:
                f = float(v)
            except (TypeError, ValueError):
                continue
            if 0.5 < f < 200.0:
                return f
    return None


def _build_tgw_layout_from_joint_summary(
    parsed_matrix: Optional[Dict[str, Any]],
    gwd_to_seam_by_source: Dict[str, Dict[int, float]],
    sources: List[str],
) -> Optional[Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]]:
    """
    Place four girth welds on Distance-from-TGW axis using Joint Summary joint lengths:

        ch(target)=0, ch(upstream)=0−L_us, ch(ds1)=0+L_ds1, ch(ds2)=L_ds1+L_ds2

    Three joint bodies lie between the four red lines. Longseam for each span uses the
    **upstream** GWD of that span (same convention as matrix seam_welds).

    Returns (girth_welds, seam_welds) or None if data is insufficient.
    """
    if not parsed_matrix or not gwd_to_seam_by_source:
        return None
    joint_lengths: Dict[int, float] = parsed_matrix.get("joint_lengths") or {}
    gwd_order = list(parsed_matrix.get("gwd_order") or [])
    if len(gwd_order) < 4 or len(joint_lengths) < 1:
        return None

    target_h = parsed_matrix.get("target_gwd_from_header")
    if target_h is not None and target_h in gwd_order:
        it = gwd_order.index(target_h)
    else:
        # Centre target so four consecutive GWDs fit: for len=4 need it=1, not len//2 (=2).
        it = (len(gwd_order) - 1) // 2
    if it < 1 or it + 2 >= len(gwd_order):
        return None

    g0, g1, g2, g3 = gwd_order[it - 1], gwd_order[it], gwd_order[it + 1], gwd_order[it + 2]
    L01 = _segment_joint_length_m(joint_lengths, g0, g1)
    L12 = _segment_joint_length_m(joint_lengths, g1, g2)
    L23 = _segment_joint_length_m(joint_lengths, g2, g3)
    if L01 is None or L12 is None or L23 is None:
        return None

    ch1 = 0.0
    ch0 = ch1 - L01
    ch2 = ch1 + L12
    ch3 = ch2 + L23

    first_js = next(iter(gwd_to_seam_by_source.keys()), "")
    seam0 = gwd_to_seam_by_source.get(first_js, {})
    feat_src = sources[0] if sources else ""

    def _primary_longseam_label(gn: int) -> Optional[str]:
        oh = _get_seam_hours_for_gwd(seam0, gn)
        return format_orientation_hours(oh) if oh is not None else None

    girths = [
        {
            "chainage": ch0,
            "gwd_number": g0,
            "label": f"GWD {g0}",
            "source": feat_src,
            "joint_summary_layout": True,
            "longseam_label_primary": _primary_longseam_label(g0),
        },
        {
            "chainage": ch1,
            "gwd_number": g1,
            "label": f"GWD {g1} (TGW)",
            "source": feat_src,
            "joint_summary_layout": True,
            "longseam_label_primary": _primary_longseam_label(g1),
        },
        {
            "chainage": ch2,
            "gwd_number": g2,
            "label": f"GWD {g2}",
            "source": feat_src,
            "joint_summary_layout": True,
            "longseam_label_primary": _primary_longseam_label(g2),
        },
        {
            "chainage": ch3,
            "gwd_number": g3,
            "label": f"GWD {g3}",
            "source": feat_src,
            "joint_summary_layout": True,
            "longseam_label_primary": _primary_longseam_label(g3),
        },
    ]

    seam_welds: List[Dict[str, Any]] = []
    seen_span_source = set()
    for g_left, c_a, c_b in ((g0, ch0, ch1), (g1, ch1, ch2), (g2, ch2, ch3)):
        for js_src, seam_map in gwd_to_seam_by_source.items():
            oh = _get_seam_hours_for_gwd(seam_map, g_left)
            if oh is None:
                continue
            span_key = (c_a, c_b, js_src)
            if span_key in seen_span_source:
                continue
            seen_span_source.add(span_key)
            seam_welds.append({
                "chainage_start": c_a,
                "chainage_end": c_b,
                "orientation_hours": oh,
                "orientation_label": format_orientation_hours(oh),
                "source": js_src,
                "feature_source": feat_src,
                "gwd_number": g_left,
            })

    if not seam_welds:
        return None
    return girths, seam_welds


def _build_joint_context_tgw_synthetic(
    girths_four: List[Dict[str, Any]],
    gwd_to_seam_by_source: Dict[str, Dict[int, float]],
    sources: List[str],
) -> Dict[str, Dict[str, Any]]:
    """Upstream / target / downstream context when girth welds come from Joint Summary TGW layout."""
    if len(girths_four) < 3:
        return {}
    ordered = sorted(girths_four, key=lambda g: g.get("chainage", 0) or 0)
    g_us, g_tgt, g_ds = ordered[0], ordered[1], ordered[2]
    joint_sources = list(gwd_to_seam_by_source.keys())
    contexts: Dict[str, Dict[str, Any]] = {}

    def _payload(role: str, gw: Dict[str, Any], seam_map: Dict[int, float]) -> Dict[str, Any]:
        gn = _normalize_gwd_number(gw.get("gwd_number"))
        sh = _get_seam_hours_for_gwd(seam_map, gn)
        return {
            "role": role,
            "gwd_number": gn,
            "chainage": gw.get("chainage"),
            "label": gw.get("label"),
            "longseam_hours": sh,
            "longseam_label": format_orientation_hours(sh) if sh is not None else None,
        }

    for feat_src in sources:
        js = _match_joint_source_name(feat_src, joint_sources) if joint_sources else None
        sm = gwd_to_seam_by_source.get(js, {}) if js else next(iter(gwd_to_seam_by_source.values()), {})
        contexts[feat_src] = {
            "feature_source": feat_src,
            "joint_source": js,
            "upstream": _payload("upstream", g_us, sm),
            "target": _payload("target", g_tgt, sm),
            "downstream": _payload("downstream", g_ds, sm),
        }
    return contexts


def _dedupe_girth_welds(girth_welds: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen = set()
    deduped = []
    for gw in sorted(girth_welds, key=lambda item: item.get("chainage", 0) or 0):
        chainage = gw.get("chainage")
        key = (
            round(float(chainage), 6) if isinstance(chainage, (int, float)) else chainage,
            _normalize_gwd_number(gw.get("gwd_number")),
            _normalize_text(gw.get("source")),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(gw)
    return deduped


def _build_joint_context(
    girth_welds: List[Dict[str, Any]],
    seam_map_by_joint_source: Dict[str, Dict[int, float]],
    feature_sources: List[str],
) -> Dict[str, Dict[str, Any]]:
    """Build upstream/target/downstream joint context for each feature source."""
    contexts: Dict[str, Dict[str, Any]] = {}
    if not girth_welds:
        return contexts

    deduped = _dedupe_girth_welds(girth_welds)
    joint_sources = list(seam_map_by_joint_source.keys())

    for feature_source in feature_sources:
        source_girths = [gw for gw in deduped if _normalize_text(gw.get("source")) == _normalize_text(feature_source)]
        if not source_girths:
            continue

        source_girths.sort(key=lambda item: abs(item.get("chainage", 0) or 0))
        target_gw = min(source_girths, key=lambda item: abs(item.get("chainage", 0) or 0))
        target_chainage = target_gw.get("chainage", 0) or 0
        ordered = sorted(source_girths, key=lambda item: item.get("chainage", 0) or 0)
        upstream = None
        downstream = None
        for gw in ordered:
            ch = gw.get("chainage")
            if ch is None:
                continue
            if ch < target_chainage:
                upstream = gw
            elif ch > target_chainage and downstream is None:
                downstream = gw
        joint_source = _match_joint_source_name(feature_source, joint_sources)
        seam_map = seam_map_by_joint_source.get(joint_source, {}) if joint_source else {}

        def _joint_payload(role: str, gw: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
            if not gw:
                return None
            gwd_number = _normalize_gwd_number(gw.get("gwd_number"))
            seam_hours = _get_seam_hours_for_gwd(seam_map, gwd_number)
            return {
                "role": role,
                "gwd_number": gwd_number,
                "chainage": gw.get("chainage"),
                "label": gw.get("label"),
                "longseam_hours": seam_hours,
                "longseam_label": format_orientation_hours(seam_hours) if seam_hours is not None else None,
            }

        contexts[feature_source] = {
            "feature_source": feature_source,
            "joint_source": joint_source,
            "upstream": _joint_payload("upstream", upstream),
            "target": _joint_payload("target", target_gw),
            "downstream": _joint_payload("downstream", downstream),
        }
    return contexts


def build_feature_map_from_dig_package(
    file_content: bytes,
) -> Tuple[List[Dict], Optional[Dict], List[str], Dict[str, Any], List[Dict[str, Any]], Dict[str, Any]]:
    """
    Parse dig package Excel and build feature map data for ILI Visual Tool.

    Returns:
        (features, scatter_data, sources, column_mapping, joint_summary_parsed, feature_summary_raw)
    """
    feature_df, joint_df, metadata = parse_dig_package_excel(file_content)

    if feature_df is None or feature_df.empty:
        raise ValueError(
            "Could not find 'Feature summary' section in the dig package. "
            "Ensure the Excel has a section header containing 'Feature summary'."
        )

    # Use dig-package-specific distance column (prefer Distance from TGW)
    ili_cols = identify_ili_columns(feature_df)
    dist_col = _get_dig_package_distance_column(feature_df)
    if not dist_col:
        dist_col = ili_cols.get("distance")
    if not dist_col:
        raise ValueError(
            "No distance/chainage column found. Dig packages typically have 'Distance from TGW (m)'."
        )
    ili_cols["distance"] = dist_col

    # Parse Joint Summary for longseam orientation
    # Supports two structures:
    # 1) Row-based: one row per GWD, columns for Long Seam Orientation
    # 2) Matrix: GWDs as columns, rows "Long Seam - 2022 Rosen", "Long Seam - 2025 TDW"
    # U/S seam = longseam at upstream GWD (left of target). D/S seam = longseam at this GWD.
    # For span [GWD_i, GWD_{i+1}]: use D/S of GWD_i = value at left boundary.
    # "2022 Rosen" applies to all Rosen (MFL-A, MFL-C, EMAT).
    joint_summary_parsed = []
    gwd_to_seam_by_source: Dict[str, Dict[int, float]] = {}  # source -> {gwd_number -> hours} (matrix)
    gwd_to_seam: Dict[float, float] = {}  # fallback: chainage -> hours (row-based)
    gwd_to_chainage: Dict[int, float] = {}
    use_gwd_lookup = False  # True when matrix structure (no chainage in Joint Summary)
    parsed_matrix: Optional[Dict[str, Any]] = None   # kept in scope for positional GWD mapping below

    if joint_df is not None and not joint_df.empty:
        logger.info(f"[Joint Summary] Columns: {list(joint_df.columns)}, shape={joint_df.shape}")

        # Try matrix structure: GWDs as column headers, longseam rows per ILI source
        parsed_matrix = _parse_joint_summary_matrix(joint_df, parse_orientation_to_hours, logger)
        if parsed_matrix:
            gwd_to_seam_by_source = parsed_matrix.get("gwd_by_source", parsed_matrix.get("by_source", {}))
            joint_summary_parsed = parsed_matrix.get("parsed", [])
            gwd_to_chainage = parsed_matrix.get("gwd_to_chainage", {})
            use_gwd_lookup = True
            logger.info(
                f"[Joint Summary] Matrix parsed: sources={list(gwd_to_seam_by_source.keys())}, "
                f"GWDs={list(next(iter(gwd_to_seam_by_source.values()), {}).keys())[:7] if gwd_to_seam_by_source else []}..., "
                f"target_gwd_header={parsed_matrix.get('target_gwd_from_header')}"
            )
        else:
            # Fallback: row-based structure
            dist_to_target_col = find_column_names(
                joint_df,
                ["Distance from TGW (m)", "Distance to Target", "Distance to target", "Distance from TGW"],
            )
            chainage_col = find_column_names(joint_df, ["Chainage", "ILI Chainage (m)", "Odometer"])
            orient_col = find_column_names(
                joint_df,
                [
                    "Long Seam Orientation for the Target",
                    "Longseam Orientation for the Target",
                    "Long Seam Orientation (hh:mm)",
                    "Longseam Orientation",
                    "Seam Orientation",
                    "Longseam Orient",
                    "Orientation (hh:mm)",
                    "Orientation (clock)",
                ],
            )
            gwd_col = find_column_names(joint_df, ["GWD", "Joint", "Joint Number", "Joint No", "Girth Weld No"])

            if orient_col:
                rows_data = []
                for _, row in joint_df.iterrows():
                    gwd_val = row.get(gwd_col) if gwd_col else None
                    ch = pd.to_numeric(row.get(chainage_col or dist_to_target_col), errors="coerce")
                    dist_to_target = pd.to_numeric(row.get(dist_to_target_col), errors="coerce") if dist_to_target_col else None
                    orient = row.get(orient_col)
                    hours = parse_orientation_to_hours(orient) if orient is not None else None
                    if pd.notna(ch):
                        rows_data.append({
                            "gwd": gwd_val,
                            "chainage": float(ch),
                            "distance_to_target": float(dist_to_target) if pd.notna(dist_to_target) else None,
                            "longseam_hours": hours,
                            "longseam_raw": str(orient) if orient is not None else None,
                        })
                        joint_summary_parsed.append({
                            "GWD": gwd_val,
                            "Chainage": float(ch),
                            "Distance to Target": dist_to_target,
                            "Long Seam Orientation": str(orient) if orient is not None else None,
                            "Longseam (hours)": round(hours, 2) if hours is not None else None,
                        })

                rows_data.sort(key=lambda r: r["chainage"])
                for i, r in enumerate(rows_data):
                    dist = r.get("distance_to_target")
                    if dist is not None and dist < 0 and i > 0:
                        effective_hours = rows_data[i - 1].get("longseam_hours")
                        used_from = "GWD-1 (left)"
                    else:
                        effective_hours = r.get("longseam_hours")
                        used_from = "this row"
                    if effective_hours is not None:
                        gwd_to_seam[r["chainage"]] = effective_hours
                        if r.get("gwd") is not None:
                            try:
                                gwd_num = int(float(r["gwd"]))
                                gwd_to_chainage[gwd_num] = r["chainage"]
                            except (TypeError, ValueError):
                                pass
                    for jp in joint_summary_parsed:
                        if jp.get("Chainage") == r["chainage"]:
                            jp["Effective Longseam (hours)"] = round(effective_hours, 2) if effective_hours is not None else None
                            jp["Used from"] = used_from
                            break
                logger.info(f"[Joint Summary] Row-based parsed: {len(gwd_to_seam)} chainage->seam entries")
            else:
                logger.warning("[Joint Summary] No orientation column found for row-based structure")

    features, scatter_data, sources = build_feature_map_from_df(feature_df, ili_cols)

    girth_sorted = sorted(
        scatter_data.get("girth_welds", []) if scatter_data else [],
        key=lambda gw: gw.get("chainage", 0) or 0,
    )

    target_gwd: Optional[int] = None
    target_longseam_hours: Optional[float] = None
    joint_context_by_source: Dict[str, Dict[str, Any]] = {}

    tgw_layout = None
    if use_gwd_lookup and gwd_to_seam_by_source and parsed_matrix:
        tgw_layout = _build_tgw_layout_from_joint_summary(
            parsed_matrix, gwd_to_seam_by_source, sources
        )

    if tgw_layout:
        syn_girth, syn_seam = tgw_layout
        scatter_data = scatter_data or {}
        scatter_data["girth_welds"] = syn_girth
        scatter_data["seam_welds"] = syn_seam
        scatter_data["joint_summary_tgw_layout"] = True
        girth_sorted = list(syn_girth)
        joint_context_by_source = _build_joint_context_tgw_synthetic(
            syn_girth, gwd_to_seam_by_source, sources
        )
        ordered_syn = sorted(syn_girth, key=lambda g: g.get("chainage", 0) or 0)
        if len(ordered_syn) >= 2:
            target_gwd = _normalize_gwd_number(ordered_syn[1].get("gwd_number"))
        first_map = next(iter(gwd_to_seam_by_source.values()), {})
        if target_gwd is not None and target_gwd in first_map:
            target_longseam_hours = first_map[target_gwd]
        if joint_context_by_source and scatter_data:
            role_labels = {"upstream": "U/S", "target": "Target", "downstream": "D/S"}
            for gw in scatter_data.get("girth_welds", []):
                gwd_num = _normalize_gwd_number(gw.get("gwd_number"))
                done = False
                for ctx in joint_context_by_source.values():
                    for role in ["upstream", "target", "downstream"]:
                        item = ctx.get(role)
                        if item and item.get("gwd_number") == gwd_num:
                            gw["joint_role"] = role
                            seam_l = item.get("longseam_label")
                            base = f"{role_labels[role]} GWD {gwd_num}"
                            gw["label"] = f"{base} @ {seam_l}" if seam_l else base
                            done = True
                            break
                    if done:
                        break
        logger.info(
            "[Joint Summary] TGW layout from joint lengths: 4 girth welds, %s seam records",
            len(syn_seam),
        )

    elif girth_sorted and (gwd_to_seam_by_source or gwd_to_seam):
        # ── Positional GWD mapping ────────────────────────────────────────────
        # When the Feature Summary has no GWD number column (gwd_number = None on
        # all girth welds), we match each girth weld to a GWD from the matrix by
        # sorted position, anchored on the target GWD.
        #
        # The target girth weld = closest to chainage 0 (Distance from TGW = 0).
        # The target GWD in the matrix = explicitly marked "Target" column header,
        # or the middle GWD in sorted order.
        gwd_order_from_matrix: List[int] = sorted(
            next(iter(gwd_to_seam_by_source.values()), {}).keys()
        ) if gwd_to_seam_by_source else []

        girth_positional_gwd: Dict[int, int] = {}
        if gwd_order_from_matrix:
            # Identify target girth weld index (closest to chainage 0)
            tgt_girth_idx = min(
                range(len(girth_sorted)),
                key=lambda k: abs(girth_sorted[k].get("chainage", 0) or 0),
            ) if girth_sorted else 0

            # Identify target GWD index in the matrix order
            # Priority: "Target" column header → middle of sorted list
            target_gwd_header = None
            if parsed_matrix:
                target_gwd_header = parsed_matrix.get("target_gwd_from_header")
            if target_gwd_header is not None and target_gwd_header in gwd_order_from_matrix:
                tgt_matrix_idx = gwd_order_from_matrix.index(target_gwd_header)
            else:
                tgt_matrix_idx = len(gwd_order_from_matrix) // 2

            for j in range(len(girth_sorted)):
                offset = j - tgt_girth_idx
                matrix_idx = tgt_matrix_idx + offset
                if 0 <= matrix_idx < len(gwd_order_from_matrix):
                    girth_positional_gwd[j] = gwd_order_from_matrix[matrix_idx]

            logger.debug(
                f"[Joint Summary] Positional GWD map: "
                f"target_girth_idx={tgt_girth_idx}, target_matrix_idx={tgt_matrix_idx}, "
                f"mapping={girth_positional_gwd}"
            )

        seam_welds = []
        seen_span_source = set()
        for i in range(len(girth_sorted) - 1):
            ch_start = girth_sorted[i].get("chainage")
            ch_end = girth_sorted[i + 1].get("chainage")
            gwd_start = girth_sorted[i].get("gwd_number")
            feat_source = girth_sorted[i].get("source", "")

            if ch_start is None or ch_end is None:
                continue

            # Resolve the GWD integer for this span (with positional fallback).
            gwd_int = _normalize_gwd_number(gwd_start)
            if gwd_int is None and i in girth_positional_gwd:
                gwd_int = girth_positional_gwd[i]
                logger.debug(f"[Joint Summary] Span {i}: positional GWD fallback -> {gwd_int}")

            # ── Build a seam weld entry for EVERY joint source ────────────────
            # Engineers need all ILI sources' longseam data per span so they can
            # compare and decide which reading is most reliable.
            if gwd_to_seam_by_source and use_gwd_lookup:
                for js_src, seam_map in gwd_to_seam_by_source.items():
                    oh = _get_seam_hours_for_gwd(seam_map, gwd_int)
                    if oh is not None:
                        span_key = (ch_start, ch_end, js_src)
                        if span_key not in seen_span_source:
                            seen_span_source.add(span_key)
                            seam_welds.append({
                                "chainage_start": ch_start,
                                "chainage_end": ch_end,
                                "orientation_hours": oh,
                                "orientation_label": format_orientation_hours(oh),
                                "source": js_src,
                                "feature_source": feat_source,
                            })

            elif gwd_to_seam_by_source and not use_gwd_lookup:
                for js_src, seam_map in gwd_to_seam_by_source.items():
                    best = min(seam_map.keys(), key=lambda k: abs(k - ch_start), default=None) if seam_map else None
                    if best is not None and abs(best - ch_start) < 2.0:
                        oh = seam_map[best]
                        span_key = (ch_start, ch_end, js_src)
                        if span_key not in seen_span_source:
                            seen_span_source.add(span_key)
                            seam_welds.append({
                                "chainage_start": ch_start,
                                "chainage_end": ch_end,
                                "orientation_hours": oh,
                                "orientation_label": format_orientation_hours(oh),
                                "source": js_src,
                                "feature_source": feat_source,
                            })

            elif gwd_to_seam:
                best = min(gwd_to_seam.keys(), key=lambda k: abs(k - ch_start), default=None)
                if best is not None and abs(best - ch_start) < 2.0:
                    oh = gwd_to_seam[best]
                    span_key = (ch_start, ch_end, "")
                    if span_key not in seen_span_source:
                        seen_span_source.add(span_key)
                        seam_welds.append({
                            "chainage_start": ch_start,
                            "chainage_end": ch_end,
                            "orientation_hours": oh,
                            "orientation_label": format_orientation_hours(oh),
                            "source": "",
                            "feature_source": feat_source,
                        })

        joint_context_by_source = _build_joint_context(
            girth_sorted,
            gwd_to_seam_by_source if gwd_to_seam_by_source else {},
            sources,
        )

        if joint_context_by_source and scatter_data:
            role_labels = {"upstream": "U/S", "target": "Target", "downstream": "D/S"}
            for gw in scatter_data.get("girth_welds", []):
                src = _normalize_text(gw.get("source"))
                gwd_num = _normalize_gwd_number(gw.get("gwd_number"))
                for ctx in joint_context_by_source.values():
                    if _normalize_text(ctx.get("feature_source")) != src:
                        continue
                    for role in ["upstream", "target", "downstream"]:
                        item = ctx.get(role)
                        if item and item.get("gwd_number") == gwd_num:
                            gw["joint_role"] = role
                            gw["label"] = f"{role_labels[role]} GWD {gwd_num}"
                            break

        # Prefer the first available source context for summary metadata.
        preferred_context = next(iter(joint_context_by_source.values()), None)
        if preferred_context and preferred_context.get("target"):
            target_gwd = preferred_context["target"].get("gwd_number")
            target_longseam_hours = preferred_context["target"].get("longseam_hours")
        else:
            for gw in girth_sorted:
                ch = gw.get("chainage")
                if ch is not None and abs(ch) < 0.001:
                    gwd_val = gw.get("gwd_number")
                    gwd_num = _normalize_gwd_number(gwd_val)
                    if gwd_num is not None:
                        target_gwd = gwd_num
                        seam_map = next(iter(gwd_to_seam_by_source.values()), {}) if gwd_to_seam_by_source else {}
                        if target_gwd in seam_map:
                            target_longseam_hours = seam_map[target_gwd]
                        elif gwd_to_seam:
                            target_longseam_hours = next(iter(gwd_to_seam.values()), None)
                        break

        if seam_welds:
            scatter_data = scatter_data or {}
            scatter_data["seam_welds"] = seam_welds
            logger.info(f"[Joint Summary] Built {len(seam_welds)} per-span seam welds (longseam changes at each GWD)")

    # Fallback: when < 2 girth welds OR no per-span seam welds built, draw single blue line at target GWD longseam
    has_seam_welds = bool(scatter_data and scatter_data.get("seam_welds"))
    if not has_seam_welds and (gwd_to_seam_by_source or gwd_to_seam) and scatter_data:
        fallback_longseam = None
        fallback_gwd = target_gwd
        if gwd_to_seam_by_source:
            gwd_order = sorted(next(iter(gwd_to_seam_by_source.values()), {}).keys())
            fallback_gwd = gwd_order[len(gwd_order) // 2] if gwd_order else None
            if fallback_gwd is not None:
                fallback_longseam = next(iter(gwd_to_seam_by_source.values()), {}).get(fallback_gwd)
        if fallback_longseam is None and gwd_to_seam:
            fallback_longseam = next(iter(gwd_to_seam.values()), None)
        if fallback_longseam is not None:
            target_gwd = fallback_gwd
            target_longseam_hours = fallback_longseam
            scatter_data["seam_welds"] = [{
                "chainage_start": None,
                "chainage_end": None,
                "orientation_hours": fallback_longseam,
                "orientation_label": format_orientation_hours(fallback_longseam),
                "source": next(iter(gwd_to_seam_by_source.keys()), "") if gwd_to_seam_by_source else "",
            }]
            logger.info(f"[Joint Summary] Fallback: single blue line at {format_orientation_hours(fallback_longseam)} (target GWD {fallback_gwd})")

    # Ensure x_column reflects Distance from TGW when used
    if scatter_data and dist_col:
        scatter_data["x_column"] = dist_col
        if joint_context_by_source:
            scatter_data["joint_context_by_source"] = joint_context_by_source
        # Expose joint lengths and GWD order from the Joint Summary matrix so the
        # visualisation layer can label joints and compute correct axis spans.
        if parsed_matrix:
            jl = parsed_matrix.get("joint_lengths", {})
            if jl:
                scatter_data["joint_lengths_by_gwd"] = jl
            go_order = parsed_matrix.get("gwd_order", [])
            if go_order:
                scatter_data["gwd_order"] = go_order

    column_mapping = {k: v for k, v in ili_cols.items() if v}

    # Build feature_summary_raw for data tracing (which Excel cells/columns produced each value)
    feature_summary_raw = None
    if feature_df is not None and not feature_df.empty:
        def _json_safe(val):
            if pd.isna(val):
                return None
            try:
                f = float(val)
                return int(f) if f == int(f) else f
            except (TypeError, ValueError):
                return str(val)

        sample = feature_df.head(50)
        sample_rows = [
            {k: _json_safe(v) for k, v in row.items()}
            for row in sample.to_dict(orient="records")
        ]
        feature_summary_raw = {
            "sheet": metadata.get("feature_sheet"),
            "header_row": metadata.get("feature_header_row"),
            "columns": list(feature_df.columns),
            "column_mapping_used": {k: v for k, v in ili_cols.items() if v},
            "sample_rows": sample_rows,
        }
        # Merge target GWD longseam into Feature Summary (single blue line, no separate Joint Summary)
        if target_gwd is not None:
            feature_summary_raw["target_gwd"] = target_gwd
        if target_longseam_hours is not None:
            feature_summary_raw["target_longseam_hours"] = target_longseam_hours
            feature_summary_raw["target_longseam_label"] = format_orientation_hours(target_longseam_hours)
        if joint_context_by_source:
            feature_summary_raw["joint_context_by_source"] = joint_context_by_source

    return features, scatter_data, sources, column_mapping, joint_summary_parsed, feature_summary_raw


