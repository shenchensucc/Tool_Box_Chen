import io
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

from backend.logging_config import get_logger

logger = get_logger("backend.pipeline.ili_reader")

# Version string - change when fixing column matching; appears in logs on reload
ILI_READER_VERSION = "v3-early-exit"
logger.info(f"ili_reader loaded ({ILI_READER_VERSION})")

# Configurable keywords for column identification
# Users can manually update these lists to support different ILI vendors.
# Note: new data-format-specific keys (ds_distance, wall_thickness, pipe_grade)
# return None for anomaly files — harmless — and are used only by the pipe-tally builder.
COLUMN_KEYWORDS = {
    "depth": [
        "depth", "defect depth", "Max. Depth", "dimp", "depth (%)", "depth (mm)",
        "Peak Depth", "Peak Depth (% WT)", "Feature Depth", "Max Depth",
        "Max. Depth (%)", "Depth (%)", "As-Reported Anomaly Depth (%WT)",
        "Feature Depth (%WT for Corrosion & Cracks, %OD for Dents)"
    ],
    "length": [
        "length", "Length", "defect length", "Limp", "length (mm)",
        "Feature Length", "Feature Length (mm)", "Length (in)", "Length (in.)",
        # Pipe-tally joint lengths (matched here so basic header scoring still works)
        "Joint Length", "Joint Length (m)", "Pipe Length", "Pipe Length (m)", "Jt Length",
    ],
    "width": [
        "width", "Width", "defect width", "Wimp", "width (mm)",
        "Feature Width", "Feature Width (mm)", "Width (in)", "Width (in.)"
    ],
    "distance": [
        "distance", "Distance", "Odometer", "Log Distance", "Chainage",
        "Wheel Count (ft)", "Wheel Count (ft.)", "Log Dist.", "ILI Chainage (m)",
        "Odometer (m)", "ILI Chainage/Odometer (m)", "ILI Chainage", "ILI Distance (m)",
        "Distance from TGW (m)", "Distance from TGW",
        # Pipe-tally US-end chainage variants (treated as the primary x position)
        "US GWD ILI Chainage (m)", "US GWD ILI Chainage", "US GWD Chainage (m)",
        "US ILI Chainage (m)", "US Chainage (m)", "US Chainage",
        "US Odometer (m)", "US Odometer", "US Odo (m)", "US Odo",
        "U/S GWD ILI Chainage (m)", "U/S GWD Chainage (m)", "U/S Chainage (m)",
        "U/S Odometer (m)", "U/S Odometer", "U/S Odo (m)", "U/S Odo",
        "Upstream Chainage (m)", "Upstream Odometer (m)", "Upstream Odo (m)",
    ],
    "feature_id": [
        "feature id", "feature", "id", "f_id", "Feature Number", "Anomaly ID",
        "ID", "FeatureID", "Target ID", "ID#", "Feature Identifier", "ILI Feature ID"
    ],
    "feature_type": ["Feature Type", "Feature", "Event", "Anomaly Type", "Type"],
    "feature_desc": ["Feature Description", "Description", "Anomaly Description", "Anomaly", "Event"],
    "orientation": [
        "Orientation", "Clock Orientation", "O'Clock", "Orientation (clock)",
        "Clock Orient.", "Orientation (hh:mm)", "Feature Orientation",
        "Feature Orientation (Center of feature) (hh:mm)", "Feature Orientation (deg. or clock)",
        "(Degree)", "o'clock",
        # Pipe-tally seam / longseam orientation variants
        "Seam Orientation", "Seam Orient.", "Seam Position",
        "Longseam Orientation", "Longseam Orient.", "Longseam (o'clock)",
        "LS Orientation", "LS Orient.", "LS Position", "Longseam Position",
    ],
    "joint_number": [
        "Joint", "Joint Number", "Weld Number", "Joint No. or US GW No.",
        "Joint No", "US GW No", "PNG Joint Number", "Client Jno.",
        # Pipe-tally GWD / joint number variants
        "GWD No.", "GWD No", "GWD #", "GWD#", "GWD",
        "US GWD", "U/S GWD", "PNG GWD", "PNG GWD No.", "Client GWD No.",
        "Weld No.", "Weld No", "Jt No.", "Jt No",
    ],
    "source": [
        "ILI source", "Source", "ILI Vendor", "Vendor", "ILI Source"
    ],
    # ── Pipe-tally-specific keys ──────────────────────────────────────────
    # These are ignored by the anomaly builder; only the pipe-tally builder reads them.
    "ds_distance": [
        "DS GWD ILI Chainage (m)", "DS GWD ILI Chainage", "DS GWD Chainage (m)",
        "DS ILI Chainage (m)", "DS Chainage (m)", "DS Chainage",
        "DS Odometer (m)", "DS Odometer", "DS Odo (m)", "DS Odo",
        "D/S GWD ILI Chainage (m)", "D/S GWD Chainage (m)", "D/S Chainage (m)",
        "D/S Odometer (m)", "D/S Odometer", "D/S Odo (m)", "D/S Odo",
        "Downstream Chainage (m)", "Downstream Odometer (m)", "Downstream Odo (m)",
    ],
    "wall_thickness": [
        "Wall Thickness", "Wall Thickness (mm)", "Nom. WT", "Nom. WT (mm)",
        "Nominal WT", "Nominal WT (mm)", "Nominal Wall Thickness (mm)",
        "WT", "WT (mm)", "Thickness (mm)", "Wall Thk", "Wall Thk (mm)",
    ],
    "pipe_grade": [
        "Grade", "Pipe Grade", "Material Grade", "Mat. Grade",
        "API Grade", "Specification", "Pipe Specification",
        "Material Specification", "Mat. Spec.", "Grade/Spec",
    ],
}

def find_column_names(df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
    """
    Find a column name in the DataFrame that matches one of the possible names (case-insensitive).
    Tries exact match first, then substring match (longest keywords first) for robustness.
    """
    df_columns_lower = [str(col).lower().strip() for col in df.columns]
    # Try longest keywords first so "As-Reported Anomaly Depth (%WT)" matches before "depth"
    sorted_names = sorted(possible_names, key=lambda n: len(n), reverse=True)
    for name in sorted_names:
        name_lower = name.lower().strip()
        if name_lower in df_columns_lower:
            idx = df_columns_lower.index(name_lower)
            return df.columns[idx]
    # Fallback: substring match (column contains keyword) for vendor-specific names
    for name in sorted_names:
        name_lower = name.lower().strip()
        if len(name_lower) < 4:
            continue  # Skip very short keywords to avoid false matches
        for i, col_lower in enumerate(df_columns_lower):
            if name_lower in col_lower:
                return df.columns[i]
    return None

def identify_ili_columns(df: pd.DataFrame, custom_keywords: Optional[Dict[str, List[str]]] = None) -> Dict[str, Optional[str]]:
    """
    Automatically identify ILI columns based on keywords.
    
    Returns a dictionary mapping standardized keys (depth, length, etc.) 
    to the actual column names found in the DataFrame.
    """
    keywords = custom_keywords or COLUMN_KEYWORDS
    results = {}
    for key, possible_names in keywords.items():
        results[key] = find_column_names(df, possible_names)
    found = {k: v for k, v in results.items() if v is not None}
    logger.debug(f"identify_ili_columns: Found {len(found)} columns: {found}")
    return results


def merge_keyword_sets(
    base_keywords: Dict[str, List[str]],
    override_keywords: Optional[Dict[str, List[str]]] = None,
) -> Dict[str, List[str]]:
    """
    Merge keyword maps while preserving order and de-duplicating values.
    Override keywords are prepended so vendor/file-specific names win first.
    """
    merged: Dict[str, List[str]] = {key: list(values) for key, values in base_keywords.items()}
    for key, values in (override_keywords or {}).items():
        combined = list(values) + merged.get(key, [])
        seen = set()
        deduped = []
        for value in combined:
            norm = str(value).strip().lower()
            if not norm or norm in seen:
                continue
            seen.add(norm)
            deduped.append(value)
        merged[key] = deduped
    return merged


def _normalize_excel_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def select_workbook_sheet(workbook, keywords: Optional[List[str]] = None) -> str:
    """
    Select the best worksheet by exact match first, then substring match.
    Falls back to the first sheet when no keyword matches are found.
    """
    if not keywords:
        return workbook.sheetnames[0]

    normalized_keywords = [_normalize_excel_text(keyword) for keyword in keywords if _normalize_excel_text(keyword)]
    exact_matches = []
    partial_matches = []

    for sheet_name in workbook.sheetnames:
        normalized_sheet = _normalize_excel_text(sheet_name)
        if normalized_sheet in normalized_keywords:
            exact_matches.append(sheet_name)
            continue
        if any(keyword in normalized_sheet or normalized_sheet in keyword for keyword in normalized_keywords):
            partial_matches.append(sheet_name)

    if exact_matches:
        return exact_matches[0]
    if partial_matches:
        return partial_matches[0]
    return workbook.sheetnames[0]


def _score_row_against_keywords(headers: List[str], keyword_map: Dict[str, List[str]]) -> int:
    """
    Fast in-place scoring of a row against keyword_map without allocating a DataFrame.
    Returns the count of keyword groups with at least one match.
    """
    headers_lower = [h.lower().strip() for h in headers]
    score = 0
    for possible_names in keyword_map.values():
        for name in possible_names:
            name_lower = name.lower().strip()
            if name_lower in headers_lower:
                score += 1
                break
            if len(name_lower) >= 4 and any(name_lower in h for h in headers_lower):
                score += 1
                break
    return score


def detect_header_row(
    worksheet,
    keyword_map: Dict[str, List[str]],
    max_scan_rows: int = 60,
    min_matches: int = 2,
    early_exit_score: int = 5,
) -> Optional[int]:
    """
    Detect the most likely header row by scoring rows against expected column keywords.

    Uses a fast pre-score pass to avoid creating DataFrames for every row.
    Exits early once a row with early_exit_score matched columns is found — this
    covers the typical case where the header is found in the first 10 rows.
    """
    best_row = None
    best_score = -1
    best_non_empty = -1
    max_row = min(max_scan_rows, worksheet.max_row)

    for row_idx in range(1, max_row + 1):
        row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        headers = [str(value).strip() if value is not None else f"_col{i}" for i, value in enumerate(row)]
        non_empty = sum(1 for value in row if _normalize_excel_text(value))
        if non_empty < 2:
            continue

        # Fast pre-score — no DataFrame allocation.
        score = _score_row_against_keywords(headers, keyword_map)

        if score < min_matches:
            continue
        if score > best_score or (score == best_score and non_empty > best_non_empty):
            best_row = row_idx
            best_score = score
            best_non_empty = non_empty

        # Early exit: we found a strong enough header row — no need to scan more.
        if best_score >= early_exit_score:
            break

    return best_row


def read_worksheet_as_dataframe(worksheet, header_row: int) -> pd.DataFrame:
    """
    Read a worksheet into a DataFrame using the provided header row.
    Stops after three consecutive empty rows once data has started.
    """
    header_cells = list(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    headers = [str(value).strip() if value is not None else f"_col{i}" for i, value in enumerate(header_cells)]

    data = []
    empty_streak = 0
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        row_values = list(row[: len(headers)])
        if len(row_values) < len(headers):
            row_values.extend([None] * (len(headers) - len(row_values)))

        if not any(_normalize_excel_text(value) for value in row_values):
            if data:
                empty_streak += 1
                if empty_streak >= 3:
                    break
            continue

        empty_streak = 0
        data.append(row_values)

    return pd.DataFrame(data, columns=headers)


def read_excel_with_detected_header(
    file_content: bytes,
    keyword_map: Dict[str, List[str]],
    sheet_keywords: Optional[List[str]] = None,
    max_scan_rows: int = 60,
    min_matches: int = 2,
    good_enough_score: int = 6,
    max_sheets_to_scan: int = 10,
) -> tuple[pd.DataFrame, Dict[str, Optional[str]], str, int]:
    """
    Read an Excel worksheet using keyword-based sheet selection and header detection.

    Performance notes:
    - Tries the keyword-selected primary sheet first. If it scores >= good_enough_score
      the search stops immediately — no other sheets are scanned. This is the typical
      path for well-structured vendor files.
    - Falls back to scanning up to max_sheets_to_scan additional sheets only when the
      primary sheet scores below good_enough_score.
    - detect_header_row uses a fast pre-score (no DataFrame per row) and early-exits
      once it hits early_exit_score matched columns.

    Returns:
        (df, column_mapping, sheet_name, header_row)
    """
    workbook = load_workbook(io.BytesIO(file_content), data_only=True)
    try:
        primary_sheet = select_workbook_sheet(workbook, sheet_keywords)
        # Always try the primary (keyword-selected) sheet first.
        remaining = [name for name in workbook.sheetnames if name != primary_sheet]
        candidate_sheet_names = [primary_sheet] + remaining[:max_sheets_to_scan - 1]

        best_result = None
        best_score = -1
        for sheet_name in candidate_sheet_names:
            worksheet = workbook[sheet_name]
            header_row = detect_header_row(
                worksheet,
                keyword_map=keyword_map,
                max_scan_rows=max_scan_rows,
                min_matches=min_matches,
            )
            if header_row is None:
                continue

            df = read_worksheet_as_dataframe(worksheet, header_row)
            mapping = identify_ili_columns(df, keyword_map)
            score = len([col for col in mapping.values() if col is not None])
            if score > best_score:
                best_result = (df, mapping, sheet_name, header_row)
                best_score = score

            # Primary sheet scored well enough — stop immediately, don't scan siblings.
            if best_score >= good_enough_score:
                break

        if best_result is not None:
            return best_result

        fallback_worksheet = workbook[primary_sheet]
        fallback_header_row = 1
        fallback_df = read_worksheet_as_dataframe(fallback_worksheet, fallback_header_row)
        fallback_mapping = identify_ili_columns(fallback_df, keyword_map)
        return fallback_df, fallback_mapping, primary_sheet, fallback_header_row
    finally:
        workbook.close()

def parse_pasted_ili_text(text: str) -> pd.DataFrame:
    """
    Parse pasted tabular text (e.g. from Excel copy) into a DataFrame.
    Tries tab separator first (Excel default), then comma.
    """
    import io
    text = text.strip()
    if not text:
        logger.warning("parse_pasted_ili_text: empty input")
        return pd.DataFrame()

    for sep, name in [("\t", "tab"), (",", "comma")]:
        try:
            df = pd.read_csv(io.StringIO(text), sep=sep, dtype=str)
            if df.shape[1] > 1:
                logger.info(f"parse_pasted_ili_text: parsed {df.shape} with {name} separator")
                return df
        except Exception as e:
            logger.debug(f"parse_pasted_ili_text: {name} sep failed: {e}")
            continue
    logger.warning("parse_pasted_ili_text: could not parse input")
    return pd.DataFrame()


def detect_data_format(df: pd.DataFrame, mapping: Dict[str, Optional[str]]) -> str:
    """
    Detect the data format from a column mapping and DataFrame.

    Returns one of the registered format names (currently ``"anomaly"`` or
    ``"pipe_tally"``).  New formats can be added by extending this function.

    Detection heuristics (higher specificity wins):

    * **pipe_tally**: has a DS-distance column  OR  has wall-thickness + joint
      number but no depth column  OR  the column names contain common pipe-tally
      keywords and there is no anomaly-style depth column.
    * **anomaly**: has a depth column or feature-type column (default).
    """
    has_ds = bool(mapping.get("ds_distance"))
    has_wt = bool(mapping.get("wall_thickness"))
    has_grade = bool(mapping.get("pipe_grade"))
    has_joint = bool(mapping.get("joint_number"))
    has_depth = bool(mapping.get("depth"))
    has_ftype = bool(mapping.get("feature_type"))

    # Heuristic scan of raw column names for pipe-tally keywords
    col_str = " ".join(str(c).lower() for c in df.columns)
    tally_col_hints = any(
        kw in col_str
        for kw in [
            "wall thickness", "wt (mm)", "nom. wt", "ds chainage",
            "ds odometer", "ds odo", "d/s chainage", "d/s odo",
        ]
    )

    if has_ds:
        return "pipe_tally"
    if (has_wt or has_grade) and has_joint and not has_depth:
        return "pipe_tally"
    if tally_col_hints and has_joint and not has_depth and not has_ftype:
        return "pipe_tally"
    return "anomaly"


def read_ili_data(file_path_or_buffer, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Read an ILI dataset from an Excel file.
    """
    logger.debug(f"read_ili_data: sheet_name={sheet_name}, input_type={type(file_path_or_buffer).__name__}")
    try:
        df = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name)
    except Exception as e:
        logger.error(f"read_ili_data failed: {type(e).__name__}: {e}")
        raise

    # If the result is a dictionary (multiple sheets), take the first sheet
    if isinstance(df, dict):
        if not df:
            logger.warning("read_ili_data: Excel returned empty dict of sheets")
            return pd.DataFrame()
        first_sheet = list(df.keys())[0]
        df = df[first_sheet]
        logger.debug(f"read_ili_data: Using first sheet '{first_sheet}', shape={df.shape}")

    # Header-row auto-detection: some workbooks (e.g. IDP master anomaly lists)
    # carry 1-3 banner/metadata rows above the real column header, so header=0
    # yields "Unnamed: N" columns and depth/length detection fails.  When that
    # happens, scan the first rows for one containing both a depth-like and a
    # length-like label and re-header the frame from there.
    cols = identify_ili_columns(df)
    if not cols.get("depth") or not cols.get("length"):
        try:
            if hasattr(file_path_or_buffer, "seek"):
                file_path_or_buffer.seek(0)
            raw = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name, header=None)
            if isinstance(raw, dict):
                raw = raw[list(raw.keys())[0]]
            for hdr_row in range(min(10, len(raw))):
                row_text = [str(v).lower() for v in raw.iloc[hdr_row].tolist() if pd.notna(v)]
                has_depth = any("depth" in t for t in row_text)
                has_length = any("length" in t for t in row_text)
                if has_depth and has_length:
                    candidate = raw.iloc[hdr_row + 1:].copy()
                    candidate.columns = raw.iloc[hdr_row].tolist()
                    candidate = candidate.reset_index(drop=True)
                    if identify_ili_columns(candidate).get("depth"):
                        logger.info(
                            f"read_ili_data: header auto-detected on row {hdr_row} "
                            f"(0-indexed); {len(candidate)} data rows."
                        )
                        df = candidate
                        break
        except Exception as e:
            logger.warning(f"read_ili_data: header auto-detection failed, keeping original frame: {e}")

    logger.info(f"read_ili_data: Loaded shape={df.shape}, columns={list(df.columns)[:10]}...")
    return df
