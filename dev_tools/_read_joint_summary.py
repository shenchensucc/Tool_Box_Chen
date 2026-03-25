import openpyxl

path = (
    r"c:\Users\cshen\trisummit\PNG - Asset Integrity - Documents\General"
    r"\003_TIMP\003_West_ML\000_Common\IDP\2026_Planning"
    r"\09-Dig Notifications and packages\ID216\Dig Packages"
    r"\ID6006_R1R2_MP31_NPS10_GW30930_ML_DP_R0.xlsx"
)

wb = openpyxl.load_workbook(path, data_only=True)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print(f"\n=== Sheet: {sheet_name} ({len(rows)} rows) ===")

    # Find the Joint Summary section
    joint_start = None
    for i, row in enumerate(rows):
        row_text = " ".join(str(c) for c in row if c is not None).strip().lower()
        if "joint" in row_text and "summary" in row_text:
            joint_start = i
            print(f"  --> Joint Summary header at row {i+1}: {' '.join(str(c) for c in row if c is not None)[:120]}")
            break

    if joint_start is not None:
        print(f"\n  --- Rows {joint_start+1} to {min(joint_start+40, len(rows))} ---")
        for idx in range(joint_start, min(joint_start + 40, len(rows))):
            vals = rows[idx]
            non_none = [c for c in vals if c is not None]
            if non_none:
                print(f"  [{idx+1:3d}]", vals[:20])
    else:
        # Print first 5 rows as sample
        print("  (No Joint Summary found — first 5 rows:)")
        for idx in range(min(5, len(rows))):
            print(f"  [{idx+1:3d}]", rows[idx][:20])
