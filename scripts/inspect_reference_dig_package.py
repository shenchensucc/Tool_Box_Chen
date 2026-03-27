"""
Dump Excel structure for Reference dig package folders (local paths).
Run from repo root:

  python scripts/inspect_reference_dig_package.py

Default base: C:\\Users\\cshen\\Documents\\Reference dig package
Override: python scripts/inspect_reference_dig_package.py "D:\\path\\Reference dig package"

Writes: reference_dig_package/STRUCTURE_REPORT.md (gitignored data stays local).
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
REPORT_PATH = ROOT / "reference_dig_package" / "STRUCTURE_REPORT.md"


def dump_workbook(path: Path, max_rows: int = 25) -> list[str]:
    lines: list[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [f"- (openpyxl missing) {path.name}\n"]

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [f"- **ERROR** reading `{path.name}`: {e}\n"]

    lines.append(f"### `{path.name}`\n")
    lines.append(f"- Path: `{path}`\n")
    lines.append(f"- Sheets: {wb.sheetnames}\n")
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True, max_row=max_rows))
        lines.append(f"\n**Sheet `{sn}`** — first {len(rows)} rows (up to {max_rows}):\n")
        for i, row in enumerate(rows, 1):
            preview = [str(c)[:80] if c is not None else "" for c in (row[:15] if row else [])]
            lines.append(f"  {i}: {preview}\n")
        lines.append("\n")
    try:
        wb.close()
    except Exception:
        pass
    return lines


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "base",
        nargs="?",
        default=r"C:\Users\cshen\Documents\Reference dig package",
        help="Folder containing 1-Input dig list, 2-Input ILI, 3-Exported dig package",
    )
    args = parser.parse_args()
    base = Path(args.base)
    if not base.is_dir():
        print(f"Not a directory: {base}", file=sys.stderr)
        return 1

    sub = {
        "1_input_dig_list": base / "1-Input dig list",
        "2_input_ili": base / "2-Input ILI",
        "3_exported": base / "3-Exported dig package",
    }

    out: list[str] = []
    out.append("# Reference dig package — structure dump\n\n")
    out.append(f"Base: `{base}`\n\n")

    for label, folder in sub.items():
        out.append(f"## {label}\n\n")
        if not folder.is_dir():
            out.append(f"_Missing: {folder}_\n\n")
            continue
        xlsx = sorted(folder.glob("*.xlsx")) + sorted(folder.glob("*.xlsm"))
        if not xlsx:
            out.append("_No Excel files found._\n\n")
            continue
        for p in xlsx:
            out.extend(dump_workbook(p))
            out.append("\n---\n\n")

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text("".join(out), encoding="utf-8")
    print(f"Wrote {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
