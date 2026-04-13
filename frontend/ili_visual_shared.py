import asyncio
import base64
import hashlib
from typing import Literal, Optional

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as st_components

from frontend_utils import (
    BACKEND_URL,
    DIG_PACKAGE_ILI_FORMAT_OPTIONS,
    call_excel_to_pdf_api,
    call_parse_paste_api,
    call_preview_api,
    call_process_dig_package_api,
    call_process_feature_map_api,
    fu_key,
)

# NPS (Nominal Pipe Size) to OD (mm) - ASME/ISO standard
NPS_TO_OD_MM = {
    4: 114.3,
    5: 141.3,
    6: 168.3,
    8: 219.1,
    10: 273.0,
    12: 323.8,
    14: 355.6,
    16: 406.4,
    18: 457.0,
    20: 508.0,
    24: 610.0,
    28: 711.0,
    32: 813.0,
}

# RGB tuples for multi-source longseam lines (2D and 3D views).
# Index matches the alphabetical position of the joint-summary source name.
_SEAM_SOURCE_COLORS: list[tuple[int, int, int]] = [
    (0, 100, 220),    # blue   — primary Rosen / MFL-A type
    (160, 0, 220),    # purple — TDW / secondary tool
    (0, 160, 150),    # teal   — EMAT / 3rd tool
    (210, 90,  0),    # orange — 4th source
    (0, 160,  60),    # green  — 5th source
]


def _seam_src_rgba(source: str, sorted_sources: list[str], alpha: float = 1.0) -> str:
    """Return rgba() colour string for a joint-summary source name."""
    try:
        idx = sorted_sources.index(source)
    except ValueError:
        idx = 0
    r, g, b = _SEAM_SOURCE_COLORS[idx % len(_SEAM_SOURCE_COLORS)]
    return f"rgba({r},{g},{b},{alpha})"


def _seam_trace_alpha(sw: dict, filter_by_source: bool, source_filter: set) -> float:
    """Dim longseam lines when joint source does not match the active filter (2D/3D)."""
    if not filter_by_source or not source_filter:
        return 1.0
    src = sw.get("source", "")
    feat_src = sw.get("feature_source", "")
    for s in source_filter:
        if s in src or src in s or s in feat_src or feat_src in s:
            return 1.0
    return 0.30


# RGB values for depth colors (for rgba with alpha)
_DEPTH_RGB = {
    "lightgrey": (211, 211, 211),
    "green": (34, 139, 34),
    "yellow": (255, 255, 0),
    "orange": (255, 165, 0),
    "red": (220, 20, 60),
}

# Reset axes / double-click return to the layout ranges set on the figure (feature x extent, fixed y).
_FEATURE_MAP_PLOTLY_CONFIG: dict = {"doubleClick": "reset"}


def _depth_color(d, alpha: float = 1.0):
    """Map depth % to color for feature boxes. alpha=1.0 full, 0.2 for 20% visible."""
    if d is None or (isinstance(d, float) and pd.isna(d)):
        name = "lightgrey"
    elif d < 20:
        name = "green"
    elif d < 40:
        name = "yellow"
    elif d < 60:
        name = "orange"
    else:
        name = "red"
    r, g, b = _DEPTH_RGB.get(name, _DEPTH_RGB["lightgrey"])
    return f"rgba({r},{g},{b},{alpha})"


def _get_x_axis_title(x_column: str) -> str:
    """Return appropriate x-axis label. Uses Distance from TGW when chainage not available."""
    xc = (x_column or "").lower()
    if "distance" in xc and "tgw" in xc:
        return "Distance from TGW (m)"
    if "chainage" in xc or "odometer" in xc or "distance" in xc:
        return "ILI Chainage (m)"
    return f"{x_column} (m)" if x_column else "Distance (m)"


def _format_joint_context_summary(joint_context: dict | None) -> str:
    if not joint_context:
        return ""
    parts = []
    role_labels = [("upstream", "U/S"), ("target", "Target"), ("downstream", "D/S")]
    for role_key, role_label in role_labels:
        item = joint_context.get(role_key)
        if not item:
            continue
        gwd = item.get("gwd_number")
        seam = item.get("longseam_label")
        if gwd is None:
            continue
        text = f"{role_label} GWD {gwd}"
        if seam:
            text += f" @ {seam}"
        parts.append(text)
    joint_source = joint_context.get("joint_source")
    if joint_source and parts:
        return f"{joint_source}: " + " | ".join(parts)
    return " | ".join(parts)


def _build_feature_map_figure(
    features: list,
    scatter_data: dict,
    pipe_circ_mm: float,
    source_filter: set,
    filter_by_source: bool,
    x_column: str,
    title: str,
    height: int = 450,
    use_opacity_overlay: bool = False,
) -> tuple:
    """
    Build a Plotly figure for the feature map. Returns (fig, filtered_count).
    When use_opacity_overlay=True, deselected sources are shown at 20% opacity instead of hidden.

    Initial x-axis range uses defect box extents only (chainage ± half length), when available.
    """

    def _is_selected(f_or_gw):
        # Joint Summary TGW layout girth welds are structural (same for all ILI sources).
        if f_or_gw.get("joint_summary_layout"):
            return True
        src = f_or_gw.get("feature_source") or f_or_gw.get("source", "") or ""
        if not filter_by_source:
            return True
        if not source_filter:
            return True
        if src in source_filter:
            return True
        # Seam welds use Joint Summary source (e.g. "2022 Rosen"); features use full source (e.g. "2022 Rosen MFL-A")
        # Match when one contains the other (Rosen MFL-A selected -> show 2022 Rosen seam line)
        for s in source_filter:
            if src in s or s in src:
                return True
        return False

    def _alpha(f_or_gw):
        return 0.2 if use_opacity_overlay and not _is_selected(f_or_gw) else 1.0

    # Overlay mode: show all; filter mode: show only selected
    if use_opacity_overlay:
        plot_features = [
            f
            for f in features
            if "girth" not in (f.get("feature_type") or "").lower()
            and "seam" not in (f.get("feature_type") or "").lower()
            and "gwd" not in (f.get("feature_type") or "").lower()
        ]
        girth_list = scatter_data.get("girth_welds", [])
        selected_count = sum(1 for f in features if _is_selected(f))
    else:
        filtered_features = [f for f in features if _is_selected(f)]
        plot_features = [
            f
            for f in filtered_features
            if "girth" not in (f.get("feature_type") or "").lower()
            and "seam" not in (f.get("feature_type") or "").lower()
            and "gwd" not in (f.get("feature_type") or "").lower()
        ]
        girth_list = [gw for gw in scatter_data.get("girth_welds", []) if _is_selected(gw)]
        selected_count = len(filtered_features)

    # Seam welds: always show ALL sources for cross-ILI comparison (longseam is
    # a structural property, not a defect feature to be filtered out).
    seam_list = scatter_data.get("seam_welds", [])
    all_seam_sources: list[str] = sorted(set(sw.get("source", "") for sw in seam_list))

    x_values = scatter_data.get("x_values", [f["x"] for f in features])
    orient_hours = scatter_data.get("orientation_hours")
    use_orientation = orient_hours and len(orient_hours) == len(features)
    y_values = orient_hours if use_orientation else [f["y"] for f in features]
    joint_context_by_source = scatter_data.get("joint_context_by_source", {})

    fig = go.Figure()

    nde_x_bounds: Optional[tuple[float, float]] = None
    nde_reg = scatter_data.get("nde_region") if scatter_data else None
    if nde_reg and nde_reg.get("x0") is not None and nde_reg.get("x1") is not None:
        nx0 = float(nde_reg["x0"])
        nx1 = float(nde_reg["x1"])
        if nx0 > nx1:
            nx0, nx1 = nx1, nx0
        nde_x_bounds = (nx0, nx1)
        fig.add_vrect(
            x0=nx0,
            x1=nx1,
            fillcolor="rgba(100, 100, 100, 0.22)",
            layer="below",
            line_width=0,
        )

    # Draw deselected first (lower layer), then selected (on top) for clearer overlap
    for layer_alpha, layer_features in [
        (0.2, [f for f in plot_features if not _is_selected(f)]),
        (1.0, [f for f in plot_features if _is_selected(f)]),
    ]:
        for f in layer_features:
            cx = f["x"]
            cy = f.get("orientation_hours", 6.0) if use_orientation else f["y"]
            ln_mm = max(f.get("length", 0) or 0.001, 0.001)
            wd_mm = max(f.get("width", 0) or 0.001, 0.001)
            ln_m = ln_mm / 1000
            wd_hr = (wd_mm / pipe_circ_mm) * 12
            x0, x1 = cx - ln_m / 2, cx + ln_m / 2
            y0, y1 = cy - wd_hr / 2, cy + wd_hr / 2
            alpha = layer_alpha if use_opacity_overlay else 1.0
            color = _depth_color(f.get("depth"), alpha=alpha)
            line_alpha = alpha
            fig.add_shape(
                type="rect",
                x0=x0,
                x1=x1,
                y0=y0,
                y1=y1,
                line=dict(color=f"rgba(0,0,0,{line_alpha})", width=2.5),
                fillcolor=color,
            )

    for gw in girth_list:
        a = _alpha(gw)
        fig.add_vline(x=gw.get("chainage"), line_color=f"rgba(220,20,60,{a})", line_width=2.5)

    if nde_x_bounds is not None:
        nxa, nxb = nde_x_bounds
        for xv in (nxa, nxb):
            fig.add_vline(
                x=xv,
                line_width=2,
                line_dash="dash",
                line_color="rgba(45, 45, 45, 0.9)",
            )

    def _feature_xaxis_range() -> Optional[tuple[float, float]]:
        """Defect boxes only: chainage ± half length."""
        edges: list[float] = []
        for f in plot_features:
            try:
                cx = float(f["x"])
            except (TypeError, ValueError):
                continue
            hl = max((f.get("length") or 0) / 1000.0, 0.001) / 2.0
            edges.extend([cx - hl, cx + hl])
        if not edges:
            return None
        lo, hi = min(edges), max(edges)
        span = max(hi - lo, 0.5)
        pad = max(span * 0.05, 0.3)
        return (lo - pad, hi + pad)

    def _feature_centers_xaxis_range() -> Optional[tuple[float, float]]:
        """Fallback: defect chainage centres only (no box half-width)."""
        xs: list[float] = []
        for f in plot_features:
            try:
                xs.append(float(f["x"]))
            except (TypeError, ValueError):
                continue
        if not xs:
            return None
        lo, hi = min(xs), max(xs)
        span = max(hi - lo, 0.5)
        pad = max(span * 0.05, 0.3)
        return (lo - pad, hi + pad)

    def _scatter_xaxis_range() -> Optional[tuple[float, float]]:
        """Last resort: full scatter x column extent."""
        xv = scatter_data.get("x_values") or []
        if not xv:
            return None
        lo, hi = float(min(xv)), float(max(xv))
        span = max(hi - lo, 0.5)
        pad = max(span * 0.05, 0.3)
        return (lo - pad, hi + pad)

    def _merge_nde_into_x_range(rng: Optional[tuple[float, float]]) -> Optional[tuple[float, float]]:
        if nde_x_bounds is None:
            return rng
        nx0, nx1 = nde_x_bounds
        pad = max(abs(nx1 - nx0) * 0.1, 0.25)
        if rng is None:
            return (nx0 - pad, nx1 + pad)
        lo, hi = rng
        return (min(lo, nx0 - pad), max(hi, nx1 + pad))

    # Default x view for initial render and Plotly "Reset axes" (needs fixed range + autorange off).
    x_range_choice = _merge_nde_into_x_range(
        _feature_xaxis_range()
        or _feature_centers_xaxis_range()
        or _scatter_xaxis_range()
    )

    x_min_plot = min(x_values) if x_values else 0
    x_max_plot = max(x_values) if x_values else 1
    y_min_plot = min(y_values) if y_values else 0
    y_max_plot = max(y_values) if y_values else 12
    for sw in seam_list:
        if sw.get("chainage_start") is None or sw.get("chainage_end") is None:
            continue
        oh = sw.get("orientation_hours", 6.0)
        ch_start = float(sw["chainage_start"])
        ch_end = float(sw["chainage_end"])
        js_src = sw.get("source", "")
        a = _seam_trace_alpha(sw, filter_by_source, source_filter)
        color = _seam_src_rgba(js_src, all_seam_sources, a)
        lbl = sw.get("orientation_label", f"{oh:.2f}")
        hover_src = f" [{js_src}]" if js_src else ""
        gwd_h = sw.get("gwd_number")
        gwd_ht = f"GWD {gwd_h}<br>" if gwd_h is not None else ""
        fig.add_trace(
            go.Scatter(
                x=[ch_start, ch_end],
                y=[oh, oh],
                mode="lines",
                line=dict(color=color, width=2.5),
                showlegend=False,
                hovertemplate=f"<b>Longseam</b><br>{gwd_ht}{lbl}{hover_src}<extra></extra>",
            )
        )

    hover_texts = [f.get("hover_text", "") for f in plot_features]
    plot_x = [f["x"] for f in plot_features]
    plot_y = [f.get("orientation_hours", 6.0) if use_orientation else f["y"] for f in plot_features]
    fig.add_trace(
        go.Scatter(
            x=plot_x,
            y=plot_y,
            mode="markers",
            marker=dict(size=4, color="rgba(0,0,0,0)"),
            text=hover_texts,
            hoverinfo="text",
        )
    )

    if nde_x_bounds is not None:
        fig.add_trace(
            go.Scatter(
                x=[float("nan")],
                y=[float("nan")],
                mode="markers",
                marker=dict(
                    size=12,
                    color="rgba(100, 100, 100, 0.45)",
                    symbol="square",
                    line=dict(color="rgb(75, 75, 75)", width=1),
                ),
                name="NDE Area",
                showlegend=True,
                hoverinfo="skip",
            )
        )

    annotations = [
        dict(
            x=1.02,
            y=0.98,
            xref="paper",
            yref="paper",
            text="Depth: 0-20% green, 20-40% yellow, 40-60% orange, >60% red",
            showarrow=False,
            font=dict(size=9),
            xanchor="left",
        ),
        dict(
            x=1.02,
            y=0.88,
            xref="paper",
            yref="paper",
            text="Box size: length (mm) × width (mm)",
            showarrow=False,
            font=dict(size=8),
            xanchor="left",
        ),
    ]
    for gw in girth_list:
        lbl = gw.get("label", "")
        seam_p = gw.get("longseam_label_primary")
        if lbl and seam_p and "@ " not in lbl:
            lbl = f"{lbl} @ {seam_p}"
        if lbl:
            y_pos = y_min_plot + 0.7 * (y_max_plot - y_min_plot) if y_values else 6
            annotations.append(
                dict(
                    x=gw["chainage"],
                    y=y_pos,
                    text=lbl,
                    showarrow=False,
                    font=dict(size=9, color="darkred"),
                    textangle=-90,
                    xanchor="right",
                    yanchor="middle",
                )
            )
    for sw in seam_list:
        lbl = sw.get("orientation_label", "")
        ch_s, ch_e = sw.get("chainage_start"), sw.get("chainage_end")
        if not lbl or ch_s is None or ch_e is None:
            continue
        mid_x = (float(ch_s) + float(ch_e)) / 2
        js_src = sw.get("source", "")
        ann_color = _seam_src_rgba(js_src, all_seam_sources, 1.0)
        # Stack labels from multiple sources vertically so they don't overlap
        src_offset = all_seam_sources.index(js_src) * 0.28 if js_src in all_seam_sources else 0
        gwd_n = sw.get("gwd_number")
        gwd_prefix = f"GWD {gwd_n} — " if gwd_n is not None else ""
        ann_text = f"{gwd_prefix}{lbl}" if not js_src else f"{gwd_prefix}{lbl} [{js_src}]"
        annotations.append(
            dict(
                x=mid_x,
                y=sw.get("orientation_hours", 6) + 0.3 + src_offset,
                text=ann_text,
                showarrow=False,
                font=dict(size=8, color=ann_color),
                xanchor="center",
                yanchor="bottom",
            )
        )

    if nde_x_bounds is not None:
        nxa, nxb = nde_x_bounds
        x_mid_nde = (nxa + nxb) / 2.0
        ann_y_title = min(y_max_plot, 11.15) if y_max_plot > y_min_plot else 6.0
        annotations.append(
            dict(
                x=x_mid_nde,
                y=ann_y_title,
                xref="x",
                yref="y",
                text="NDE Area",
                showarrow=False,
                font=dict(size=10, color="rgb(38, 38, 38)"),
                bgcolor="rgba(220, 220, 220, 0.88)",
                bordercolor="rgba(95, 95, 95, 0.5)",
                borderwidth=1,
                borderpad=4,
                xanchor="center",
                yanchor="middle",
            )
        )
        ann_y_dist = (
            y_min_plot + 0.22 * (y_max_plot - y_min_plot) if y_max_plot > y_min_plot else 2.0
        )
        for i, xv in enumerate((nxa, nxb)):
            is_left = i == 0
            annotations.append(
                dict(
                    x=xv,
                    y=ann_y_dist,
                    xref="x",
                    yref="y",
                    text=f"{xv:.3f} m",
                    showarrow=False,
                    font=dict(size=9, color="rgb(42, 42, 42)"),
                    xanchor="right" if is_left else "left",
                    xshift=-8 if is_left else 8,
                    yanchor="middle",
                )
            )

    x_axis_title = _get_x_axis_title(x_column)
    xaxis_kwargs = dict(
        showgrid=True,
        gridcolor="rgba(0,0,0,0.2)",
        showline=True,
        linewidth=2,
        linecolor="black",
        mirror=True,
        tickformat=".3f",
        ticksuffix=" m",
    )
    if x_range_choice is not None:
        xaxis_kwargs["range"] = list(x_range_choice)
        xaxis_kwargs["autorange"] = False

    # Orientation as decimal hours: 12:48 → 12.8; axis runs 0–13 with top tick labelled 12:60.
    _y_orient_tickvals = list(range(0, 14))
    _y_orient_ticktext = [f"{h:02d}:00" for h in range(0, 13)] + ["12:60"]

    layout_extras = {}
    if nde_x_bounds is not None:
        layout_extras["showlegend"] = True
        layout_extras["legend"] = dict(
            x=1.02,
            y=0.80,
            xanchor="left",
            yanchor="top",
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="rgba(0,0,0,0.18)",
            borderwidth=1,
            font=dict(size=10),
        )

    fig.update_layout(
        title=title,
        xaxis_title=x_axis_title,
        yaxis_title="Feature Orientation (hh:mm)",
        height=height,
        hovermode="closest",
        dragmode="pan",
        template="plotly_white",
        plot_bgcolor="white",
        xaxis=xaxis_kwargs,
        **layout_extras,
        yaxis=dict(
            showgrid=True,
            gridcolor="rgba(0,0,0,0.2)",
            showline=True,
            linewidth=2,
            linecolor="black",
            mirror=True,
            tickvals=_y_orient_tickvals,
            ticktext=_y_orient_ticktext,
            range=[0, 13],
            autorange=False,
            # Box zoom / scroll zoom only change chainage (x); orientation axis stays fixed
            fixedrange=True,
        ),
        annotations=annotations,
    )
    count = len(features) if use_opacity_overlay else selected_count
    return fig, count


def _build_3d_pipeline_figure(
    features: list,
    scatter_data: dict,
    pipe_od_mm: float,
    source_filter: set,
    filter_by_source: bool,
    title: str = "3D Pipeline View — 3 Joints",
    max_joints: int = 3,
) -> go.Figure:
    """
    Build a 3D pipeline figure using the industry-standard C-scan-on-cylinder approach.

    The pipeline is a single continuous parametric cylinder:
        X = chainage (m)  — pipe longitudinal axis
        Y = R * sin(θ)    — horizontal cross-section
        Z = R * cos(θ)    — vertical cross-section (12 o'clock = Z+)

    Metal loss depth is painted onto the cylinder surface as a 2D heatmap
    (surfacecolor array = depth at each [chainage × angle] grid cell).
    Each defect footprint is projected from (chainage, orientation) onto the grid.

    Additional layers:
        - Red rings          : GWDs (girth welds) at joint boundaries
        - Coloured lines     : one longseam line per ILI source per span
                               (blue = source 1, purple = source 2, …)
        - Clock labels       : 12, 3, 6, 9 o'clock at the upstream face

    Joint selection is anchored on the target GWD (Distance from TGW ≈ 0)
    so the dig-site boundary is always visible rather than the index-middle.
    """
    pipe_r_m = pipe_od_mm / 2.0 / 1000.0          # pipe radius in metres
    pipe_circ_m = 2.0 * np.pi * pipe_r_m           # circumference in metres

    girth_welds_all = sorted(scatter_data.get("girth_welds", []), key=lambda g: g["chainage"])
    seam_welds = scatter_data.get("seam_welds", [])

    def _is_selected(f):
        src = f.get("source") or f.get("feature_source") or ""
        if not filter_by_source or not source_filter:
            return True
        if src in source_filter:
            return True
        return any(src in s or s in src for s in source_filter)

    # Defects only (exclude girth-weld and seam-weld rows)
    plot_features = [
        f for f in features
        if _is_selected(f)
        and "girth" not in (f.get("feature_type") or "").lower()
        and "gwd"   not in (f.get("feature_type") or "").lower()
        and "seam"  not in (f.get("feature_type") or "").lower()
    ]

    # ── 1. Choose which 3 joints to display ──────────────────────────────────
    # Anchor on the Target GWD (Distance from TGW ≈ 0) so the view is always
    # centred on the dig site rather than the middle of the GWD list by index.
    # Layout: 1 joint upstream of target + target GWD boundary + joints downstream.
    needed_gwds = max_joints + 1
    if len(girth_welds_all) >= needed_gwds:
        tgt_idx = min(
            range(len(girth_welds_all)),
            key=lambda k: abs(girth_welds_all[k].get("chainage") or float("inf")),
        )
        # One joint upstream of target (tgt_idx − 1); clamp so we always fit needed_gwds
        start_idx = max(0, tgt_idx - 1)
        start_idx = min(start_idx, len(girth_welds_all) - needed_gwds)
        selected_gwds = girth_welds_all[start_idx: start_idx + needed_gwds]
    elif len(girth_welds_all) >= 2:
        selected_gwds = girth_welds_all
    else:
        # No GWD data — synthesise evenly spaced joint boundaries
        all_x = [f["x"] for f in plot_features]
        if not all_x:
            return go.Figure()
        x_min_f, x_max_f = min(all_x), max(all_x)
        jlen = (x_max_f - x_min_f) / max(max_joints, 1)
        selected_gwds = [
            {"chainage": x_min_f + i * jlen, "label": f"GWD ~{i}", "gwd_number": None}
            for i in range(max_joints + 1)
        ]

    view_x_min = selected_gwds[0]["chainage"]
    view_x_max = selected_gwds[-1]["chainage"]
    x_span     = max(view_x_max - view_x_min, 1.0)

    visible_features = [
        f for f in plot_features if view_x_min - 0.1 <= f["x"] <= view_x_max + 0.1
    ]

    # ── 2. Build cylinder grid ────────────────────────────────────────────────
    # Higher resolution → smoother surface; 120 × 300 is a good balance
    N_THETA = 120    # circumferential
    N_X     = 300    # longitudinal

    theta   = np.linspace(0, 2 * np.pi, N_THETA, endpoint=False)
    x_arr   = np.linspace(view_x_min, view_x_max, N_X)

    # Meshgrid shape: (N_X, N_THETA) — row = chainage slice, col = angle slice
    THETA, X_GRID = np.meshgrid(theta, x_arr)
    Y_GRID = pipe_r_m * np.sin(THETA)   # horizontal (3 o'clock = +Y)
    Z_GRID = pipe_r_m * np.cos(THETA)   # vertical   (12 o'clock = +Z)

    # ── 3. Paint metal-loss depth onto the cylinder surface ───────────────────
    # depth_grid[i, j] = max depth (% WT) of any defect covering grid cell (i,j)
    depth_grid = np.zeros((N_X, N_THETA), dtype=float)

    for feat in visible_features:
        feat_depth = float(feat.get("depth", 0) or 0)
        if feat_depth <= 0:
            continue

        feat_x_c = feat["x"]
        feat_oh  = feat.get("orientation_hours", 6.0)
        feat_theta_c = (feat_oh / 12.0) * 2.0 * np.pi

        # Physical footprint (mm → m); enforce a visible minimum
        feat_len_m = max((feat.get("length") or 0) / 1000.0, x_span * 0.012)
        feat_wid_m = max((feat.get("width")  or 0) / 1000.0, pipe_circ_m * 0.05)

        half_len   = feat_len_m / 2.0
        half_theta = (feat_wid_m / pipe_circ_m) * np.pi  # angular half-width (rad)

        # Longitudinal mask
        x_mask = np.abs(X_GRID - feat_x_c) <= half_len

        # Angular mask with wrap-around handling
        d_theta = THETA - feat_theta_c
        d_theta = (d_theta + np.pi) % (2.0 * np.pi) - np.pi   # fold to [-π, π]
        theta_mask = np.abs(d_theta) <= half_theta

        hit = x_mask & theta_mask
        depth_grid[hit] = np.maximum(depth_grid[hit], feat_depth)

    # ── 4. Colorscale — matches the 2D feature map exactly ───────────────────
    # green 0-20 %, yellow 20-40 %, orange 40-60 %, red 60-100 %
    depth_colorscale = [
        [0.00, "rgb(210,240,210)"],   # very light green  — 0 % (bare steel)
        [0.19, "rgb(34,139,34)"],     # forest green      — ~20 %
        [0.20, "rgb(255,230,50)"],    # yellow            — 20 %
        [0.39, "rgb(255,200,0)"],     # amber             — ~40 %
        [0.40, "rgb(255,140,0)"],     # orange            — 40 %
        [0.59, "rgb(220,60,20)"],     # orange-red        — ~60 %
        [0.60, "rgb(180,0,0)"],       # dark red          — 60 %
        [1.00, "rgb(80,0,0)"],        # very dark red     — 100 %
    ]

    fig = go.Figure()

    # ── 5. Main cylinder surface with depth heatmap ───────────────────────────
    fig.add_trace(go.Surface(
        x=X_GRID,
        y=Y_GRID,
        z=Z_GRID,
        surfacecolor=depth_grid,
        colorscale=depth_colorscale,
        cmin=0,
        cmax=100,
        showscale=True,
        colorbar=dict(
            title=dict(text="Depth (% WT)", side="right", font=dict(size=12)),
            thickness=16,
            len=0.55,
            y=0.5,
            tickvals=[0, 20, 40, 60, 80, 100],
            ticktext=["0%", "20%", "40%", "60%", "80%", "100%"],
            tickfont=dict(size=10),
        ),
        opacity=0.95,
        name="Pipeline",
        showlegend=False,
        # Lighting gives the cylinder a 3-D rounded look
        lighting=dict(ambient=0.65, diffuse=0.85, specular=0.15, roughness=0.6),
        lightposition=dict(x=2, y=-3, z=3),
    ))

    # ── 6. GWD rings — red circumferential band at each joint boundary ────────
    ring_theta = np.linspace(0, 2.0 * np.pi, 180)
    ring_y = pipe_r_m * np.sin(ring_theta)
    ring_z = pipe_r_m * np.cos(ring_theta)
    first_gwd = True
    for gw in selected_gwds:
        ch  = gw["chainage"]
        lbl = gw.get("label") or (
            f"GWD {gw['gwd_number']}" if gw.get("gwd_number") is not None else "GWD"
        )
        fig.add_trace(go.Scatter3d(
            x=np.full_like(ring_theta, ch),
            y=ring_y,
            z=ring_z,
            mode="lines",
            line=dict(color="red", width=7),
            name="GWDs (Girth Welds)" if first_gwd else lbl,
            showlegend=first_gwd,
            legendgroup="gwds",
            hovertemplate=f"<b>{lbl}</b><br>Chainage: {ch:.3f} m<extra></extra>",
        ))
        # GWD label at 12 o'clock above the ring
        fig.add_trace(go.Scatter3d(
            x=[ch], y=[0.0], z=[pipe_r_m * 1.4],
            mode="text",
            text=[lbl],
            textfont=dict(size=9, color="darkred"),
            showlegend=False,
            hoverinfo="skip",
        ))
        first_gwd = False

    # ── 7. Clock-position labels on the upstream face ─────────────────────────
    CLOCK_POS = [(0, "12:00"), (3, "03:00"), (6, "06:00"), (9, "09:00")]
    label_r = pipe_r_m * 1.40
    for clock_h, clock_lbl in CLOCK_POS:
        tc = (clock_h / 12.0) * 2.0 * np.pi
        fig.add_trace(go.Scatter3d(
            x=[view_x_min],
            y=[label_r * np.sin(tc)],
            z=[label_r * np.cos(tc)],
            mode="text",
            text=[clock_lbl],
            textfont=dict(size=11, color="rgb(80,80,80)"),
            showlegend=False,
            hoverinfo="skip",
        ))

    # ── 8. End-cap discs (upstream and downstream) ────────────────────────────
    # Discs close the pipe ends so the cylinder looks like real pipe joints
    r_disc = np.array([0.0, pipe_r_m])
    theta_disc = np.linspace(0, 2.0 * np.pi, 90)
    R_D, T_D = np.meshgrid(r_disc, theta_disc)
    Y_D = R_D * np.sin(T_D)
    Z_D = R_D * np.cos(T_D)
    for x_cap in [view_x_min, view_x_max]:
        fig.add_trace(go.Surface(
            x=np.full_like(Y_D, x_cap),
            y=Y_D,
            z=Z_D,
            surfacecolor=np.zeros_like(Y_D),
            colorscale=[[0, "rgb(180,190,200)"], [1, "rgb(180,190,200)"]],
            cmin=0, cmax=1,
            showscale=False,
            opacity=0.55,
            hoverinfo="skip",
            showlegend=False,
        ))

    # ── 9. Longseam welds (drawn last so end caps / surface do not bury them) ─
    # Same colours and source-filter dimming as 2D; skip rows without both chainages.
    all_seam_sources_3d: list[str] = sorted(set(sw.get("source", "") for sw in seam_welds))
    seam_legend_shown: set = set()
    r_ls = pipe_r_m * 1.085
    for sw in seam_welds:
        if sw.get("chainage_start") is None or sw.get("chainage_end") is None:
            continue
        oh = sw.get("orientation_hours", 6.0)
        theta_sw = (oh / 12.0) * 2.0 * np.pi
        ch_s = max(float(sw["chainage_start"]), view_x_min)
        ch_e = min(float(sw["chainage_end"]), view_x_max)
        if ch_s >= ch_e:
            continue
        js_src = sw.get("source", "")
        a = _seam_trace_alpha(sw, filter_by_source, source_filter)
        color = _seam_src_rgba(js_src, all_seam_sources_3d, a)
        y_ls = r_ls * np.sin(theta_sw)
        z_ls = r_ls * np.cos(theta_sw)
        orient_lbl = sw.get("orientation_label", f"{oh:.2f}")
        hover_src = f" [{js_src}]" if js_src else ""
        gwd_h = sw.get("gwd_number")
        gwd_ht = f"GWD {gwd_h}<br>" if gwd_h is not None else ""
        show_leg = js_src not in seam_legend_shown
        seam_legend_shown.add(js_src)
        legend_name = f"Longseam: {js_src}" if js_src else "Longseam"
        fig.add_trace(go.Scatter3d(
            x=[ch_s, ch_e],
            y=[y_ls, y_ls],
            z=[z_ls, z_ls],
            mode="lines",
            line=dict(color=color, width=2.5),
            name=legend_name,
            showlegend=show_leg,
            legendgroup=f"longseam_{js_src}",
            hovertemplate=f"<b>Longseam</b><br>{gwd_ht}{orient_lbl}{hover_src}<extra></extra>",
        ))

    # ── 10. Layout ────────────────────────────────────────────────────────────
    # Compress the x-axis so the pipe looks physically round, not needle-thin.
    # Natural ratio is x_span / diameter; we divide by ~7 to compress,  cap at 8.
    natural_ratio = x_span / (pipe_r_m * 2.0)
    x_ratio = float(np.clip(natural_ratio / 7.0, 2.5, 8.0))

    pad = pipe_r_m * 0.45   # padding around the cylinder cross-section

    fig.update_layout(
        title=dict(text=title, font=dict(size=13)),
        scene=dict(
            xaxis=dict(
                title="Chainage (m)",
                showgrid=True, gridcolor="rgba(0,0,0,0.06)",
                range=[view_x_min - x_span * 0.02, view_x_max + x_span * 0.02],
            ),
            yaxis=dict(
                title="Y (m)  ← 9:00 | 3:00 →",
                showgrid=True, gridcolor="rgba(0,0,0,0.06)",
                range=[-pipe_r_m - pad, pipe_r_m + pad],
            ),
            zaxis=dict(
                title="Z (m)  ↑ 12:00 top",
                showgrid=True, gridcolor="rgba(0,0,0,0.06)",
                range=[-pipe_r_m - pad, pipe_r_m + pad],
            ),
            aspectmode="manual",
            aspectratio=dict(x=x_ratio, y=1.0, z=1.0),
            camera=dict(
                eye=dict(x=1.4, y=-2.1, z=0.85),
                up=dict(x=0, y=0, z=1),
            ),
            bgcolor="rgb(245,248,252)",
        ),
        height=650,
        template="plotly_white",
        margin=dict(l=0, r=20, b=0, t=50),
        legend=dict(
            x=1.01, y=0.75, xanchor="left",
            bgcolor="rgba(255,255,255,0.85)",
            bordercolor="rgba(0,0,0,0.18)",
            borderwidth=1,
            font=dict(size=11),
        ),
        paper_bgcolor="white",
    )
    return fig


@st.fragment
def render_feature_map_fragment(
    fm: dict,
    total_before_filter: int | None = None,
    key_prefix: str = "ili",
    *,
    source_filter_layout: Literal["columns", "stack"] = "columns",
):
    """
    Fragment wrapper: when source checkboxes (or NPS) change, only this reruns—not the full page.

    Important: ``fm`` is already the parsed API result from session state. This path does not
    call FastAPI (no re-upload, no /process-dig-package, /process-feature-map, etc.); it only
    filters in Python and rebuilds Plotly figures.
    """
    render_feature_map(
        fm,
        total_before_filter=total_before_filter,
        key_prefix=key_prefix,
        source_filter_layout=source_filter_layout,
    )


def render_feature_map(
    fm: dict,
    total_before_filter: int | None = None,
    key_prefix: str = "ili",
    *,
    source_filter_layout: Literal["columns", "stack"] = "columns",
):
    """
    Render the unwrapped pipe feature map from FeatureMapResponse data.
    Used by both paste and upload modes.
    """
    features = fm.get("features", [])
    mapping = fm.get("column_mapping", {})
    scatter_data = fm.get("scatter_data") or {}
    all_sources = fm.get("sources", [])
    feature_summary_raw = fm.get("feature_summary_raw")

    st.markdown("### 🗺️ Feature Map")
    caption_parts = [
        "Unwrapped pipe view: X = chainage (m) or Distance from TGW (m); Y = orientation (o'clock). "
        "Feature boxes proportional to length (mm) × width (mm). Depth: green 0-20% WT, yellow 20-40%, orange 40-60%, red >60%. "
        "Red vertical lines = girth welds. Blue horizontal lines = longseam per span (changes at each GWD)."
    ]
    if all_sources:
        caption_parts.append("Deselected sources shown at 20% opacity for overlap comparison.")
    st.caption(" ".join(caption_parts))

    nps_options = sorted(NPS_TO_OD_MM.keys())
    nps_default = 10 if 10 in nps_options else nps_options[0]
    nps = st.selectbox(
        "**Pipe NPS (Nominal Pipe Size)**",
        options=nps_options,
        index=nps_options.index(nps_default),
        format_func=lambda x: f"NPS {x}",
        help="Select pipe size for width scaling. OD is used to convert feature width (mm) to circumferential position.",
        key=f"{key_prefix}_nps_select",
    )
    pipe_od_mm = NPS_TO_OD_MM.get(nps, 273.0)
    pipe_circ_mm = 3.14159 * pipe_od_mm

    # Source checkboxes only filter ``features`` / scatter_data already in memory (``fm``).
    # No backend request—fragment reruns stay inside ``render_feature_map_fragment``.
    selected_sources = set()
    if all_sources:
        st.markdown("**Filter by ILI Source** — check/uncheck to compare:")
        if source_filter_layout == "stack":
            # Vertical checkboxes avoid st.columns-in-columns (e.g. dig package maps beside workbook).
            for src in all_sources:
                if st.checkbox(src, value=True, key=f"{key_prefix}_source_{src}"):
                    selected_sources.add(src)
        else:
            cols = st.columns(min(len(all_sources), 6))
            for i, src in enumerate(all_sources):
                with cols[i % len(cols)]:
                    if st.checkbox(src, value=True, key=f"{key_prefix}_source_{src}"):
                        selected_sources.add(src)
    filter_by_source = bool(all_sources)

    x_column = scatter_data.get("x_column", "ILI Chainage (m)")
    joint_context_by_source = scatter_data.get("joint_context_by_source", {})

    def _selected_joint_contexts(active_sources: set[str]) -> list[dict]:
        if not joint_context_by_source:
            return []
        if active_sources:
            return [
                joint_context_by_source[src]
                for src in all_sources
                if src in active_sources and src in joint_context_by_source
            ]
        return [joint_context_by_source[src] for src in all_sources if src in joint_context_by_source]

    main_title = "ILI Feature Map (Unwrapped Pipe View)"
    if all_sources and selected_sources:
        main_title += f" — {' + '.join(sorted(selected_sources))}"

    fig, filtered_count = _build_feature_map_figure(
        features,
        scatter_data,
        pipe_circ_mm,
        selected_sources,
        filter_by_source,
        x_column,
        main_title,
        height=450,
        use_opacity_overlay=filter_by_source,
    )
    st.plotly_chart(fig, width="stretch", config=_FEATURE_MAP_PLOTLY_CONFIG)

    combined_contexts = _selected_joint_contexts(selected_sources)
    if combined_contexts:
        st.caption(
            "Joint context: "
            + " || ".join(
                _format_joint_context_summary(ctx)
                for ctx in combined_contexts
                if _format_joint_context_summary(ctx)
            )
        )

    filter_msg = ""
    if total_before_filter is not None and filtered_count != total_before_filter:
        filter_msg = f" (filtered from {total_before_filter})"
    elif filtered_count != len(features):
        filter_msg = f" (filtered from {len(features)})"
    st.info(f"**{filtered_count} features** visualized" + filter_msg)

    # ── 3D Pipeline View ──────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🔩 3D Pipeline View")
    girth_welds_count = len(scatter_data.get("girth_welds", []))
    joint_count = max(min(girth_welds_count - 1, 3), 0) if girth_welds_count >= 2 else 0
    st.caption(
        f"Interactive 3D cylinder view showing **{joint_count or 'up to 3'} joint(s)** centred in the data. "
        "Metal loss depth is painted directly onto the pipe surface as a heatmap (C-scan rewrapped). "
        "Rotate with mouse drag · Zoom with scroll. "
        "**Heatmap** = depth % WT (green <20 % → yellow → orange → red >60 %) · "
        "**Red rings** = GWDs (girth welds) · "
        "**Coloured lines** = longseam per ILI source (blue = source 1, purple = source 2, …) · "
        "**Clock labels** = pipe orientation (12 o'clock = top of pipe). "
        "All sources' longseam lines are shown simultaneously — upstream/downstream joints show "
        "longseam orientation even when they contain no defects."
    )
    fig_3d = _build_3d_pipeline_figure(
        features=features,
        scatter_data=scatter_data,
        pipe_od_mm=pipe_od_mm,
        source_filter=selected_sources,
        filter_by_source=filter_by_source,
        title=f"3D Pipeline View — {joint_count or '≤3'} Joint(s) | NPS {nps} (OD {pipe_od_mm:.1f} mm)",
        max_joints=3,
    )
    st.plotly_chart(fig_3d, width="stretch")

    if len(all_sources) >= 2:
        st.markdown("---")
        st.markdown("#### 📊 Breakdown by ILI Source")
        st.caption("Individual feature maps for each ILI source.")
        for src in all_sources:
            fig_breakdown, count = _build_feature_map_figure(
                features,
                scatter_data,
                pipe_circ_mm,
                {src},
                True,
                x_column,
                f"Source: {src}",
                height=380,
            )
            fig_breakdown.update_layout(title=f"Source: {src} ({count} features)")
            st.plotly_chart(
                fig_breakdown,
                width="stretch",
                config=_FEATURE_MAP_PLOTLY_CONFIG,
            )
            src_context = joint_context_by_source.get(src)
            if src_context:
                st.caption("Joint context: " + _format_joint_context_summary(src_context))

    with st.expander("🔧 Column mapping", expanded=False):
        st.json(mapping)


def _prepare_data_for_api(df: pd.DataFrame) -> str:
    """If first row looks like ILI headers, use it as column names."""
    if df.empty or len(df) < 1:
        return ""
    first_row = df.iloc[0].astype(str).str.strip()
    header_keywords = [
        "ili",
        "chainage",
        "distance",
        "depth",
        "feature",
        "length",
        "width",
        "orientation",
        "odometer",
    ]
    first_row_lower = " ".join(first_row.str.lower())
    if len(df) >= 2 and any(keyword in first_row_lower for keyword in header_keywords):
        new_cols = [str(value).strip() or f"_col{i}" for i, value in enumerate(first_row)]
        out = df.iloc[1:].copy()
        column_count = len(out.columns)
        out.columns = [new_cols[i] if i < len(new_cols) else f"_col{i}" for i in range(column_count)]
        return out.to_csv(index=False).strip()
    return df.to_csv(index=False).strip()


def _init_ili_session_state() -> None:
    st.session_state.setdefault("ili_preview_data", None)
    st.session_state.setdefault("ili_feature_map_data", None)
    st.session_state.setdefault("ili_upload_file_name", None)
    st.session_state.setdefault("ili_upload_feature_map_data", None)
    st.session_state.setdefault("ili_pasted_df", pd.DataFrame([[""] * 15]))


def _excel_to_html(file_bytes: bytes, max_rows_per_sheet: int = 300) -> str:
    """
    Convert an Excel workbook to a self-contained HTML string showing every sheet
    as a scrollable table.  Uses openpyxl (data_only mode) so formula values are
    preserved without recalculation.

    Merged cells are rendered with the correct rowspan/colspan so the visual
    layout matches what the user sees in Excel.
    """
    import io
    import openpyxl
    from openpyxl.utils import get_column_letter  # noqa: F401

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return f"<p style='color:red'>Could not read workbook: {exc}</p>"

    css = """
<style>
  body { font-family: Calibri, Arial, sans-serif; font-size: 11px; margin: 0; padding: 8px; background:#fff; }
  .sheet-name {
    display: inline-block; padding: 4px 14px; margin: 0 4px 0 0;
    background: #4472c4; color: #fff; border-radius: 4px 4px 0 0;
    font-weight: bold; font-size: 11px; letter-spacing: .3px;
  }
  .sheet-block { margin-bottom: 28px; }
  table {
    border-collapse: collapse; width: auto; min-width: 100%;
    table-layout: fixed; font-size: 11px;
  }
  th, td {
    border: 1px solid #c8cdd2; padding: 2px 7px;
    white-space: pre-wrap; word-break: break-word;
    max-width: 240px; vertical-align: middle;
  }
  th { background: #dce6f1; font-weight: bold; text-align: center; }
  td { text-align: left; }
  tr:nth-child(even) td { background: #f5f8fd; }
  .row-trunc { font-style: italic; color: #888; text-align: center; }
</style>
"""

    parts = [css]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_column is None:
            continue

        # Build a set of "hidden" merged cells and a map of anchor→(rowspan, colspan)
        hidden: set[tuple[int, int]] = set()
        span_map: dict[tuple[int, int], tuple[int, int]] = {}
        for mr in ws.merged_cells.ranges:
            rs = mr.max_row - mr.min_row + 1
            cs = mr.max_col - mr.min_col + 1
            span_map[(mr.min_row, mr.min_col)] = (rs, cs)
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if (r, c) != (mr.min_row, mr.min_col):
                        hidden.add((r, c))

        max_col = ws.max_column
        max_row = min(ws.max_row, max_rows_per_sheet)
        truncated = ws.max_row > max_rows_per_sheet

        parts.append(f'<div class="sheet-block">')
        parts.append(f'<span class="sheet-name">📄 {sheet_name}</span>')
        parts.append("<table>")

        for row_idx in range(1, max_row + 1):
            row_html_cells = []
            all_empty = True
            for col_idx in range(1, max_col + 1):
                if (row_idx, col_idx) in hidden:
                    continue
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                if val is not None:
                    all_empty = False

                # Format value
                if val is None:
                    display = ""
                elif isinstance(val, float):
                    # Show up to 4 significant figures; strip trailing zeros
                    display = f"{val:.4g}"
                elif hasattr(val, "strftime"):
                    display = str(val)
                else:
                    raw = str(val).strip()
                    display = raw[:200] + ("…" if len(raw) > 200 else "")
                # HTML-escape
                display = display.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

                # rowspan / colspan
                span = span_map.get((row_idx, col_idx))
                attrs = ""
                if span:
                    rs, cs = span
                    if rs > 1:
                        attrs += f' rowspan="{rs}"'
                    if cs > 1:
                        attrs += f' colspan="{cs}"'

                # Use <th> for bold cells (typically headers)
                is_bold = cell.font and cell.font.bold
                tag = "th" if is_bold else "td"
                row_html_cells.append(f"<{tag}{attrs}>{display}</{tag}>")

            if all_empty:
                continue  # skip completely blank rows
            parts.append("<tr>" + "".join(row_html_cells) + "</tr>")

        if truncated:
            parts.append(
                f'<tr><td colspan="{max_col}" class="row-trunc">'
                f"⚠ Showing first {max_rows_per_sheet} of {ws.max_row} rows"
                f"</td></tr>"
            )

        parts.append("</table></div>")

    return "\n".join(parts)


def _stable_dig_pdf_dom_id(filename: str) -> str:
    """Stable element id for PDF.js DOM (avoids full re-init on unrelated reruns)."""
    h = hashlib.md5(filename.encode("utf-8"), usedforsecurity=False).hexdigest()[:12]
    return f"pdf{h}"


# ── Dig Package Floating PDF Panel ─────────────────────────────────────────────
_DIG_PANEL_WIDTH_VW  = 40
_DIG_PANEL_STYLE_ID  = "dig-pdf-panel-style"
_DIG_PANEL_DATA_ATTR = "data-dig-pdf-panel"
_DIG_PANEL_HTML_VER  = 2


def _dig_pdf_floating_panel_html(panel_id: str, pdfjs_base: str, filename: str, b64: str) -> str:
    """Build the floating panel HTML — mirrors the Inspection Report panel design."""
    pw = _DIG_PANEL_WIDTH_VW
    short = filename.rsplit(".", 1)[0] if "." in filename else filename
    short_js = short[:20] + "\u2026" if len(short) > 22 else short
    return f"""<!DOCTYPE html>
<html><head><style>html,body{{margin:0;padding:0;background:transparent;overflow:hidden;}}</style></head>
<body><script>
(function() {{
  var pid     = {repr(panel_id)};
  var pjsBase = {repr(pdfjs_base)};
  var b64data = {repr(b64)};
  var fname   = {repr(filename)};
  var shortName = {repr(short_js)};
  var pw      = {pw};
  var P       = window.parent.document;

  /* ── Layout CSS — push main content left ── */
  var styleId = {repr(_DIG_PANEL_STYLE_ID)};
  if (!P.getElementById(styleId)) {{
    var st = P.createElement('style');
    st.id  = styleId;
    st.textContent =
      'div[data-testid="stMainBlockContainer"],' +
      'div[data-testid="stAppViewBlockContainer"] {{' +
      '  max-width:none !important;' +
      '  padding-right:calc(' + pw + 'vw + 28px) !important;' +
      '  box-sizing:border-box !important;' +
      '  transition:padding-right .2s ease;' +
      '}}';
    P.head.appendChild(st);
  }}

  /* ── Guard: remove stale panels, skip if same panel already visible ── */
  P.querySelectorAll('[{_DIG_PANEL_DATA_ATTR}]').forEach(function(el) {{
    if (el.id !== pid) el.remove();
  }});
  if (P.getElementById(pid)) {{
    P.getElementById(pid).style.display = '';
    return;
  }}

  /* ══════════════════ BUILD PANEL ══════════════════ */
  var panel = P.createElement('div');
  panel.id  = pid;
  panel.setAttribute('{_DIG_PANEL_DATA_ATTR}', '1');
  panel.style.cssText = [
    'position:fixed','right:0','top:58px',
    'width:' + pw + 'vw','height:calc(100vh - 58px)',
    'z-index:9000','background:#1e1e1e',
    'border-left:2px solid #3c3c3c',
    'display:flex','flex-direction:column','overflow:hidden',
    'box-shadow:-4px 0 20px rgba(0,0,0,.6)'
  ].join(';');

  /* ── Tab bar ── */
  var tabBar = P.createElement('div');
  tabBar.style.cssText = [
    'display:flex','align-items:flex-end',
    'background:#252526','border-bottom:1px solid #3c3c3c',
    'overflow-x:auto','flex-shrink:0',
    'scrollbar-width:thin','padding-top:4px'
  ].join(';');

  var tab = P.createElement('div');
  tab.title = fname;
  tab.style.cssText = [
    'display:flex','align-items:center','gap:5px',
    'padding:5px 12px 6px','cursor:default',
    'border-right:1px solid #3c3c3c',
    'font:12px system-ui,sans-serif',
    'white-space:nowrap','flex-shrink:0',
    'background:#1e1e1e','color:#ffffff',
    'border-top:2px solid #007acc'
  ].join(';');
  var ico = P.createElement('span'); ico.textContent = '\uD83D\uDCC4';
  var lbl = P.createElement('span'); lbl.textContent = shortName;
  tab.append(ico, lbl);
  tabBar.appendChild(tab);

  var closeBtn = P.createElement('button');
  closeBtn.style.cssText = [
    'margin-left:auto','flex-shrink:0',
    'background:transparent','border:none',
    'color:#858585','cursor:pointer',
    'font-size:18px','line-height:1',
    'padding:4px 10px','align-self:center'
  ].join(';');
  closeBtn.title = 'Dismiss';
  closeBtn.textContent = '\u00d7';
  tabBar.appendChild(closeBtn);

  /* ── Zoom toolbar (drag to resize) ── */
  var zoomBar = P.createElement('div');
  zoomBar.style.cssText = [
    'padding:4px 12px','background:#2d2d2d',
    'border-bottom:1px solid #3c3c3c',
    'display:flex','align-items:center','gap:8px','flex-shrink:0',
    'user-select:none'
  ].join(';');
  var zpct = P.createElement('span');
  zpct.style.cssText = 'color:#ccc;min-width:42px;font:12px ui-monospace,monospace;';
  zpct.textContent = '130%';
  var pgInfo = P.createElement('span');
  pgInfo.style.cssText = 'color:#888;font:11px system-ui;margin-left:auto;white-space:nowrap;';
  function mkBtn(t) {{
    var b = P.createElement('button');
    b.textContent = t;
    b.style.cssText = 'padding:1px 9px;border-radius:4px;border:1px solid #666;background:#444;color:#fff;cursor:pointer;font-size:13px;';
    return b;
  }}
  var zOut = mkBtn('\u2212'), zIn = mkBtn('+'), zRst = mkBtn('Reset');
  zRst.style.fontSize = '11px';
  var zLabel = P.createElement('span');
  Object.assign(zLabel, {{style:'color:#aaa;font:600 11px system-ui;', textContent:'Zoom'}});
  zoomBar.append(zLabel, zOut, zpct, zIn, zRst, pgInfo);

  /* ── Brighter scrollbar CSS for the PDF wrap area ── */
  (function() {{
    var sbId = pid + '-sb';
    if (!P.getElementById(sbId)) {{
      var sbStyle = P.createElement('style');
      sbStyle.id = sbId;
      sbStyle.textContent =
        '#' + pid + ' div[style*="overflow:auto"]::-webkit-scrollbar {{' +
        '  width:10px;height:10px;}}' +
        '#' + pid + ' div[style*="overflow:auto"]::-webkit-scrollbar-track {{' +
        '  background:#1e1e1e;}}' +
        '#' + pid + ' div[style*="overflow:auto"]::-webkit-scrollbar-thumb {{' +
        '  background:#6b7280;border-radius:5px;border:2px solid #1e1e1e;}}' +
        '#' + pid + ' div[style*="overflow:auto"]::-webkit-scrollbar-thumb:hover {{' +
        '  background:#9ca3af;}}';
      P.head.appendChild(sbStyle);
    }}
  }})();

  /* ── Scrollable PDF area ── */
  var wrap = P.createElement('div');
  wrap.style.cssText = 'flex:1;overflow:auto;background:#3c3c3c;padding:8px;scrollbar-color:#6b7280 #1e1e1e;scrollbar-width:thin;';
  var pages = P.createElement('div');
  pages.style.cssText = 'width:max-content;min-width:100%;';
  var errEl = P.createElement('div');
  errEl.style.cssText = 'display:none;color:#f88;padding:12px;font:13px system-ui;';
  wrap.append(pages, errEl);

  panel.append(tabBar, zoomBar, wrap);
  P.body.appendChild(panel);

  /* ══════════════════ ZOOM & RENDER ══════════════════ */
  var baseScale = 1.1, curScale = baseScale * 1.3;
  var pdfDoc = null;
  function zPct() {{ return Math.round((curScale / baseScale) * 100); }}
  function updZ() {{ zpct.textContent = zPct() + '%'; }}

  function renderDoc(preserveScroll) {{
    if (!pdfDoc) return;
    var saved = preserveScroll ? wrap.scrollTop : 0;
    pages.innerHTML = '';
    errEl.style.display = 'none';
    var chain = Promise.resolve();
    for (var p = 1; p <= pdfDoc.numPages; p++) {{
      (function(n) {{
        chain = chain.then(function() {{
          return pdfDoc.getPage(n).then(function(page) {{
            var vp  = page.getViewport({{scale: curScale}});
            var cv  = P.createElement('canvas');
            var ctx = cv.getContext('2d');
            var dpr = window.parent.devicePixelRatio || 1;
            cv.width  = Math.floor(vp.width  * dpr);
            cv.height = Math.floor(vp.height * dpr);
            cv.style.cssText = 'display:block;margin:0 auto 10px;box-shadow:0 1px 6px rgba(0,0,0,.5);' +
              'width:' + Math.floor(vp.width) + 'px;height:' + Math.floor(vp.height) + 'px;';
            pages.appendChild(cv);
            return page.render({{
              canvasContext: ctx, viewport: vp,
              transform: dpr !== 1 ? [dpr,0,0,dpr,0,0] : null
            }}).promise;
          }});
        }});
      }})(p);
    }}
    chain.then(function() {{ if (saved) wrap.scrollTop = saved; }});
  }}

  zOut.onclick = function() {{ curScale = Math.max(baseScale*.35, curScale/1.2); updZ(); renderDoc(true); }};
  zIn.onclick  = function() {{ curScale = Math.min(baseScale*12,  curScale*1.2); updZ(); renderDoc(true); }};
  zRst.onclick = function() {{ curScale = baseScale*1.3; updZ(); renderDoc(false); }};
  updZ();

  /* ══════════════════ DRAG-TO-RESIZE ══════════════════ */
  var EDGE = 14;
  function beginResize(e) {{
    if (e.button !== 0) return;
    e.preventDefault();
    var startX = e.clientX, startW = panel.getBoundingClientRect().width;
    function onMove(ev) {{
      var newW = Math.max(280, startW - (ev.clientX - startX));
      var pct  = Math.round((newW / window.parent.innerWidth) * 100);
      panel.style.width = pct + 'vw';
      var st2 = P.getElementById({repr(_DIG_PANEL_STYLE_ID)});
      if (st2) st2.textContent =
        'div[data-testid="stMainBlockContainer"],' +
        'div[data-testid="stAppViewBlockContainer"] {{' +
        'max-width:none !important;padding-right:calc(' + pct + 'vw + 28px) !important;' +
        'box-sizing:border-box !important;}}';
    }}
    function onUp() {{
      P.removeEventListener('mousemove', onMove);
      P.removeEventListener('mouseup', onUp);
      wrap.style.cursor = '';
    }}
    P.addEventListener('mousemove', onMove);
    P.addEventListener('mouseup', onUp);
  }}
  zoomBar.addEventListener('mousedown', function(e) {{
    if (e.target.closest && e.target.closest('button')) return;
    beginResize(e);
  }});
  wrap.addEventListener('mousedown', function(e) {{
    if (e.clientX - panel.getBoundingClientRect().left > EDGE) return;
    beginResize(e);
  }});
  wrap.addEventListener('mousemove', function(e) {{
    wrap.style.cursor = (e.clientX - panel.getBoundingClientRect().left <= EDGE) ? 'ew-resize' : '';
  }});
  wrap.addEventListener('mouseleave', function() {{ wrap.style.cursor = ''; }});

  /* ══════════════════ CLOSE ══════════════════ */
  closeBtn.onclick = function() {{
    panel.style.display = 'none';
    var st2 = P.getElementById({repr(_DIG_PANEL_STYLE_ID)});
    if (st2) st2.remove();
  }};

  /* ══════════════════ LOAD PDF ══════════════════ */
  function loadPdf() {{
    var lib = window.parent.pdfjsLib;
    if (!lib) {{
      errEl.style.display = 'block';
      errEl.textContent = 'PDF.js not loaded — ensure the backend is running.';
      return;
    }}
    lib.GlobalWorkerOptions.workerSrc = pjsBase + '/pdf.worker.min.js';
    pages.innerHTML = '<div style="color:#aaa;padding:16px;font:13px system-ui;">Loading\u2026</div>';
    var raw = atob(b64data), bytes = new Uint8Array(raw.length);
    for (var i = 0; i < raw.length; i++) bytes[i] = raw.charCodeAt(i);
    lib.getDocument({{data: bytes}}).promise.then(function(pdf) {{
      pdfDoc = pdf;
      var n = pdf.numPages;
      pgInfo.textContent = n === 1 ? '1 page' : n + ' pages';
      renderDoc(false);
    }}).catch(function(e) {{
      errEl.style.display = 'block';
      errEl.textContent = 'PDF error: ' + (e && e.message || String(e));
    }});
  }}

  if (window.parent.pdfjsLib) {{
    loadPdf();
  }} else {{
    var s = P.createElement('script');
    s.src    = pjsBase + '/pdf.min.js';
    s.onload = loadPdf;
    s.onerror = function() {{
      errEl.style.display = 'block';
      errEl.textContent = 'Failed to load PDF.js from the backend.';
    }};
    P.head.appendChild(s);
  }}
}})();
</script></body></html>"""


def _render_dig_pdf_floating_panel(filename: str, pdf_bytes: bytes) -> None:
    """Inject the floating workbook PDF panel into the parent DOM."""
    pdfjs_base = f"{BACKEND_URL.rstrip('/')}/static/pdfjs"
    fp = hashlib.md5(pdf_bytes[:16384], usedforsecurity=False).hexdigest()[:16]
    panel_id = "dig-pdf-" + fp

    b64_key = f"dig_pdf_b64_{fp}"
    if b64_key not in st.session_state:
        st.session_state[b64_key] = base64.b64encode(pdf_bytes).decode()
    b64 = st.session_state[b64_key]

    html_key = f"dig_pdf_panel_html_{fp}_v{_DIG_PANEL_HTML_VER}"
    done_key = f"dig_pdf_panel_done_{fp}_v{_DIG_PANEL_HTML_VER}"

    if html_key not in st.session_state:
        st.session_state[html_key] = _dig_pdf_floating_panel_html(
            panel_id, pdfjs_base, filename, b64
        )

    if not st.session_state.get(done_key):
        st_components.html(st.session_state[html_key], height=0, scrolling=False)
        st.session_state[done_key] = True


def _cleanup_dig_pdf_panel() -> None:
    """Remove the floating dig package PDF panel and layout CSS from the parent DOM."""
    for key in list(st.session_state.keys()):
        if key.startswith("dig_pdf_panel_done_"):
            del st.session_state[key]
    st_components.html(
        f"""<script>
        var P = window.parent.document;
        P.querySelectorAll('[{_DIG_PANEL_DATA_ATTR}]').forEach(function(el){{el.remove();}});
        var st = P.getElementById({repr(_DIG_PANEL_STYLE_ID)});
        if (st) st.remove();
        </script>""",
        height=0,
        scrolling=False,
    )


def _dig_pdf_js_viewer_html(
    vid: str,
    pdfjs_base: str,
    b64: str,
    base_scale: float,
    wrap_max_h: str,
    *,
    default_zoom_pct: float = 140.0,
) -> str:
    """Self-contained PDF.js viewer HTML: zoom buttons + reset; scroll panes when zoom exceeds column width."""
    return f"""<div id="pdf-shell-{vid}" style="border:1px solid #5a5f66;border-radius:8px;padding:8px;background:#3a3d42;box-sizing:border-box;">
<style>
#pdf-wrap-{vid}::-webkit-scrollbar{{width:10px;height:10px;}}
#pdf-wrap-{vid}::-webkit-scrollbar-track{{background:#2d2f33;}}
#pdf-wrap-{vid}::-webkit-scrollbar-thumb{{background:#6b7280;border-radius:5px;border:2px solid #2d2f33;}}
#pdf-wrap-{vid}::-webkit-scrollbar-thumb:hover{{background:#9ca3af;}}
</style>
<div id="pdf-toolbar-{vid}" style="display:flex;align-items:center;gap:8px;margin-bottom:8px;flex-wrap:wrap;">
  <span style="color:#e8eaed;font:600 12px system-ui,sans-serif;">Zoom</span>
  <button type="button" id="pdf-zout-{vid}" style="padding:4px 12px;border-radius:6px;border:1px solid #888;background:#555;color:#fff;cursor:pointer;">−</button>
  <span id="pdf-zpct-{vid}" style="color:#e8eaed;min-width:44px;font:13px ui-monospace,monospace;">{int(default_zoom_pct)}%</span>
  <button type="button" id="pdf-zin-{vid}" style="padding:4px 12px;border-radius:6px;border:1px solid #888;background:#555;color:#fff;cursor:pointer;">+</button>
  <button type="button" id="pdf-zreset-{vid}" style="padding:4px 10px;border-radius:6px;border:1px solid #888;background:#444;color:#ddd;cursor:pointer;font-size:11px;">Reset</button>
  <span id="pdf-pginfo-{vid}" style="color:#9ca3af;font:11px system-ui;margin-left:auto;white-space:nowrap;"></span>
</div>
<div id="pdf-wrap-{vid}" style="max-height:{wrap_max_h};overflow:auto;border-radius:4px;background:#525659;padding:8px;width:100%;box-sizing:border-box;scrollbar-color:#6b7280 #2d2f33;scrollbar-width:thin;">
<div id="pdf-pages-{vid}" style="width:max-content;min-width:100%;box-sizing:border-box;"></div>
<div id="pdf-err-{vid}" style="display:none;color:#fff;padding:12px;font-family:sans-serif;"></div>
</div>
</div>
<script src="{pdfjs_base}/pdf.min.js"></script>
<script>
(function() {{
  var holder = document.getElementById("pdf-pages-{vid}");
  var errEl = document.getElementById("pdf-err-{vid}");
  var b64 = {repr(b64)};
  var pdfDoc = null;
  var baseScale = {base_scale};
  var defaultZoomFactor = {default_zoom_pct / 100.0};
  var currentScale = baseScale * defaultZoomFactor;
  function showErr(msg) {{
    errEl.style.display = "block";
    errEl.textContent = msg;
  }}
  function zPct() {{
    return Math.round((currentScale / baseScale) * 100);
  }}
  function updateZLabel() {{
    var el = document.getElementById("pdf-zpct-{vid}");
    if (el) el.textContent = zPct() + "%";
  }}
  function renderPdf() {{
    if (!pdfDoc || !holder) return Promise.resolve();
    holder.innerHTML = "";
    var chain = Promise.resolve();
    for (var p = 1; p <= pdfDoc.numPages; p++) {{
      (function(pageNum) {{
        chain = chain.then(function() {{
          return pdfDoc.getPage(pageNum).then(function(page) {{
            var viewport = page.getViewport({{ scale: currentScale }});
            var canvas = document.createElement("canvas");
            canvas.style.display = "block";
            canvas.style.margin = "0 auto 12px auto";
            canvas.style.boxShadow = "0 1px 4px rgba(0,0,0,0.4)";
            var ctx = canvas.getContext("2d");
            var dpr = window.devicePixelRatio || 1;
            canvas.width = Math.floor(viewport.width * dpr);
            canvas.height = Math.floor(viewport.height * dpr);
            canvas.style.width = Math.floor(viewport.width) + "px";
            canvas.style.height = Math.floor(viewport.height) + "px";
            var t = dpr !== 1 ? [dpr, 0, 0, dpr, 0, 0] : null;
            holder.appendChild(canvas);
            return page.render({{ canvasContext: ctx, viewport: viewport, transform: t }}).promise;
          }});
        }});
      }})(p);
    }}
    return chain;
  }}
  try {{
    if (typeof pdfjsLib === "undefined") {{
      showErr("Could not load PDF.js from the API server. Ensure the backend is running.");
      return;
    }}
    pdfjsLib.GlobalWorkerOptions.workerSrc = "{pdfjs_base}/pdf.worker.min.js";
    var raw = atob(b64);
    var len = raw.length;
    var bytes = new Uint8Array(len);
    for (var i = 0; i < len; i++) bytes[i] = raw.charCodeAt(i);
    pdfjsLib.getDocument({{ data: bytes }}).promise.then(function(pdf) {{
      pdfDoc = pdf;
      updateZLabel();
      var pgInfoEl = document.getElementById("pdf-pginfo-{vid}");
      if (pgInfoEl) {{
        var n = pdf.numPages;
        pgInfoEl.textContent = n === 1 ? "1 page" : n + " pages";
      }}
      document.getElementById("pdf-zout-{vid}").addEventListener("click", function() {{
        currentScale = Math.max(baseScale * 0.35, currentScale / 1.2);
        updateZLabel();
        renderPdf();
      }});
      document.getElementById("pdf-zin-{vid}").addEventListener("click", function() {{
        currentScale = Math.min(baseScale * 12, currentScale * 1.2);
        updateZLabel();
        renderPdf();
      }});
      document.getElementById("pdf-zreset-{vid}").addEventListener("click", function() {{
        currentScale = baseScale * defaultZoomFactor;
        updateZLabel();
        renderPdf();
      }});
      return renderPdf();
    }}).catch(function(e) {{
      showErr("PDF preview failed: " + (e && e.message ? e.message : String(e)));
    }});
  }} catch (e) {{
    showErr("PDF preview failed: " + (e && e.message ? e.message : String(e)));
  }}
}})();
</script>"""


def _render_dig_package_source_preview(
    dig_file,
    raw_bytes: bytes,
    *,
    compact: bool = False,
) -> None:
    """
    Original workbook: download buttons + PDF.js canvas preview or HTML fallback.
    ``compact`` uses a smaller render scale and height for the side-by-side layout.
    """
    scale = 1.08 if compact else 1.35
    # Taller side-by-side workbook panel so more sheet rows/pages are visible
    wrap_max_h = "min(92vh, 1120px)" if compact else "min(82vh, 880px)"
    component_h = 880 if compact else 920
    html_fallback_h = 780 if compact else 840

    fname = dig_file.name
    dl_key = f"dig_pkg_download_{fname}"

    def _excel_download() -> None:
        st.download_button(
            "⬇ Excel",
            data=raw_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=dl_key,
        )

    pdf_key = f"dig_pkg_pdf_{fname}"
    if pdf_key not in st.session_state:
        with st.spinner("Converting to PDF for exact preview (using Excel)…"):
            pdf_bytes = asyncio.run(call_excel_to_pdf_api(raw_bytes, dig_file.name))
        st.session_state[pdf_key] = pdf_bytes

    pdf_bytes = st.session_state.get(pdf_key)
    stem = fname.rsplit(".", 1)[0] if "." in fname else fname
    pdf_dl_key = f"dig_pkg_download_pdf_{fname}"

    if pdf_bytes:
        popover_label = "⬇" if compact else "⬇ Downloads"
        popover_help = "Download Excel or PDF" if compact else "Excel and PDF"
        with st.popover(popover_label, help=popover_help):
            if compact:
                st.caption("Workbook")
            _excel_download()
            st.download_button(
                "PDF",
                data=pdf_bytes,
                file_name=f"{stem}.pdf",
                mime="application/pdf",
                key=pdf_dl_key,
            )

        pdfjs_base = f"{BACKEND_URL.rstrip('/')}/static/pdfjs"
        pdf_fingerprint = hashlib.md5(
            pdf_bytes[:16384], usedforsecurity=False
        ).hexdigest()[:20]
        viewer_sig = (
            fname,
            len(pdf_bytes),
            pdf_fingerprint,
            len(raw_bytes),
            scale,
            wrap_max_h,
            pdfjs_base,
            140.0,
        )
        sig_key = f"dig_pkg_pdf_viewer_sig_{fname}"
        html_key = f"dig_pkg_pdf_viewer_html_{fname}"
        if st.session_state.get(sig_key) != viewer_sig:
            b64 = base64.b64encode(pdf_bytes).decode()
            vid = _stable_dig_pdf_dom_id(fname)
            st.session_state[html_key] = _dig_pdf_js_viewer_html(
                vid, pdfjs_base, b64, scale, wrap_max_h, default_zoom_pct=140.0
            )
            st.session_state[sig_key] = viewer_sig
        html_body = st.session_state[html_key]
        toolbar_h = 52
        st_components.html(
            html_body,
            height=component_h + toolbar_h,
            scrolling=True,
        )
    else:
        if compact:
            with st.popover("⬇", help="Download Excel"):
                st.caption("Workbook")
                _excel_download()
        else:
            _excel_download()
        st.info(
            "PDF conversion needs Microsoft Excel on this PC. "
            "Showing a best-effort HTML table preview below."
        )
        with st.spinner("Rendering Excel sheets…"):
            html_content = _excel_to_html(raw_bytes)
        scrollable_html = (
            f"<div style='overflow:auto; max-height:{wrap_max_h}; border:1px solid #dee2e6;"
            " border-radius:6px; padding:8px; background:#fff'>"
            + html_content
            + "</div>"
        )
        st_components.html(scrollable_html, height=html_fallback_h, scrolling=True)


def _init_dig_package_session_state() -> None:
    st.session_state.setdefault("dig_package_uploaded_file", None)
    st.session_state.setdefault("dig_package_feature_map_data", None)
    st.session_state.setdefault("dig_package_file_bytes", None)
    st.session_state.setdefault("dig_pdf_panel_visible", True)
    st.session_state.setdefault("dig_split_left_frac", 0.5)


def _dig_package_feature_map_kwargs(fm: dict) -> dict:
    """Shared args for dig package maps (stacked source filters; same fragment everywhere)."""
    return {
        "fm": fm,
        "total_before_filter": fm.get("total_rows"),
        "key_prefix": "dig_package",
        "source_filter_layout": "stack",
    }


def render_dig_package_visual_tool() -> None:
    _init_dig_package_session_state()

    st.caption(
        "Upload a dig package Excel with **Feature summary** and optional **Joint Summary**. "
        "Features and longseam lines use **Distance from TGW (m)** when present; multiple ILI sources are supported."
    )

    dig_file = st.file_uploader(
        "Choose a dig package Excel file (.xlsx)",
        type=["xlsx"],
        key=fu_key("ili", "dig_pkg"),
        help="Sectioned Excel with 'Feature summary' and optionally 'Joint Summary'",
    )

    if dig_file is not None:
        if st.session_state.dig_package_uploaded_file != dig_file.name:
            # New file — clear cached result so processing runs automatically below
            st.session_state.dig_package_uploaded_file = dig_file.name
            st.session_state.dig_package_feature_map_data = None
            st.session_state.dig_package_file_bytes = None

        # Cache raw bytes (needed for PDF/HTML workbook preview)
        if st.session_state.dig_package_file_bytes is None:
            st.session_state.dig_package_file_bytes = dig_file.getvalue()

        # One-shot backend parse when a new file is loaded. Checkbox / NPS changes in
        # ``render_feature_map_fragment`` do not enter this block and do not call the API.
        if st.session_state.dig_package_feature_map_data is None:
            with st.spinner(f"Parsing **{dig_file.name}** (Feature summary, Joint Summary)…"):
                result = asyncio.run(call_process_dig_package_api(dig_file))
                if result and result.get("success"):
                    st.session_state.dig_package_feature_map_data = result
                    st.success(f"✅ Parsed **{result.get('total_rows', 0)} features** from {dig_file.name}")
                elif result and not result.get("success"):
                    st.error(result.get("error", "Process failed"))

        fm = st.session_state.dig_package_feature_map_data
        raw_bytes = st.session_state.get("dig_package_file_bytes")
        maps_ok = bool(fm and fm.get("success"))

        if maps_ok and raw_bytes:
            st.markdown("---")

            # ── Convert Excel → PDF (cached) ──────────────────────────────────
            fname = dig_file.name
            stem  = fname.rsplit(".", 1)[0] if "." in fname else fname
            pdf_key = f"dig_pkg_pdf_{fname}"
            if pdf_key not in st.session_state:
                with st.spinner("Converting workbook to PDF for preview…"):
                    _pdf = asyncio.run(call_excel_to_pdf_api(raw_bytes, fname))
                st.session_state[pdf_key] = _pdf
            pdf_bytes = st.session_state.get(pdf_key)

            # ── Downloads + PDF preview toggle (side by side) ─────────────────
            col_dl, col_tog, _ = st.columns([2, 2, 3])
            with col_dl:
                if pdf_bytes:
                    with st.popover("⬇ Downloads", help="Download Excel or PDF"):
                        st.download_button(
                            "⬇ Excel",
                            data=raw_bytes,
                            file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dig_pkg_download_{fname}",
                        )
                        st.download_button(
                            "PDF",
                            data=pdf_bytes,
                            file_name=f"{stem}.pdf",
                            mime="application/pdf",
                            key=f"dig_pkg_download_pdf_{fname}",
                        )
                else:
                    st.download_button(
                        "⬇ Excel",
                        data=raw_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dig_pkg_download_{fname}",
                    )
            with col_tog:
                if pdf_bytes:
                    panel_on = st.toggle(
                        "📄 PDF Preview",
                        value=bool(st.session_state.get("dig_pdf_panel_visible", True)),
                        key="dig_pdf_panel_visible_toggle",
                        help="Show/hide the workbook PDF panel",
                    )
                    st.session_state.dig_pdf_panel_visible = panel_on
                else:
                    panel_on = False

            # ── Floating panel inject / cleanup ───────────────────────────────
            if panel_on and pdf_bytes:
                _render_dig_pdf_floating_panel(fname, pdf_bytes)
            else:
                _cleanup_dig_pdf_panel()

            # ── Maps — always full-width ──────────────────────────────────────
            render_feature_map_fragment(**_dig_package_feature_map_kwargs(fm))

        elif maps_ok:
            st.markdown("---")
            render_feature_map_fragment(**_dig_package_feature_map_kwargs(fm))
        elif raw_bytes:
            st.markdown("---")
            st.markdown("### 📄 Original workbook")
            st.caption(
                "Parse did not return features — you can still review the file. "
                "PDF.js canvas preview (no browser PDF plug-in required)."
            )
            with st.container(border=True):
                _render_dig_package_source_preview(dig_file, raw_bytes, compact=False)

    else:
        st.info(
            "👆 Upload a **dig package** Excel (sections **Feature summary** / **Joint Summary**). "
            "Parsing runs automatically after upload."
        )


def _render_ili_paste_mode() -> None:
    st.markdown("### 📋 Paste ILI Data")
    st.caption(
        "Copy a table from Excel (including header row) and paste directly into the table below. "
        "Click Generate to visualize."
    )

    col_config = {
        str(i): st.column_config.TextColumn(label=" ", width="small") for i in range(20)
    }
    edited_df = st.data_editor(
        st.session_state.ili_pasted_df,
        width="stretch",
        num_rows="dynamic",
        key="ili_paste_editor",
        hide_index=True,
        column_config=col_config,
    )
    st.session_state.ili_pasted_df = edited_df

    if st.button("🗺️ Generate Feature Map", type="primary", key="ili_paste_generate"):
        to_send = _prepare_data_for_api(edited_df) if not edited_df.empty else ""
        if not to_send or edited_df.empty:
            st.warning("Please paste data into the table above (copy from Excel and paste).")
        else:
            with st.spinner("Parsing and mapping columns..."):
                result = asyncio.run(call_parse_paste_api(to_send))
                if result and result.get("success"):
                    st.session_state.ili_feature_map_data = result
                    st.success(f"✅ Parsed {result.get('total_rows', 0)} features!")
                elif result and not result.get("success"):
                    st.error(result.get("error", "Parse failed"))

    if st.session_state.ili_feature_map_data and st.session_state.ili_feature_map_data.get("success"):
        st.markdown("---")
        render_feature_map_fragment(st.session_state.ili_feature_map_data, key_prefix="ili_paste")
    else:
        st.info(
            """
            👆 **Paste your ILI data into the table above**

            Copy a table from Excel (with headers) and paste directly into the table.
            Column format is preserved. Click **Generate Feature Map** to visualize.
            """
        )


def _render_ili_upload_mode() -> None:
    st.markdown("### 📁 Step 1: Upload Excel File")
    uploaded_file = st.file_uploader(
        "Choose an Excel file (.xlsx or .xls)",
        type=["xlsx", "xls"],
        help="Maximum file size: 100 MB",
        key=fu_key("ili", "excel"),
    )

    if uploaded_file is not None:
        if st.session_state.ili_upload_file_name != uploaded_file.name:
            st.session_state.ili_upload_file_name = uploaded_file.name
            st.session_state.ili_preview_data = None
            st.session_state.ili_upload_feature_map_data = None

        st.success(f"✅ File uploaded: **{uploaded_file.name}**")

        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            if st.button("🔍 Preview File", type="primary", key="ili_preview_button"):
                with st.spinner("Loading file preview..."):
                    preview_data = asyncio.run(call_preview_api(uploaded_file))
                    if preview_data:
                        st.session_state.ili_preview_data = preview_data

        if st.session_state.ili_preview_data:
            st.markdown("---")
            st.markdown("### 📋 Step 2: File Preview")
            preview = st.session_state.ili_preview_data
            col1, col2 = st.columns(2)
            with col1:
                st.info(f"**Sheets found:** {len(preview['sheet_names'])}")
            with col2:
                st.info(f"**Total rows:** {sum(preview['row_counts'].values())}")
            for sheet_name in preview["sheet_names"]:
                with st.expander(f"📄 Sheet: **{sheet_name}** ({preview['row_counts'][sheet_name]} rows)"):
                    columns = preview["columns"][sheet_name]
                    st.write(f"**Columns ({len(columns)}):**")
                    st.write(", ".join(columns))

            st.markdown("---")
            st.markdown("### ⚙️ Step 3: Process and Visualize")
            st.caption(
                "Columns are auto-identified. Use **Choose sheet** for a specific tab, or **Auto-detect** "
                "to match Dig Package ILI parsing (vendor sheet + header). Optionally zoom by GWD range or center ±3."
            )

            ili_sheet_mode = st.radio(
                "ILI table source",
                options=["choose_sheet", "vendor_auto"],
                format_func=lambda x: (
                    "Choose sheet from preview"
                    if x == "choose_sheet"
                    else "Auto-detect (vendor format — same as Dig Package)"
                ),
                horizontal=True,
                key="ili_sheet_mode",
            )

            selected_sheet = None
            selected_vendor = DIG_PACKAGE_ILI_FORMAT_OPTIONS[1]
            if ili_sheet_mode == "choose_sheet":
                selected_sheet = st.selectbox(
                    "Select sheet to process",
                    options=preview["sheet_names"],
                    help="Choose which sheet contains your ILI data",
                    key="ili_selected_sheet",
                )
            else:
                selected_vendor = st.selectbox(
                    "ILI vendor / layout",
                    options=list(DIG_PACKAGE_ILI_FORMAT_OPTIONS),
                    index=1,
                    help="Uses the same sheet and header detection as Dig Package Generator (Rosen anomalies vs pipetally, etc.).",
                    key="ili_vendor_format_visual",
                )

            gwd_numbers = (st.session_state.ili_upload_feature_map_data or {}).get("gwd_numbers", [])
            zoom_mode = st.radio(
                "**Zoom to section**",
                options=["Show all", "GWD range (start–end)", "Center GWD ±3"],
                horizontal=True,
                help="Filter by GWD numbers. Process once with 'Show all' to load available GWDs.",
                key="ili_zoom_mode",
            )
            gwd_start, gwd_end, gwd_center = None, None, None
            gwd_opts = [str(gwd) for gwd in gwd_numbers]
            if zoom_mode == "GWD range (start–end)" and gwd_opts:
                c1, c2 = st.columns(2)
                with c1:
                    gwd_start = st.selectbox(
                        "Start GWD",
                        options=[""] + gwd_opts,
                        format_func=lambda x: x or "—",
                        key="ili_gwd_start",
                    )
                    gwd_start = int(gwd_start) if gwd_start and gwd_start.isdigit() else None
                with c2:
                    gwd_end = st.selectbox(
                        "End GWD",
                        options=[""] + gwd_opts,
                        format_func=lambda x: x or "—",
                        key="ili_gwd_end",
                    )
                    gwd_end = int(gwd_end) if gwd_end and gwd_end.isdigit() else None
            elif zoom_mode == "Center GWD ±3" and gwd_opts:
                gwd_center = st.selectbox(
                    "Center GWD",
                    options=[""] + gwd_opts,
                    format_func=lambda x: x or "—",
                    key="ili_gwd_center",
                )
                gwd_center = int(gwd_center) if gwd_center and gwd_center.isdigit() else None

            if st.button("🚀 Process Data", type="primary", key="ili_process_button"):
                with st.spinner("Processing data (columns auto-identified)..."):
                    if ili_sheet_mode == "vendor_auto":
                        result = asyncio.run(
                            call_process_feature_map_api(
                                uploaded_file,
                                vendor_format=selected_vendor,
                                gwd_start=gwd_start,
                                gwd_end=gwd_end,
                                gwd_center=gwd_center,
                            )
                        )
                    else:
                        result = asyncio.run(
                            call_process_feature_map_api(
                                uploaded_file,
                                sheet_name=selected_sheet,
                                gwd_start=gwd_start,
                                gwd_end=gwd_end,
                                gwd_center=gwd_center,
                            )
                        )
                    if result and result.get("success"):
                        st.session_state.ili_upload_feature_map_data = result
                        st.success(f"✅ Parsed {result.get('total_rows', 0)} features!")
                    elif result and not result.get("success"):
                        st.error(result.get("error", "Process failed"))

            if (
                st.session_state.ili_upload_feature_map_data
                and st.session_state.ili_upload_feature_map_data.get("success")
            ):
                st.markdown("---")
                fm = st.session_state.ili_upload_feature_map_data
                total_before = fm.get("total_rows")
                render_feature_map_fragment(
                    fm,
                    total_before_filter=total_before,
                    key_prefix="ili_upload",
                )
    else:
        st.info(
            """
            👆 **Upload an Excel file** with ILI (In-Line Inspection) data

            The tool visualizes pipeline features by chainage and depth.
            No assessment or statistics — visualization only.
            """
        )


def render_ili_visual_tool() -> None:
    _init_ili_session_state()

    input_mode = st.radio(
        "**Input format**",
        options=["Upload Excel File", "Paste from Clipboard"],
        horizontal=True,
        help="Upload a raw ILI Excel file or paste tabular ILI data from the clipboard.",
        key="ili_input_mode",
    )

    if input_mode == "Paste from Clipboard":
        _render_ili_paste_mode()
    else:
        _render_ili_upload_mode()
