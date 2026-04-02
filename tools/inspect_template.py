"""
Template Inspector — run this to dump ALL named ranges, cell addresses, and
sheet layout from the 2026 Dig Package Template.

Usage:
    uv run python tools/inspect_template.py

Output: docs/TEMPLATE_NAMED_RANGES.md
"""

import io
import sys
from pathlib import Path

# Allow running from repo root
_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

# Prefer the template shipped with the backend; fall back to a local Reference path.
_BUNDLED = (
    _REPO_ROOT
    / "backend"
    / "static"
    / "templates"
    / "dig_package"
    / "2026 Dig Package Template.xlsx"
)
_LEGACY = Path(r"C:\Users\cshen\Documents\Reference dig package\3-Dig Package Template\2026 Dig Package Template.xlsx")
TEMPLATE_PATH = _BUNDLED if _BUNDLED.is_file() else _LEGACY
OUTPUT_PATH = _REPO_ROOT / "docs" / "TEMPLATE_NAMED_RANGES.md"


def run():
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not found — run: pip install openpyxl")
        sys.exit(1)

    if not TEMPLATE_PATH.exists():
        print(f"Template not found at: {TEMPLATE_PATH}")
        sys.exit(1)

    wb = load_workbook(str(TEMPLATE_PATH), data_only=False)

    lines = ["# 2026 Dig Package Template — Named Ranges & Sheet Layout", ""]

    # -----------------------------------------------------------------------
    # Named ranges
    # -----------------------------------------------------------------------
    lines.append("## Named Ranges")
    lines.append("")
    lines.append("| Name | Sheet | Cell | Current Value |")
    lines.append("|------|-------|------|---------------|")

    named_ranges_found = []
    for nr_name, nr in wb.defined_names.items():
        destinations = list(nr.destinations)
        if not destinations:
            lines.append(f"| `{nr_name}` | (no destination) | — | — |")
            named_ranges_found.append(nr_name)
            continue
        for sheet_name, cell_ref in destinations:
            try:
                ws = wb[sheet_name]
                cell = ws[cell_ref]
                val = cell.value
                val_str = repr(val) if val is not None else "(empty)"
                lines.append(f"| `{nr_name}` | {sheet_name} | {cell_ref} | {val_str} |")
            except Exception as e:
                lines.append(f"| `{nr_name}` | {sheet_name} | {cell_ref} | ERROR: {e} |")
            named_ranges_found.append(nr_name)

    lines.append("")
    lines.append(f"**Total named ranges: {len(named_ranges_found)}**")
    lines.append("")

    # -----------------------------------------------------------------------
    # Sheet overview
    # -----------------------------------------------------------------------
    lines.append("## Sheets")
    lines.append("")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"### Sheet: `{sheet_name}`")
        lines.append(f"- Dimensions: {ws.dimensions}")
        lines.append(f"- Max row: {ws.max_row}, Max col: {ws.max_column}")
        lines.append("")

        # First 30 non-empty rows
        lines.append("| Row | Col | Value | Formula |")
        lines.append("|-----|-----|-------|---------|")
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    is_formula = str(cell.value).startswith("=") if isinstance(cell.value, str) else False
                    formula_str = str(cell.value) if is_formula else ""
                    val_preview = repr(cell.value)[:80] if not is_formula else "(formula)"
                    lines.append(f"| {cell.row} | {cell.column} ({cell.column_letter}) | {val_preview} | {formula_str} |")
                    count += 1
                    if count >= 200:
                        break
            if count >= 200:
                lines.append("| ... | ... | (first 200 cells shown) | |")
                break
        lines.append("")

    # -----------------------------------------------------------------------
    # Write output
    # -----------------------------------------------------------------------
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Written to: {OUTPUT_PATH}")
    print(f"Named ranges found: {len(named_ranges_found)}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    run()
