"""
MDL Inspector — reads the sample Dig Notification Log and reports:
  1. All sheet names
  2. Every column header found on the detected header row
  3. Which MDL_COLUMN_KEYWORDS map to which actual column names (and which are MISSING)
  4. First 3 data rows for manual review
  5. All unique Dig IDs and whether each passes is_valid_dig_id()

Usage:
    python tools/inspect_mdl.py
    python tools/inspect_mdl.py "path/to/other_mdl.xlsx"
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

DEFAULT_PATH = r"C:\Users\cshen\Documents\Reference dig package\1-Input dig list\ID216_Dig Notification Log_Integrity Dig Program_20260226.xlsx"


def run(mdl_path: str = DEFAULT_PATH):
    try:
        from openpyxl import load_workbook
        import pandas as pd
    except ImportError:
        print("openpyxl / pandas not found. Run: pip install openpyxl pandas")
        sys.exit(1)

    path = Path(mdl_path)
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    from backend.pipeline.dig_package import (
        MDL_COLUMN_KEYWORDS,
        MDL_WORKSHEET_KEYWORDS,
        is_valid_dig_id,
    )
    from backend.pipeline.ili_reader import read_excel_with_detected_header

    print(f"\n{'='*70}")
    print(f"MDL INSPECTOR: {path.name}")
    print(f"{'='*70}")

    # ── 1. Sheet names ────────────────────────────────────────────────────────
    wb = load_workbook(str(path), read_only=True, data_only=True)
    print(f"\n[1] Sheet names: {wb.sheetnames}")
    kw_lower = [k.lower() for k in MDL_WORKSHEET_KEYWORDS]
    for name in wb.sheetnames:
        match = "✅ MATCH" if any(k in name.lower() for k in kw_lower) else "  (no match)"
        print(f"    {name!r}  {match}")
    wb.close()

    # ── 2. Parse with the real reader ─────────────────────────────────────────
    file_bytes = path.read_bytes()
    df, col_map, sheet_name, header_row = read_excel_with_detected_header(
        file_content=file_bytes,
        keyword_map=MDL_COLUMN_KEYWORDS,
        sheet_keywords=MDL_WORKSHEET_KEYWORDS,
        min_matches=3,
    )

    print(f"\n[2] Parsed sheet: {sheet_name!r}  |  header at row {header_row}")
    print(f"    DataFrame shape: {df.shape}  ({df.shape[0]} data rows, {df.shape[1]} columns)")

    # ── 3. Column mapping results ─────────────────────────────────────────────
    print(f"\n[3] Column Mapping  ({len(col_map)} of {len(MDL_COLUMN_KEYWORDS)} keys mapped)")
    print(f"    {'MDL Key':<30} {'Actual Column Name':<40} {'Status'}")
    print(f"    {'-'*30} {'-'*40} {'-'*10}")
    for key in MDL_COLUMN_KEYWORDS:
        actual = col_map.get(key)
        status = "✅" if actual else "❌ MISSING"
        print(f"    {key:<30} {str(actual or ''):<40} {status}")

    # ── 4. All actual columns (for discovering unmapped ones) ─────────────────
    print(f"\n[4] All actual columns in the sheet ({len(df.columns)} total):")
    mapped_actuals = set(col_map.values())
    for col in df.columns:
        tag = "  (mapped)" if col in mapped_actuals else "  ← NOT MAPPED"
        print(f"    {col!r}{tag}")

    # ── 5. First 3 data rows ──────────────────────────────────────────────────
    print(f"\n[5] First 3 data rows (mapped columns only):")
    show_cols = [v for v in col_map.values() if v in df.columns]
    preview = df[show_cols].head(3)
    for i, (_, row) in enumerate(preview.iterrows(), 1):
        print(f"\n  Row {i}:")
        for col in show_cols:
            key = next((k for k, v in col_map.items() if v == col), col)
            print(f"    [{key}] {col!r}: {row[col]!r}")

    # ── 6. Dig ID validation ──────────────────────────────────────────────────
    dig_id_col = col_map.get("dig_id")
    if dig_id_col and dig_id_col in df.columns:
        print(f"\n[6] Dig IDs  (column: {dig_id_col!r})")
        all_ids = df[dig_id_col].dropna().unique()
        for did in sorted(all_ids, key=str):
            valid = is_valid_dig_id(did)
            tag = "✅ valid" if valid else "❌ FAILS is_valid_dig_id()"
            print(f"    {did!r:>20}  {tag}")
    else:
        print("\n[6] ❌ No Dig ID column mapped — check MDL_COLUMN_KEYWORDS['dig_id']")

    print(f"\n{'='*70}\n")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    run(path)
