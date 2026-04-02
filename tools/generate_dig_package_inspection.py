#!/usr/bin/env python3
"""
Generate dig-package template inspection artifacts under dev/dig_package_inspection/generated/
(JSON + Markdown) so tools/agents can read results without copy-paste.

Template resolution (first hit wins):
  1. --template PATH
  2. env DIG_PACKAGE_TEMPLATE_PATH
  3. backend/static/templates/dig_package/2026 Dig Package Template.xlsx
  4. Legacy path (Windows), same as tools/inspect_template.py

Usage (repo root):
  python tools/generate_dig_package_inspection.py
  python tools/generate_dig_package_inspection.py --template "C:\\path\\to\\template.xlsx"

Safe to delete: entire folder dev/dig_package_inspection/generated/
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

OUT_DIR = REPO / "dev" / "dig_package_inspection" / "generated"

# Match dig_package_layout scanner
SCAN_MAX_ROW = 300
SCAN_MAX_COL = 40

_LEGACY_TEMPLATE = Path(
    r"C:\Users\cshen\Documents\Reference dig package\3-Dig Package Template\2026 Dig Package Template.xlsx"
)
_BUNDLED = (
    REPO
    / "backend"
    / "static"
    / "templates"
    / "dig_package"
    / "2026 Dig Package Template.xlsx"
)


def _resolve_template(cli_path: Optional[Path]) -> tuple[Optional[Path], List[str]]:
    tried: List[str] = []
    candidates: List[Optional[Path]] = [
        cli_path,
        Path(os.environ.get("DIG_PACKAGE_TEMPLATE_PATH", "").strip()) if os.environ.get("DIG_PACKAGE_TEMPLATE_PATH") else None,
        _BUNDLED,
        _LEGACY_TEMPLATE,
    ]
    for c in candidates:
        if c is None:
            continue
        p = Path(c)
        tried.append(str(p.resolve()))
        if p.is_file():
            return p, tried
    return None, tried


def _dump_sheet_cells(ws) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in range(1, SCAN_MAX_ROW + 1):
        for col in range(1, SCAN_MAX_COL + 1):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            s = str(cell.value).strip()
            if not s:
                continue
            flat = " ".join(s.split())
            out.append(
                {
                    "row": row,
                    "col": col,
                    "col_letter": cell.column_letter,
                    "value": flat,
                }
            )
    return out


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def _write_md_summary(
    path: Path,
    *,
    status: Dict[str, Any],
    verify: Optional[Dict[str, Any]],
    cell_count: int,
) -> None:
    lines = [
        "# Dig package template inspection (generated)",
        "",
        f"- **Generated (UTC):** {status.get('generated_at_utc', '')}",
        f"- **Template:** `{status.get('template_path', '')}`",
        f"- **Template found:** {status.get('template_found', False)}",
        f"- **Non-empty cells (Dig Package):** {cell_count}",
        "",
    ]
    if verify:
        lines.extend(
            [
                "## Layout verify (`dig_package_layout.json`)",
                "",
                f"- **OK:** {verify.get('ok_count', 0)}  **Failed:** {verify.get('fail_count', 0)}  **Total:** {verify.get('total', 0)}",
                "",
                "| Status | Field | Error |",
                "|--------|-------|-------|",
            ]
        )
        for r in verify.get("results", []):
            st = "OK" if r.get("ok") else "FAIL"
            err = (r.get("error") or "").replace("|", "\\|")
            lines.append(f"| {st} | `{r.get('field_id', '')}` | {err} |")
        lines.append("")
    else:
        lines.extend(["## Layout verify", "", "_Skipped (no workbook)._", ""])
    lines.append("---")
    lines.append("")
    lines.append("Artifacts: `inspection_status.json`, `dig_package_sheet_cells.json`, `verify_layout.json`, `INSPECTION_SUMMARY.md`")
    lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=None, help="Path to .xlsx")
    args = parser.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl required", file=sys.stderr)
        return 1

    from backend.pipeline.dig_package_layout import (
        default_layout_manifest_path,
        load_layout_manifest,
        verify_layout_against_workbook,
    )

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    template_path, tried = _resolve_template(args.template)
    manifest_path = default_layout_manifest_path()

    status: Dict[str, Any] = {
        "generated_at_utc": generated_at,
        "template_path": str(template_path) if template_path else None,
        "template_found": template_path is not None,
        "paths_tried": tried,
        "manifest_path": str(manifest_path),
        "output_dir": str(OUT_DIR.resolve()),
    }

    verify_payload: Optional[Dict[str, Any]] = None
    cells: List[Dict[str, Any]] = []
    sheet_used = "Dig Package"

    if template_path is None:
        status["error"] = "template_not_found"
        _write_json(OUT_DIR / "inspection_status.json", status)
        _write_json(OUT_DIR / "dig_package_sheet_cells.json", {"sheet": sheet_used, "cells": []})
        _write_json(
            OUT_DIR / "verify_layout.json",
            {
                "generated_at_utc": generated_at,
                "skipped": True,
                "reason": "no_template",
                "results": [],
            },
        )
        _write_md_summary(
            OUT_DIR / "INSPECTION_SUMMARY.md",
            status=status,
            verify=None,
            cell_count=0,
        )
        print(f"Wrote (no template): {OUT_DIR}")
        return 0

    # Match tools/verify_dig_package_layout.py (data_only=False) so anchors resolve the same way.
    wb = load_workbook(str(template_path), data_only=False)
    manifest = load_layout_manifest()

    if sheet_used not in wb.sheetnames:
        status["error"] = f"sheet_missing:{sheet_used}"
        status["sheetnames"] = wb.sheetnames
        _write_json(OUT_DIR / "inspection_status.json", status)
        _write_json(OUT_DIR / "dig_package_sheet_cells.json", {"sheet": sheet_used, "cells": []})
        _write_json(
            OUT_DIR / "verify_layout.json",
            {
                "generated_at_utc": generated_at,
                "skipped": True,
                "reason": "missing_dig_package_sheet",
                "workbook_sheets": wb.sheetnames,
                "results": [],
            },
        )
        _write_md_summary(OUT_DIR / "INSPECTION_SUMMARY.md", status=status, verify=None, cell_count=0)
        print(f"Wrote (missing sheet): {OUT_DIR}")
        return 0

    ws = wb[sheet_used]
    cells = _dump_sheet_cells(ws)

    results = verify_layout_against_workbook(wb, manifest)
    ok_count = sum(1 for r in results if r["ok"])
    fail_count = len(results) - ok_count

    verify_payload = {
        "generated_at_utc": generated_at,
        "template_path": str(template_path.resolve()),
        "manifest_path": str(manifest_path.resolve()),
        "total": len(results),
        "ok_count": ok_count,
        "fail_count": fail_count,
        "results": results,
    }

    status["sheet_used"] = sheet_used
    status["non_empty_cell_count"] = len(cells)
    _write_json(OUT_DIR / "inspection_status.json", status)
    _write_json(
        OUT_DIR / "dig_package_sheet_cells.json",
        {
            "generated_at_utc": generated_at,
            "template_path": str(template_path.resolve()),
            "sheet": sheet_used,
            "scan": {"max_row": SCAN_MAX_ROW, "max_col": SCAN_MAX_COL},
            "cell_count": len(cells),
            "cells": cells,
        },
    )
    _write_json(OUT_DIR / "verify_layout.json", verify_payload)
    _write_md_summary(
        OUT_DIR / "INSPECTION_SUMMARY.md",
        status=status,
        verify=verify_payload,
        cell_count=len(cells),
    )

    print(f"Wrote: {OUT_DIR}")
    print(f"  cells: {len(cells)}, verify OK: {ok_count}, FAIL: {fail_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
