#!/usr/bin/env python3
"""
Dev tool: verify every field in dig_package_layout.json resolves on a template .xlsx.

Usage (from repo root):
  python tools/verify_dig_package_layout.py path/to/template.xlsx
  python tools/verify_dig_package_layout.py   # uses bundled default template if present

Exit code 0 = all anchors found; 1 = one or more failures.
"""

from __future__ import annotations

import argparse
import io
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from openpyxl import load_workbook

from backend.pipeline.dig_package_layout import (
    default_layout_manifest_path,
    load_layout_manifest,
    verify_layout_against_workbook,
)


def main() -> int:
    p = argparse.ArgumentParser(description="Verify dig package template vs layout manifest")
    p.add_argument(
        "template",
        nargs="?",
        default=None,
        help="Path to .xlsx (default: bundled 2026 template if it exists)",
    )
    p.add_argument(
        "--manifest",
        type=Path,
        default=None,
        help="Override layout JSON (default: backend/static/templates/dig_package/dig_package_layout.json)",
    )
    args = p.parse_args()

    manifest = load_layout_manifest(args.manifest)

    if args.template:
        path = Path(args.template)
        if not path.is_file():
            print(f"ERROR: file not found: {path}", file=sys.stderr)
            return 1
        wb = load_workbook(path, data_only=False)
    else:
        bundled = (
            REPO
            / "backend"
            / "static"
            / "templates"
            / "dig_package"
            / "2026 Dig Package Template.xlsx"
        )
        if not bundled.is_file():
            print(
                f"ERROR: No template path given and bundled file missing:\n  {bundled}\n"
                f"Pass a .xlsx path. Default manifest: {default_layout_manifest_path()}",
                file=sys.stderr,
            )
            return 1
        wb = load_workbook(bundled, data_only=False)

    results = verify_layout_against_workbook(wb, manifest)
    ok_count = sum(1 for r in results if r["ok"])
    fail_count = len(results) - ok_count

    print(f"Layout manifest: {args.manifest or default_layout_manifest_path()}")
    print(f"Results: {ok_count} OK, {fail_count} failed (total checks {len(results)})\n")

    for r in results:
        status = "OK " if r["ok"] else "FAIL"
        extra = "" if r["ok"] else f" — {r['error']}"
        print(f"  [{status}] {r['field_id']}{extra}")

    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
